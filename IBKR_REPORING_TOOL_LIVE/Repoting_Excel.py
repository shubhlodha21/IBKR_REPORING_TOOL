"""
Fetch IBKR Flex Query trades and write an Excel file with sheets:
  - Dashboard         (account info + 7-day & prev-day P&L summary)
  - Pending Order     (live from TWS)
  - Running Positions (net open positions across all trades)
  - PreviousDay
  - Past7Days
  - Past30Days

Usage:
    pip install requests openpyxl ibapi
    python Repoting_Excel.py

Notes:
  - The Flex Query must cover at least 30 days.
  - TWS or IB Gateway must be running for the Pending Order sheet.
  - Set IBKR_TOKEN and IBKR_QUERY_ID as environment variables (via .env).
"""

import os
import sys
import json
import time
import threading
import datetime as dt
import xml.etree.ElementTree as ET
from collections import defaultdict, deque

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.execution import ExecutionFilter

# ----------------------------------------------------------------------
# CONFIG  (set IBKR_TOKEN and IBKR_QUERY_ID as environment variables)
# ----------------------------------------------------------------------
TOKEN    = os.environ.get("IBKR_TOKEN", "")
QUERY_ID = os.environ.get("IBKR_QUERY_ID", "")

if not TOKEN or not QUERY_ID:
    sys.exit("Error: IBKR_TOKEN and IBKR_QUERY_ID must be set as environment variables.\n"
             "Copy .env.example to .env and fill in your credentials, then run via run.ps1.")

_REPORTS_DIR = "reports"
os.makedirs(_REPORTS_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(_REPORTS_DIR, dt.date.today().strftime("MIS_%d%b%Y") + ".xlsx")

# Fixed "old account" reference workbook. These sheets belong to a closed
# account, never change, and are copied VERBATIM (values + formatting) into
# every report. Each entry is (report_title, [candidate source tab names]) —
# the first candidate found in OLD_REFERENCE.xlsx is copied. Multiple candidates
# are listed so a tab rename in the source (e.g. "Open Position old" vs
# "Open Position old AC") doesn't silently drop the sheet. Matching is done on
# stripped/lower-cased names, so trailing spaces and case don't matter.
OLD_REFERENCE_FILE   = os.path.join(_REPORTS_DIR, "OLD_REFERENCE.xlsx")
OLD_REFERENCE_SHEETS = [
    ("Dashboard old AC",     ["Dashboard old AC"]),
    ("Open Position old AC", ["Open Position old AC", "Open Position old"]),
    ("All Trades old AC",    ["All Trades old AC",    "All Trades old"]),
    ("Trade Summary old AC", ["Trade Summary old AC", "Trade Summary old"]),
]

# Persistent order ledger — accumulates trigger/limit from live open orders on
# every run, so trades can be matched to the prices set when the order was placed.
# (IBKR does NOT retain this on executed trades, so we must capture it ourselves.)
_BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
LEDGER_FILE = os.path.join(_BASE_DIR, "orders_ledger.json")

# Persistent trades ledger — accumulates EVERY live TWS execution (fill) across
# runs, keyed by IBKR's unique execId. IBKR's reqExecutions only returns the
# current day and only while TWS is connected, and the Flex batch lags ~a day, so
# without this a missed/late run would lose that day's orders permanently. With
# it, every day's fills are captured the moment they're seen and always shown.
TRADES_LEDGER_FILE = os.path.join(_BASE_DIR, "trades_ledger.json")

BASE     = "https://ndcdyn.interactivebrokers.com/AccountManagement/FlexWebService"
SEND_URL = f"{BASE}/SendRequest"
GET_URL  = f"{BASE}/GetStatement"
VERSION  = "3"

MAX_POLL_ATTEMPTS = 12
POLL_WAIT_SECONDS = 5

# TWS live connection. Host/port/client id come from the environment (.env) so
# you can point the report at paper TWS for testing without editing code — e.g.
# set IBKR_TWS_PORT=7497 for paper. Defaults are the live-TWS values.
HOST       = os.environ.get("IBKR_TWS_HOST", "127.0.0.1")
PORT       = int(os.environ.get("IBKR_TWS_PORT", "7496"))  # 7496 live TWS | 7497 paper | 4001/4002 Gateway
CLIENT_ID  = int(os.environ.get("IBKR_TWS_CLIENT_ID", "998"))
ORDER_WAIT = 8           # seconds to wait for open-order callbacks
ACCT_WAIT  = 8           # seconds to wait for account-summary callbacks
POS_WAIT   = 8           # seconds to wait for live-position callbacks
EXEC_WAIT  = 15          # seconds to wait for executions (0.5 s timer + fills + commissions)

_UNSET = (0, 1.7976931348623157e308)

# IBKR sends Double.MAX_VALUE as the "unset" sentinel for portfolio figures that
# aren't available (e.g. realizedPNL before any close). Treat those as missing.
_SENTINEL = 1.7976931348623157e308


def _ibkr_has(v):
    """True when an IBKR numeric value is real (not None and not the MAX_VALUE
    sentinel IBKR uses to mean 'no value')."""
    try:
        return v is not None and abs(float(v)) < _SENTINEL
    except (ValueError, TypeError):
        return False

# No date cutoff — the report includes every trade IBKR returns.

# Risk limits shown on the Dashboard (hardcoded).
MAX_EXPOSURE   = 150000
DAILY_MAX_LOSS = 2000

# Cumulative loss attributed to bugs — a manually-maintained figure shown on the
# Dashboard. Update this value here when it changes; it stays fixed otherwise.
LOSS_DUE_TO_BUGS = 0

# IST offset for report timestamps
_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))

# Account summary tags to request from TWS
ACCOUNT_SUMMARY_TAGS = (
    "NetLiquidation,TotalCashValue,SettledCash,AccruedCash,"
    "BuyingPower,AvailableFunds,ExcessLiquidity,"
    "InitMarginReq,MaintMarginReq,"
    "GrossPositionValue,UnrealizedPnL,RealizedPnL"
)

# Display labels for each account tag (label, tag_key)
ACCOUNT_DISPLAY = [
    ("Net Liquidation (Total Funds)", "NetLiquidation"),
    ("Total Cash Value",              "TotalCashValue"),
    ("Settled Cash",                  "SettledCash"),
    ("Accrued Cash",                  "AccruedCash"),
    ("Buying Power",                  "BuyingPower"),
    ("Available Funds",               "AvailableFunds"),
    ("Excess Liquidity",              "ExcessLiquidity"),
    ("Initial Margin Req.",           "InitMarginReq"),
    ("Maintenance Margin Req.",       "MaintMarginReq"),
    ("Gross Position Value",          "GrossPositionValue"),
    ("Unrealized PnL",                "UnrealizedPnL"),
    ("Realized PnL",                  "RealizedPnL"),
]


# ----------------------------------------------------------------------
# TWS: fetch live pending orders + account summary in one connection
# ----------------------------------------------------------------------
class IBOrderApp(EWrapper, EClient):
    def __init__(self):
        EClient.__init__(self, self)
        self.orders_by_symbol = defaultdict(list)
        self.account_data     = {}          # tag -> (value, currency)
        self.executions       = {}          # execId -> execution dict
        self.completed_orders = []          # today's filled/cancelled orders
        self.positions        = []          # live open positions (reqPositions)
        self.portfolio        = {}          # conId -> live portfolio dict (updatePortfolio)
        self.account_code     = ""          # managed account for reqAccountUpdates
        self._orders_done     = threading.Event()
        self._acct_done       = threading.Event()
        self._exec_done       = threading.Event()
        self._completed_done  = threading.Event()
        self._positions_done  = threading.Event()
        self._portfolio_done  = threading.Event()
        self.connected_ok     = False

    def error(self, reqId, errorCode, errorString, advancedOrderRejectJson=""):
        if errorCode in (502, 504, 1100, 1300):
            print(f"[TWS] {errorCode}: {errorString}", file=sys.stderr)
        elif errorCode not in (2104, 2106, 2107, 2119, 2158, 2100):
            print(f"[TWS] id={reqId} code={errorCode}: {errorString}", file=sys.stderr)

    def connectAck(self):
        self.connected_ok = True

    def nextValidId(self, orderId):
        self.connected_ok = True
        self.reqAllOpenOrders()
        self.reqAccountSummary(9001, "All", ACCOUNT_SUMMARY_TAGS)
        # Live open positions across the account — the real broker holdings,
        # so the Open Position sheet reflects what IBKR actually shows open
        # rather than netting the (possibly incomplete) Flex trade history.
        self.reqPositions()
        # Delay executions request by 0.5 s so the first two requests
        # don't congest the message loop, and set a 24-hour window so
        # pre-market fills and all intraday trades are captured.
        threading.Timer(0.5, self._req_executions).start()

    # ── live open positions ───────────────────────────────────────────
    def position(self, account, contract, position, avgCost):
        self.positions.append({
            "account":  account,
            "conId":    contract.conId,
            "symbol":   contract.symbol,
            "secType":  contract.secType,
            "currency": contract.currency,
            "exchange": contract.exchange or contract.primaryExchange or "",
            "name":     contract.localSymbol or contract.symbol,
            "position": position,
            "avgCost":  avgCost,
        })

    def positionEnd(self):
        print(f"[TWS] positionEnd - {len(self.positions)} live position(s).")
        self.cancelPositions()
        self._positions_done.set()

    # ── live portfolio (per-position market value / unrealized & realized PnL) ──
    def managedAccounts(self, accountsList):
        # Fires on connect. reqAccountUpdates needs a concrete account code, so we
        # subscribe here (the portfolio download then streams updatePortfolio).
        self.account_code = (accountsList or "").split(",")[0].strip()
        if self.account_code:
            self.reqAccountUpdates(True, self.account_code)
        else:
            self._portfolio_done.set()

    def updatePortfolio(self, contract, position, marketPrice, marketValue,
                        averageCost, unrealizedPNL, realizedPNL, accountName):
        self.portfolio[contract.conId] = {
            "position":      position,
            "marketPrice":   marketPrice,
            "marketValue":   marketValue,
            "averageCost":   averageCost,
            "unrealizedPNL": unrealizedPNL,
            "realizedPNL":   realizedPNL,
        }

    def accountDownloadEnd(self, accountName):
        print(f"[TWS] accountDownloadEnd - {len(self.portfolio)} portfolio item(s).")
        if self.account_code:
            self.reqAccountUpdates(False, self.account_code)   # unsubscribe
        self._portfolio_done.set()

    def _req_executions(self):
        # Empty filter → ALL of the current trading day's executions.
        # NOTE: IBKR's API only ever returns the *current day's* executions;
        # a time filter cannot reach prior days, and a malformed one can
        # suppress results entirely — so we pass an unfiltered request.
        self.reqExecutions(10001, ExecutionFilter())
        # Today's completed (filled/cancelled) orders — these carry the order's
        # limit/trigger even after they've filled, which open orders no longer do.
        self.reqCompletedOrders(False)   # False = all, not just API-placed

    # ── completed orders (today's filled/cancelled, with limit/trigger) ──
    @staticmethod
    def _pair_symbol(contract):
        """Display/keying symbol for a contract. IBKR reports an FX (CASH)
        contract.symbol as the bare BASE currency (e.g. 'USD' for USD.JPY); its
        localSymbol carries the full pair ('USD.JPY'), which is how the Flex
        trade history and the rest of the report key it. Use the pair for FX so
        pending/completed FX orders show 'USD.JPY', not just 'USD'."""
        if getattr(contract, "secType", "") == "CASH":
            if contract.localSymbol:
                return contract.localSymbol
            if contract.symbol and contract.currency:
                return f"{contract.symbol}.{contract.currency}"
        return contract.symbol

    def completedOrder(self, contract, order, orderState):
        trigger = order.auxPrice if order.auxPrice not in _UNSET else None
        limit   = order.lmtPrice if order.lmtPrice not in _UNSET else None
        oid     = getattr(order, "permId", None) or getattr(order, "orderId", None)
        self.completed_orders.append({
            "orderId":   f"C{oid}",
            "symbol":    self._pair_symbol(contract),
            "action":    order.action,
            "orderType": order.orderType,
            "trigger":   trigger,
            "limit":     limit,
            "quantity":  int(order.totalQuantity) if order.totalQuantity else 0,
            "status":    orderState.status,
            "currency":  contract.currency,
        })

    def completedOrdersEnd(self):
        print(f"[TWS] completedOrdersEnd - {len(self.completed_orders)} completed order(s).")
        self._completed_done.set()

    # ── pending orders ────────────────────────────────────────────────
    def openOrder(self, orderId, contract, order, orderState):
        trigger = order.auxPrice if order.auxPrice not in _UNSET else None
        limit   = order.lmtPrice if order.lmtPrice not in _UNSET else None
        self.orders_by_symbol[self._pair_symbol(contract)].append({
            "orderId":   orderId,
            "action":    order.action,
            "orderType": order.orderType,
            "trigger":   trigger,
            "limit":     limit,
            "quantity":  int(order.totalQuantity),
            "parentId":  order.parentId,
            # Quote currency (JPY for USD.JPY, USD for a US stock) so the USD
            # notional can be derived from the native trigger/limit price.
            "currency":  contract.currency,
            # Order lifecycle status (Submitted / PreSubmitted / Filled /
            # Cancelled / Inactive / …). reqAllOpenOrders can still momentarily
            # return an order that has just filled or been cancelled; the status
            # lets build_pending_rows drop it so a filled position is never shown
            # as still pending.
            "status":    getattr(orderState, "status", "") or "",
        })

    def openOrderEnd(self):
        self._orders_done.set()

    # ── account summary ───────────────────────────────────────────────
    def accountSummary(self, reqId, account, tag, value, currency):
        self.account_data[tag] = (value, currency)

    def accountSummaryEnd(self, reqId):
        self._acct_done.set()

    # ── today's executions ────────────────────────────────────────────
    def execDetails(self, reqId, contract, execution):
        self.executions[execution.execId] = {
            "account":     execution.acctNumber,
            "contract":    contract.symbol,
            # localSymbol is the tradeable instrument as Flex reports it — for FX
            # it's the pair ("EUR.USD") vs. the bare base ("EUR") in contract.symbol,
            # so today's live FX fills line up with the Flex trade history.
            "localSymbol": contract.localSymbol or contract.symbol,
            "secType":     contract.secType,
            "conId":       contract.conId,
            "action":      "BUY" if execution.side in ("BOT", "BUY") else "SELL",
            "quantity":    execution.shares,
            "price":       execution.price,
            "currency":    contract.currency,
            "exchange":    execution.exchange,
            "time":        execution.time,
            "commission":  None,            # filled in by commissionReport
            "commission_currency": None,    # filled in by commissionReport
        }

    def commissionReport(self, commissionReport):
        ex = self.executions.get(commissionReport.execId)
        if ex is not None:
            ex["commission"] = commissionReport.commission
            # IBKR bills forex commissions in USD even when the trade currency is
            # JPY/etc.; capturing the currency lets us avoid wrongly FX-converting
            # an already-USD commission.
            ex["commission_currency"] = getattr(commissionReport, "currency", None)

    def execDetailsEnd(self, reqId):
        print(f"[TWS] execDetailsEnd - {len(self.executions)} execution(s) received so far.")
        self._exec_done.set()


def fetch_tws_data():
    """Connect to TWS once and fetch open orders, account summary, today's
    executions, today's completed orders, and live open positions.

    The positions element is None when TWS could not be reached (so the Open
    Position sheet can fall back to netting the Flex trade history); when
    connected it is a list, which may legitimately be empty if the account is
    currently flat."""
    app = IBOrderApp()
    try:
        app.connect(HOST, PORT, CLIENT_ID)
    except Exception as e:
        print(f"[TWS] Could not connect ({e}) - live sheets will be empty.",
              file=sys.stderr)
        return {}, {}, {}, [], None
    threading.Thread(target=app.run, daemon=True).start()
    # Each feed has its own generous timeout so a slow/missing feed
    # doesn't block the others.
    app._orders_done.wait(timeout=ORDER_WAIT)
    app._acct_done.wait(timeout=ACCT_WAIT)
    app._positions_done.wait(timeout=POS_WAIT)
    app._portfolio_done.wait(timeout=POS_WAIT)
    app._exec_done.wait(timeout=EXEC_WAIT)
    app._completed_done.wait(timeout=EXEC_WAIT)
    # Commission reports arrive asynchronously after execDetailsEnd;
    # give TWS extra time to push them all through before we disconnect.
    time.sleep(1.5)
    app.disconnect()
    # Merge the live portfolio figures (market value, avg cost, unrealized &
    # realized PnL) onto each position by conId, so the Open Position sheet can
    # source those columns straight from IBKR.
    for p in app.positions:
        pf = app.portfolio.get(p.get("conId"))
        if pf:
            p.update(pf)
    # connected_ok distinguishes "connected but flat" (empty list) from
    # "never connected" (None) so the Open Position sheet knows whether to
    # trust the empty result or fall back to the trade-netted figures.
    positions = app.positions if app.connected_ok else None
    return (app.orders_by_symbol, app.account_data, app.executions,
            app.completed_orders, positions)


# ----------------------------------------------------------------------
# Persistent order ledger: capture trigger/limit from live open orders so
# trades can later be matched to the prices set when the order was placed.
# ----------------------------------------------------------------------
def _norm_type(t):
    """Normalise order type for matching (TWS 'STP LMT' vs Flex 'STPLMT')."""
    return (t or "").upper().replace(" ", "")


def load_order_ledger():
    if os.path.exists(LEDGER_FILE):
        try:
            with open(LEDGER_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            print("[ledger] existing ledger unreadable - starting fresh.", file=sys.stderr)
    return {}


def update_order_ledger(ledger, orders_by_symbol, completed_orders=None):
    """Merge live open orders AND today's completed orders into the ledger.

    Open orders cover still-working orders; completed orders cover ones that
    already filled today (which an EOD run would otherwise miss). Both carry
    the limit/trigger prices set when the order was placed.
    """
    today = dt.date.today().isoformat()

    def _merge(oid, symbol, action, order_type, limit, trigger, quantity):
        rec = ledger.get(oid, {})
        rec.update({
            "symbol":    symbol,
            "action":    (action or "").upper(),
            "orderType": _norm_type(order_type),
            "limit":     limit,
            "trigger":   trigger,
            "quantity":  quantity,
            "last_seen": today,
        })
        rec.setdefault("first_seen", today)
        ledger[oid] = rec

    for symbol, legs in orders_by_symbol.items():
        for o in legs:
            _merge(str(o["orderId"]), symbol, o["action"], o["orderType"],
                   o["limit"], o["trigger"], o["quantity"])

    for o in (completed_orders or []):
        _merge(str(o["orderId"]), o["symbol"], o["action"], o["orderType"],
               o["limit"], o["trigger"], o["quantity"])

    return ledger


def save_order_ledger(ledger):
    try:
        with open(LEDGER_FILE, "w", encoding="utf-8") as f:
            json.dump(ledger, f, indent=2)
    except OSError as e:
        print(f"[ledger] could not save: {e}", file=sys.stderr)


# ----------------------------------------------------------------------
# Persistent trades ledger: accumulate every live TWS execution across runs so
# no day's orders are ever lost (see TRADES_LEDGER_FILE note above).
# ----------------------------------------------------------------------
def load_trades_ledger():
    if os.path.exists(TRADES_LEDGER_FILE):
        try:
            with open(TRADES_LEDGER_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            print("[trades ledger] existing ledger unreadable - starting fresh.",
                  file=sys.stderr)
    return {}


def update_trades_ledger(ledger, executions):
    """Merge this session's live TWS fills into the persistent trades ledger,
    keyed by IBKR's unique execId. Existing entries are refreshed (e.g. a
    commission that arrived after the first capture) rather than duplicated."""
    today = dt.date.today().isoformat()
    for exec_id, ex in (executions or {}).items():
        rec = ledger.get(exec_id, {})
        rec.update(ex)                      # latest values win (e.g. commission)
        rec.setdefault("captured_date", today)
        ledger[exec_id] = rec
    return ledger


def save_trades_ledger(ledger):
    try:
        with open(TRADES_LEDGER_FILE, "w", encoding="utf-8") as f:
            json.dump(ledger, f, indent=2)
    except OSError as e:
        print(f"[trades ledger] could not save: {e}", file=sys.stderr)


def build_order_lookup(ledger):
    """(symbol, action, normType) -> list of (last_seen_date, limit, trigger)."""
    lookup = defaultdict(list)
    for rec in ledger.values():
        key = (rec.get("symbol"), (rec.get("action") or "").upper(),
               _norm_type(rec.get("orderType")))
        try:
            seen = dt.date.fromisoformat(rec.get("last_seen", ""))
        except (ValueError, TypeError):
            seen = dt.date.min
        lookup[key].append((seen, rec.get("limit"), rec.get("trigger")))
    return lookup


def lookup_order_prices(order_lookup, symbol, action, order_type, trade_date):
    """Return (limit, trigger) for the ledger order best matching a trade."""
    if not order_lookup:
        return None, None
    key = (symbol, (action or "").upper(), _norm_type(order_type))
    cands = order_lookup.get(key, [])
    if not cands:
        return None, None
    td = trade_date or dt.date.min
    prior = [c for c in cands if c[0] <= td]           # orders seen on/before the trade
    pick = (max(prior, key=lambda c: c[0]) if prior
            else min(cands, key=lambda c: abs((c[0] - td).days)))
    return pick[1], pick[2]


# ----------------------------------------------------------------------
# Pending order helpers
# ----------------------------------------------------------------------
def _fmt_price(v):
    return f"{v:.2f}" if v is not None else ""


def _sl_pct(entry_trigger, sl_trigger):
    if entry_trigger and sl_trigger:
        return f"{abs(entry_trigger - sl_trigger) / entry_trigger * 100:.2f}%"
    return ""


def _offset(trigger, limit_price):
    if trigger is not None and limit_price is not None:
        return f"{abs(limit_price - trigger):.2f}"
    return ""


def _usd_rate(currency, fx_rates):
    """fxRateToBase for `currency` (1.0 for USD or when no rate is known)."""
    ccy = (currency or BASE_CURRENCY).upper()
    return 1.0 if ccy == BASE_CURRENCY else (fx_rates or {}).get(ccy, 1.0)


def _total_amount(qty, trigger, limit_price, currency=None, fx_rates=None):
    """USD notional for an order/trade row. The price (trigger, else limit) is the
    instrument's native quote — for an FX pair it's in the quote currency (JPY for
    USD.JPY) — so multiply by that currency's fxRateToBase to express every Total
    Amount in USD (a USD.JPY leg then reads ~50,000 USD, not millions of JPY)."""
    price = trigger if trigger is not None else limit_price
    if price is None or qty in (None, ""):
        return ""
    try:
        return f"{float(qty) * float(price) * _usd_rate(currency, fx_rates):,.2f}"
    except (ValueError, TypeError):
        return ""


# Order statuses that mean the order is no longer working, so it must NOT be
# shown as pending even if reqAllOpenOrders still momentarily reports it (e.g. an
# entry that has just filled). Any other status (Submitted / PreSubmitted /
# PendingSubmit / ApiPending / …) is treated as still live.
_DEAD_ORDER_STATUSES = {"filled", "cancelled", "apicancelled", "inactive"}


def _is_working_order(o):
    """True when an open-order leg is still live (not filled/cancelled/inactive)."""
    return (o.get("status") or "").strip().lower() not in _DEAD_ORDER_STATUSES


def build_pending_rows(orders_by_symbol, fx_rates=None):
    rows = []
    sr = 0
    for symbol, legs in orders_by_symbol.items():
        # Drop legs that already filled / were cancelled so a position that is
        # now executed doesn't linger here as "pending"; skip the symbol entirely
        # once none of its legs are still working.
        legs = [o for o in legs if _is_working_order(o)]
        if not legs:
            continue
        sr += 1
        parents  = [o for o in legs if not o["parentId"]]
        children = [o for o in legs if o["parentId"]]
        entry = parents[0] if parents else legs[0]
        stop  = children[0] if children else (legs[1] if len(legs) > 1 else None)

        # Entry row — show Total Amount.  Name is blank: pending orders come
        # live from TWS, which doesn't carry the Flex 'description' full name.
        rows.append([
            sr, symbol, "", entry["action"], entry["orderType"],
            _fmt_price(entry["trigger"]), _fmt_price(entry["limit"]),
            _offset(entry["trigger"], entry["limit"]),
            _sl_pct(entry["trigger"], stop["trigger"] if stop else None),
            entry["quantity"],
            _total_amount(entry["quantity"], entry["trigger"], entry["limit"],
                          entry.get("currency"), fx_rates),
        ])
        # Stop/SELL row — no Total Amount. Repeat the Contract (ticker) so the
        # symbol is visible on the SELL leg too; Sr No stays blank.
        if stop:
            rows.append([
                "", symbol, "", stop["action"], stop["orderType"],
                _fmt_price(stop["trigger"]), _fmt_price(stop["limit"]),
                _offset(stop["trigger"], stop["limit"]),
                "", stop["quantity"],
                "",
            ])
    return rows


PENDING_HEADERS = ["Sr No", "Contract", "Name", "Action", "Type",
                   "Trigger", "Limit", "Offset", "SL-Percentage", "Quantity", "Total Amount"]
PENDING_WIDTHS  = [7, 12, 30, 9, 10, 10, 10, 10, 15, 10, 14]

# Trades Status sheet — Pending Order layout plus an Order_Status column,
# combining live pending orders, today's completed orders, and every filled trade.
TRADES_STATUS_HEADERS = PENDING_HEADERS + ["Order_Status"]
TRADES_STATUS_WIDTHS  = PENDING_WIDTHS + [20]
STATUS_PENDING   = "pending"            # live open orders still working
STATUS_CANCELLED = "cancelled"          # completed order the user cancelled
STATUS_REJECTED  = "did not go through" # completed order that never executed (rejected/inactive)
STATUS_FILLED    = "executed"           # filled trade (from All Trades)


def _num_cell(v):
    """Parse a formatted numeric cell ('1,234.50') back to float, or None."""
    try:
        return float(str(v).replace(",", "")) if v not in (None, "") else None
    except (ValueError, TypeError):
        return None


def build_trades_status_rows(pending_rows, completed_orders, all_trade_rows,
                             fx_rates=None):
    """Rows for the Trades_Status sheet (Pending Order layout + Order_Status), from:
         • live pending orders                         → 'pending'
         • today's cancelled completed orders          → 'cancelled'
         • today's rejected/inactive completed orders  → 'did not go through'
         • every filled trade in All Trades            → 'executed'
    Sr No is a plain running count (1,2,3,…) on every row. Every Total Amount is
    in USD (see _total_amount) so the sheet never mixes currencies."""
    rows = []
    sr   = 0

    # 1) Pending orders — reuse the exact Pending Order rows (already in USD),
    #    but stamp a sequential Sr No onto every row, including SELL legs.
    for r in pending_rows:
        sr += 1
        row = list(r)
        row[0] = sr
        rows.append(row + [STATUS_PENDING])

    # 2) Today's completed orders that did NOT fill: split into user-cancelled
    #    ('cancelled') vs rejected/inactive ('did not go through'). Filled ones
    #    are skipped here — they come from All Trades as 'executed' below.
    for o in (completed_orders or []):
        st = (o.get("status") or "").lower()
        if "fill" in st:
            continue
        label = STATUS_CANCELLED if "cancel" in st else STATUS_REJECTED
        sr += 1
        rows.append([
            sr, o["symbol"], "", o["action"], o["orderType"],
            _fmt_price(o.get("trigger")), _fmt_price(o.get("limit")),
            _offset(o.get("trigger"), o.get("limit")),
            "", o.get("quantity", ""),
            _total_amount(o.get("quantity") or 0, o.get("trigger"), o.get("limit"),
                          o.get("currency"), fx_rates),
            label,
        ])

    # 3) Filled trades — map each All Trades row into the Pending Order layout.
    #    All Trades columns: 3=Contract 4=Name 5=Action 6=Qty 7=Price
    #    8=Type 9=Trigger 10=Offset 11=Limit 12=SL%.
    #    The All Trades Price is the native quote for FX (JPY for USD.JPY) but
    #    already USD for stocks; derive the quote currency from the pair symbol
    #    ("USD.JPY" → JPY) so the USD notional is correct for both.
    for t in all_trade_rows:
        sr += 1
        contract = str(t[3] or "")
        quote_ccy = contract.split(".")[-1] if "." in contract else BASE_CURRENCY
        total = _total_amount(_num_cell(t[6]), _num_cell(t[7]), None,
                              quote_ccy, fx_rates)
        rows.append([
            sr, t[3], t[4], t[5], t[8],
            t[9], t[11], t[10], t[12],
            t[6], total,
            STATUS_FILLED,
        ])

    # Name falls back to the Contract (ticker) whenever it's blank — pending and
    # cancelled orders carry no Flex 'description', so their Name column would
    # otherwise be empty.
    for row in rows:
        if not row[2]:          # Name column
            row[2] = row[1]     # Contract column
    return rows

TODAY_HEADERS = ["Date & Time (UTC)", "Date & Time (GST)", "Sr No", "Contract", "Name",
                 "Action", "Quantity", "Price",
                 "Order Type", "Trigger Price", "Offset", "Limit Price", "SL %",
                 "Exchange", "Commission", "Realized PnL", "Account"]
TODAY_WIDTHS  = [22, 22, 7, 14, 30, 10, 12, 14, 12, 14, 12, 14, 10, 12, 14, 14, 14]

# Gulf Standard Time (Dubai) is UTC+4 — used for the second All Trades time column.
_GST_OFFSET_HOURS = 4

# Columns whose numeric cells get green (positive) / red (negative) shading.
# Commission is intentionally excluded — it is shown with no colour.
COLOR_COLUMNS_TRADES = ["Realized PnL"]


def _parse_exec_time(raw):
    """Parse IBKR execution.time string into a display-friendly datetime."""
    raw        = str(raw).strip()
    normalised = " ".join(raw.split())
    for fmt in ("%Y%m%d %H:%M:%S", "%Y%m%d"):
        try:
            return dt.datetime.strptime(normalised, fmt).strftime("%d-%b-%Y %H:%M:%S")
        except ValueError:
            continue
    return raw


def _to_gst(utc_str):
    """Convert a formatted UTC datetime string to Gulf Standard Time (Dubai, UTC+4).
    Returns the shifted 'DD-Mon-YYYY HH:MM:SS' string, or the original unchanged
    when there is no parseable time component to shift."""
    if not utc_str:
        return ""
    try:
        d = dt.datetime.strptime(str(utc_str), "%d-%b-%Y %H:%M:%S")
    except ValueError:
        return utc_str   # date-only or unparseable — leave as-is
    return (d + dt.timedelta(hours=_GST_OFFSET_HOURS)).strftime("%d-%b-%Y %H:%M:%S")


def _trade_price_key(dt_str, symbol, action, qty_str, price_str):
    """Stable identity for a single fill, used to carry forward manual prices.
    Built from the exact formatted strings written to the All Trades sheet."""
    return (str(dt_str), str(symbol), str(action), str(qty_str), str(price_str))


def build_individual_trade_rows(trade_rows, date_filter=None, order_lookup=None,
                                manual_prices=None):
    """
    Build per-trade rows from Flex Query data.

    date_filter   : a collection of dates to include (e.g. {today, yesterday}).
                    If None, ALL trades are included (full history).
    order_lookup  : optional ledger lookup (see build_order_lookup) used to fill
                    the Trigger / Limit columns from previously-captured orders.
    manual_prices : optional {trade_key: (trigger, limit)} carried forward from
                    the previous report, where the user typed prices by hand.
                    These take precedence over the ledger / Flex values.
    """
    if date_filter is not None:
        sel = [r for r in trade_rows if parse_trade_date(r) in date_filter]
    else:
        sel = list(trade_rows)

    # Sort newest-first (descending) by parsed date, then raw timestamp.
    sel.sort(key=lambda r: (parse_trade_date(r) or dt.date.min,
                            r.get("dateTime") or r.get("tradeDate") or ""),
             reverse=True)

    def _fmt(v, dp=2):
        """Absolute-value format (for qty/price/commission)."""
        try:
            return f"{abs(float(v)):.{dp}f}" if v not in (None, "") else ""
        except (ValueError, TypeError):
            return str(v) if v else ""

    def _amt(v, signed=False):
        """Numeric amount (so the cell can carry the 1,234.00 number format).
        Returns a rounded float (abs unless signed) or '' when not numeric."""
        try:
            x = round(float(v), 2) if v not in (None, "") else ""
        except (ValueError, TypeError):
            return str(v) if v else ""
        return x if (x == "" or signed) else abs(x)

    rows = []
    for sr, r in enumerate(sel, start=1):
        symbol   = r.get("symbol") or r.get("description") or ""
        raw_side = r.get("buySell", "").upper().strip()
        action   = ("BUY"  if raw_side in ("BUY", "BOT", "B") else
                    "SELL" if raw_side in ("SELL", "SLD", "S") else raw_side)
        account  = (r.get("accountId") or r.get("acctId") or
                    r.get("account") or "")
        commission = r.get("ibCommission") or r.get("commission")
        name       = r.get("description") or r.get("underlyingSymbol") or ""
        # Trigger / limit are order-level fields, absent from Flex trades.
        # Prefer the captured order ledger; fall back to any Flex fields.
        led_lim, led_trg = lookup_order_prices(order_lookup, symbol, action,
                                               r.get("orderType"), parse_trade_date(r))
        trigger = (led_trg if led_trg is not None
                   else r.get("triggerPrice") or r.get("auxPrice") or r.get("stopPrice"))
        limit   = (led_lim if led_lim is not None
                   else r.get("limitPrice") or r.get("lmtPrice") or r.get("orderLimitPrice"))

        dt_str    = parse_trade_datetime(r)
        qty_str   = _fmt(r.get("quantity"),   0)
        price_str = _fmt(r.get("tradePrice"), 4)
        trigger_cell = _fmt(trigger, 4)
        limit_cell   = _fmt(limit,   4)

        # Carry forward any trigger/limit the user typed into the previous
        # report for this same fill — manual edits win over ledger/Flex values.
        if manual_prices:
            m_trg, m_lim = manual_prices.get(
                _trade_price_key(dt_str, symbol, action, qty_str, price_str),
                (None, None))
            if m_trg not in (None, ""):
                trigger_cell = m_trg
            if m_lim not in (None, ""):
                limit_cell = m_lim

        # Offset = |limit - trigger|, mirroring the Pending Order sheet, computed
        # from whatever trigger/limit the row ended up with (ledger/Flex/manual).
        def _num(v):
            try:
                return float(str(v).replace(",", "")) if v not in (None, "") else None
            except (ValueError, TypeError):
                return None
        t_num, l_num = _num(trigger_cell), _num(limit_cell)
        offset_cell = f"{abs(l_num - t_num):.4f}" if (t_num is not None and l_num is not None) else ""
        # SL % is an entry-vs-stop relationship that only exists for a live order
        # pair, not a single executed fill, so it stays blank here (column present
        # to mirror the Pending Order layout).
        sl_pct_cell = ""

        # Realized PnL: FIFO-derived USD value (see attach_fifo_realized), which
        # is populated on squared-off fills and sums to the Trade Summary PnL.
        # Fall back to IBKR's fifoPnlRealized if the FIFO pass didn't run.
        realized = (r["_realized_usd"] if "_realized_usd" in r
                    else r.get("fifoPnlRealized"))

        rows.append([
            dt_str,             # Date & Time (UTC)
            _to_gst(dt_str),    # Date & Time (GST) — Dubai, UTC+4
            sr,
            symbol,
            name,
            action,
            qty_str,
            price_str,
            r.get("orderType", ""),   # Order Type
            trigger_cell,             # Trigger Price
            offset_cell,              # Offset
            limit_cell,               # Limit Price
            sl_pct_cell,              # SL %
            r.get("exchange", ""),    # Exchange
            _amt(commission),         # Commission
            _amt(realized, signed=True),   # Realized PnL
            account,                  # Account
        ])
    return rows


def build_today_rows_from_flex(trade_rows):
    """Today Orders fallback: per-trade rows for today and yesterday only."""
    today = dt.date.today()
    return build_individual_trade_rows(trade_rows, {today, today - dt.timedelta(days=1)})


def _list_previous_reports():
    """All prior report files in the reports folder, newest first (by the date
    encoded in the filename, then modified time), excluding today's OUTPUT_FILE.
    Accepts the current "MIS_" prefix and the legacy "Reports_" prefix."""
    try:
        candidates = []
        for fn in os.listdir(_REPORTS_DIR):
            if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
                continue
            path = os.path.join(_REPORTS_DIR, fn)
            if os.path.abspath(path) == os.path.abspath(OUTPUT_FILE):
                continue
            d = dt.date.min
            for prefix in ("MIS_", "Reports_"):
                try:
                    d = dt.datetime.strptime(fn, prefix + "%d%b%Y.xlsx").date()
                    break
                except ValueError:
                    continue
            candidates.append((d, os.path.getmtime(path), path))
    except OSError:
        return []
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
    return [c[2] for c in candidates]


# Manually-maintained sheets copied VERBATIM from the previous report — the
# header row AND every data row are reproduced exactly as the user left them,
# with no blank entry rows inserted. New entries are typed into the empty rows
# Excel always provides beneath the data.
_VERBATIM_SHEETS = ("Strategy Details", "Bugs", "Shubham_Activity",
                    "Ajay_Activity", "Pending_Task")

# Some manual sheets were first created by hand under a different tab name.
# When carrying data forward we accept those legacy names so the existing rows
# are picked up on the first run, then keep writing under the canonical name.
_SHEET_ALIASES = {
    "Pending_Task": ["Pending_Task", "Priorities task ", "Priorities task"],
}


def _resolve_sheet(wb, sheet_name):
    """Return the actual worksheet for sheet_name, honouring legacy aliases."""
    for nm in _SHEET_ALIASES.get(sheet_name, [sheet_name]):
        if nm in wb.sheetnames:
            return wb[nm]
    return None


def _capture_cell_style(c):
    """Snapshot a source cell's full styling, detached from its workbook so it
    survives the source being closed. Returns a dict applied verbatim when the
    cell is re-written, so borders, fills (highlights), fonts and alignment all
    come across exactly as the user left them."""
    from copy import copy
    return {
        "number_format": c.number_format,
        "font":          copy(c.font),
        "fill":          copy(c.fill),
        "border":        copy(c.border),
        "alignment":     copy(c.alignment),
    }


def _expand_column_widths(ws):
    """Return {1-based column index: width} for a worksheet, expanding the
    column RANGES openpyxl stores (a single dimension keyed by its first letter
    can span min..max columns). Without expanding, the trailing columns in a
    range lose their width and fall back to the default — making text wrap and
    rows grow. Requires a fully-loaded (non read-only) worksheet."""
    widths = {}
    for key, dim in ws.column_dimensions.items():
        if dim.width is None:
            continue
        lo = dim.min or column_index_from_string(key)
        hi = dim.max or lo
        for idx in range(lo, hi + 1):
            widths[idx] = dim.width
    return widths


def _read_manual_sheet(wb, sheet_name):
    """Copy a manual sheet verbatim from a previous report.

    Returns (headers, data_rows, layout):
      - headers  : the column names exactly as they appear in row 2 (so any
                   rename the user made is preserved), trailing blanks trimmed.
      - data_rows: every row from row 3 onward, as-is, with fully-blank rows
                   dropped so leftover entry rows don't pile up run after run.
                   Each cell is captured as a (value, style_dict) pair, where
                   style_dict holds the original number format AND the cell's
                   borders/fill/font/alignment — so the sheet is reproduced
                   pixel-for-pixel, including any manual highlights.
      - layout   : {"col_widths": {col_idx: width},
                    "row_heights": [height-or-None per kept data row]} so the
                   carried sheet keeps the exact column widths / row heights the
                   user set, instead of the script's default sizing.
    """
    ws = _resolve_sheet(wb, sheet_name)
    if ws is None:
        return [], [], {"col_widths": {}, "row_heights": []}
    all_rows = list(ws.iter_rows(min_row=2))
    if not all_rows:
        return [], [], {"col_widths": {}, "row_heights": []}
    headers = ["" if c.value is None else c.value for c in all_rows[0]]
    while headers and headers[-1] == "":
        headers.pop()
    n = len(headers)
    data, row_heights = [], []
    for cells in all_rows[1:]:
        cells = list(cells)[:n] if n else list(cells)
        if not cells or all(c.value in (None, "") for c in cells):
            continue
        data.append([("" if c.value is None else c.value, _capture_cell_style(c))
                     for c in cells])
        row_heights.append(ws.row_dimensions[cells[0].row].height)
    layout = {"col_widths": _expand_column_widths(ws), "row_heights": row_heights}
    return headers, data, layout


def _read_alltrades_prices(wb):
    """Build {trade_key: (trigger, limit)} from a workbook's 'All Trades' sheet.

    Columns are resolved by HEADER NAME (row 2), not fixed positions, so prices
    carry forward regardless of the file's column layout — older 14-column
    reports (single "Date & Time") and the current 15-column UTC/GST layout both
    read correctly, and future column changes won't silently break this."""
    prices = {}
    if "All Trades" not in wb.sheetnames:
        return prices
    rows = list(wb["All Trades"].iter_rows(min_row=2, values_only=True))
    if not rows:
        return prices
    header = [str(c) if c is not None else "" for c in rows[0]]

    def _col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_dt  = _col("Date & Time (UTC)", "Date & Time")
    c_con = _col("Contract")
    c_act = _col("Action")
    c_qty = _col("Quantity")
    c_prc = _col("Price")
    c_trg = _col("Trigger Price")
    c_lim = _col("Limit Price")
    if None in (c_dt, c_con, c_act, c_qty, c_prc, c_trg, c_lim):
        return prices

    for row in rows[1:]:
        if not row:
            continue
        g = lambda i: row[i] if i < len(row) else None
        trigger, limit = g(c_trg), g(c_lim)
        if trigger in (None, "") and limit in (None, ""):
            continue
        key = _trade_price_key(g(c_dt) or "", g(c_con) or "", g(c_act) or "",
                               g(c_qty) or "", g(c_prc) or "")
        prices[key] = (
            str(trigger) if trigger not in (None, "") else None,
            str(limit)   if limit   not in (None, "") else None,
        )
    return prices


def load_previous_report():
    """Open the closest prior report (e.g. MIS_04Jun2026.xlsx) and pull forward
    the data the user maintains by hand, which would otherwise be lost because
    every run writes a brand-new dated file:

      - manual Trigger/Limit prices from the 'All Trades' sheet, keyed by
        (datetime, contract, action, qty, price) so they re-attach to the same fill
      - the running rows of every manually-filled sheet: Strategy Details, Bugs,
        Shubham_Activity and Ajay_Activity

    Returns (manual_prices: dict, carried: {sheet_name: list-of-rows}).
    """
    from openpyxl import load_workbook

    reports = _list_previous_reports()
    if not reports:
        return {}, {}
    prev_path = reports[0]

    try:
        # Not read_only: openpyxl only exposes column widths / row heights on a
        # fully-loaded worksheet, and we carry those forward verbatim. These
        # daily reports are small, so the extra memory is negligible.
        wb = load_workbook(prev_path, data_only=True)
    except Exception as e:                                  # noqa: BLE001
        print(f"  [Carry-forward] could not open {prev_path}: {e}", file=sys.stderr)
        return {}, {}

    # Trigger/limit prices from the most recent report.
    manual_prices = _read_alltrades_prices(wb)

    # Manual sheets copied verbatim (headers + every data row, no blanks).
    carried = {name: _read_manual_sheet(wb, name) for name in _VERBATIM_SHEETS}

    wb.close()

    # Self-heal: if the latest report carries no trigger/limit prices at all
    # (e.g. a column-layout change once broke the chain), fall back to the most
    # recent older report that still has them so the prices aren't lost.
    if not manual_prices:
        for older in reports[1:]:
            try:
                w2 = load_workbook(older, read_only=True, data_only=True)
            except Exception:                              # noqa: BLE001
                continue
            mp = _read_alltrades_prices(w2)
            w2.close()
            if mp:
                manual_prices = mp
                print(f"  [Carry-forward] recovered {len(mp)} trigger/limit "
                      f"price(s) from {os.path.basename(older)}.")
                break
    def _n_rows(v):
        # Verbatim sheets store (headers, rows); data-only sheets store rows.
        return len(v[1]) if isinstance(v, tuple) else len(v)
    counts = ", ".join(f"{_n_rows(v)} {k}" for k, v in carried.items())
    print(f"  [Carry-forward] {len(manual_prices)} trigger/limit price(s); "
          f"{counts} from {os.path.basename(prev_path)}.")
    return manual_prices, carried


def build_trade_rows(executions):
    """Supplementary: live fills received from reqExecutions this session.
    Aligned to TODAY_HEADERS; Order Type / Realized PnL aren't carried by the
    execution feed, so those cells are left blank for live rows."""
    rows = []
    for sr, ex in enumerate(
        sorted(executions.values(), key=lambda e: str(e["time"])), start=1
    ):
        exec_time = _parse_exec_time(ex["time"])
        rows.append([
            exec_time,                                       # Date & Time (UTC)
            _to_gst(exec_time),                              # Date & Time (GST) — Dubai, UTC+4
            sr,
            ex["contract"],
            "",                                              # Name (not in exec feed)
            ex["action"],
            ex["quantity"],
            f"{ex['price']:.4f}"      if ex["price"]      is not None else "",
            "",                                              # Order Type (n/a)
            "",                                              # Trigger Price (n/a)
            "",                                              # Offset (n/a)
            "",                                              # Limit Price (n/a)
            "",                                              # SL % (n/a)
            ex["exchange"],
            round(abs(ex["commission"]), 2) if ex["commission"] is not None else "",
            "",                                              # Realized PnL (n/a)
            ex["account"],
        ])
    return rows


def build_fx_rate_map(flex_rows, open_positions=None):
    """currency -> fxRateToBase, so any non-USD amount anywhere in the report can
    be converted to USD (live TWS fills and pending orders omit the rate, and a
    currency with no rate at all would silently leak native JPY/etc. figures).

    Sourced, in priority order, keeping the most recent value per currency:
      1. fxRateToBase on the Flex trades (they carry it directly);
      2. fxRateToBase on the Flex <OpenPosition> records;
      3. FALLBACK — the rate implied by a USD-based FX pair's own price. USD.JPY
         quoted at 161.7 means '161.7 JPY per 1 USD', so 1 JPY = 1/161.7 USD.
         Only used for a currency still missing after (1)/(2), so an account that
         trades a pair but whose trades lack fxRateToBase still converts."""
    best = {}                                   # ccy -> (date, rate)

    def _consider(ccy, raw_rate, d):
        ccy = (ccy or "").upper()
        if not ccy or ccy == BASE_CURRENCY:
            return
        try:
            rate = float(raw_rate)
        except (ValueError, TypeError):
            return
        if rate <= 0:
            return
        if ccy not in best or d >= best[ccy][0]:
            best[ccy] = (d, rate)

    for r in flex_rows:
        _consider(r.get("currency"), r.get("fxRateToBase"),
                  parse_trade_date(r) or dt.date.min)
    for op in (open_positions or []):
        _consider(op.get("currency"), op.get("fxRateToBase"), dt.date.min)

    # Fallback: derive a still-missing quote currency's rate from a USD-based FX
    # pair's own trade price (only when USD is the pair's base, so the price is
    # 'quote units per 1 USD').
    derived = {}                                # ccy -> (date, rate)
    for r in flex_rows:
        sym = (r.get("symbol") or "").upper()
        base_ccy, _, quote_ccy = sym.partition(".")
        if base_ccy != BASE_CURRENCY or not quote_ccy or quote_ccy == BASE_CURRENCY:
            continue
        if quote_ccy in best:
            continue
        price = _flt(r.get("tradePrice"))
        if price <= 0:
            continue
        d = parse_trade_date(r) or dt.date.min
        if quote_ccy not in derived or d >= derived[quote_ccy][0]:
            derived[quote_ccy] = (d, 1.0 / price)

    out = {ccy: rate for ccy, (_d, rate) in best.items()}
    for ccy, (_d, rate) in derived.items():
        out.setdefault(ccy, rate)
    return out


def _executions_to_flex_rows(executions, fx_rates=None):
    """Reshape this session's live TWS fills (reqExecutions) into the same
    Flex-trade dicts parse_trades produces, so today's buy/sell fills can flow
    into All Trades / Trade Summary / Open Positions before IBKR's end-of-day
    Flex batch includes them.

    fx_rates (currency -> fxRateToBase, see build_fx_rate_map) supplies the
    conversion rate the execution feed omits, so a non-USD fill (e.g. the JPY leg
    of USD.JPY) is normalized to USD consistently with its Flex counterpart —
    otherwise one leg stays in JPY and the summed PnL is nonsense."""
    fx_rates = fx_rates or {}
    out = []
    for exec_id, ex in executions.items():
        qty   = _flt(ex.get("quantity"))
        price = _flt(ex.get("price"))
        sectype = (ex.get("secType") or "").upper()
        ccy = (ex.get("currency") or BASE_CURRENCY)
        # Tag the fill with the rate for its currency so normalize_trades_to_usd
        # converts it. USD legs keep 1.0; a missing rate falls back to 1.0 (the
        # fill stays native until the Flex batch, which carries the rate, lands).
        rate = fx_rates.get(ccy.upper()) if ccy.upper() != BASE_CURRENCY else None
        # For FX, Flex keys the trade by the pair (localSymbol, e.g. "EUR.USD"),
        # not the bare base currency; use it so the symbol, dedup key, and CASH
        # detection all match the Flex history.
        symbol = (ex.get("localSymbol") or ex.get("contract") or ""
                  ) if sectype == "CASH" else (ex.get("contract") or "")
        out.append({
            "symbol":          symbol,
            "description":     ex.get("localSymbol") or "",
            "assetCategory":   sectype,       # STK/CASH/OPT/... — drives CASH handling
            "conid":           ex.get("conId"),
            # execId lets merge dedup exactly against the Flex ibExecID once the
            # Flex batch catches up, independent of the fuzzy field match.
            "ibExecID":        exec_id,
            "buySell":         (ex.get("action") or "").upper(),
            "quantity":        qty,
            "tradePrice":      price,
            "tradeMoney":      qty * price,
            "ibCommission":    ex.get("commission"),
            # Forex commissions are billed in USD; default to USD for CASH so we
            # don't FX-convert an already-USD commission (see normalize).
            "ibCommissionCurrency": (ex.get("commission_currency")
                                     or (BASE_CURRENCY if sectype == "CASH" else ccy)),
            "fifoPnlRealized": 0.0,           # not carried by the execution feed
            "exchange":        ex.get("exchange") or "",
            "dateTime":        str(ex.get("time") or ""),
            "currency":        ccy,
            "fxRateToBase":    rate,          # None -> normalize treats as 1.0
            "orderType":       "",
            "accountId":       ex.get("account") or "",
        })
    return out


def merge_todays_executions(flex_rows, executions, fx_rates=None):
    """Splice live TWS fills (this session's + every prior day's, from the trades
    ledger) into the Flex trade rows in place. A fill is skipped only when the
    Flex batch already contains an equivalent trade (matched on
    date/symbol/side/qty/price), so nothing is double-counted once Flex catches
    up — but two genuinely-identical fills on the same day are both kept, because
    each carries a distinct IBKR execId.

    fx_rates (see build_fx_rate_map) converts non-USD fills to USD, so a live leg
    (e.g. today's JPY sell of USD.JPY) is on the same footing as its Flex leg."""
    if not executions:
        return flex_rows

    def _key(r):
        # abs() on qty/price is essential: IBKR Flex reports SELL quantities as
        # NEGATIVE (-50000) while the execution feed stores them positive, so
        # without abs the same sell never matches and gets duplicated. The side
        # (B/S) already encodes direction, so abs loses nothing.
        return (parse_trade_date(r),
                (r.get("symbol") or "").upper(),
                (r.get("buySell") or "").upper()[:1],       # B(UY)/B(OT), S(ELL)/S(LD)
                round(abs(_flt(r.get("quantity"))),   2),
                round(abs(_flt(r.get("tradePrice"))), 4))

    def _exec_id(r):
        return str(r.get("ibExecID") or r.get("execID") or r.get("ibExecId") or "").strip()

    # Dedup against the Flex rows ONLY (frozen up front): execution rows are
    # already unique by execId, so we must not suppress a second identical fill.
    flex_keys     = {_key(r) for r in flex_rows}
    flex_exec_ids = {e for e in (_exec_id(r) for r in flex_rows) if e}
    added = 0
    for r in _executions_to_flex_rows(executions, fx_rates):
        # Skip if the Flex batch already has this fill — by exact execId (when the
        # Flex query exposes ibExecID) or by the fuzzy field key otherwise.
        if _exec_id(r) in flex_exec_ids or _key(r) in flex_keys:
            continue
        flex_rows.append(r)
        added += 1
    print(f"  [Live fills] merged {added} TWS execution(s) not present in the Flex batch "
          f"(from {len(executions)} ledgered fill(s)).")
    return flex_rows


# ----------------------------------------------------------------------
# STEP 1: request the statement -> reference code
# ----------------------------------------------------------------------
def send_request():
    r = requests.get(SEND_URL, params={"t": TOKEN, "q": QUERY_ID, "v": VERSION}, timeout=30)
    r.raise_for_status()
    root = ET.fromstring(r.content)
    status = root.findtext("Status")
    if status != "Success":
        code = root.findtext("ErrorCode")
        msg  = root.findtext("ErrorMessage")
        sys.exit(f"SendRequest failed [{code}]: {msg}")
    ref = root.findtext("ReferenceCode")
    print(f"Reference code: {ref}")
    return ref


# ----------------------------------------------------------------------
# STEP 2: poll for the statement XML
# ----------------------------------------------------------------------
def get_statement(ref):
    for attempt in range(1, MAX_POLL_ATTEMPTS + 1):
        r = requests.get(GET_URL, params={"t": TOKEN, "q": ref, "v": VERSION}, timeout=60)
        r.raise_for_status()
        text = r.text
        if "<FlexQueryResponse" in text:
            print("Statement retrieved.")
            return text
        try:
            root = ET.fromstring(r.content)
            code = root.findtext("ErrorCode")
            msg  = root.findtext("ErrorMessage")
            if code and code not in ("1019",):
                sys.exit(f"GetStatement failed [{code}]: {msg}")
            print(f"  not ready yet (attempt {attempt}/{MAX_POLL_ATTEMPTS}), waiting...")
        except ET.ParseError:
            pass
        time.sleep(POLL_WAIT_SECONDS)
    sys.exit("Statement not ready after maximum retries.")


# ----------------------------------------------------------------------
# STEP 3: parse trades + account ID
# ----------------------------------------------------------------------
def parse_trades(xml_text):
    root   = ET.fromstring(xml_text)
    trades = root.findall(".//Trade")
    if not trades:
        trades = root.findall(".//Order")
    return [dict(t.attrib) for t in trades]


# Base (reporting) currency — every monetary figure in the report is expressed
# in this currency. IBKR Flex tags each trade with fxRateToBase, the rate that
# converts a value from the trade's own currency into the account base currency.
BASE_CURRENCY = "USD"

# Monetary trade fields denominated in the trade's own currency. These are the
# amounts that must be scaled by fxRateToBase so the whole report reads in USD.
# Commission is handled separately (see normalize_trades_to_usd) because IBKR
# bills it in ibCommissionCurrency, which for forex is USD — NOT the trade
# currency — so it must not be scaled by the trade's fxRateToBase.
_MONEY_FIELDS = (
    "tradeMoney", "proceeds", "netCash", "cost", "closePrice", "taxes",
    "fifoPnlRealized", "fifoPnlUnrealized", "mtmPnl", "unrealizedPnl",
)
_COMMISSION_FIELDS = ("ibCommission", "commission")


def _fx_rate_to_base(row):
    """Rate that converts this trade's currency amounts into the base currency.

    Returns 1.0 when the trade is already in the base currency or when IBKR did
    not supply a usable rate (so values pass through unchanged)."""
    if (row.get("currency") or BASE_CURRENCY).upper() == BASE_CURRENCY:
        return 1.0
    try:
        rate = float(row.get("fxRateToBase"))
    except (ValueError, TypeError):
        return 1.0
    # IBKR uses 0/blank as "no rate"; never scale by an invalid factor.
    return rate if rate > 0 else 1.0


def normalize_trades_to_usd(trade_rows, fx_rates=None):
    """Convert every trade's monetary fields into the base currency (USD) in place.

    IBKR Flex reports FX pairs (e.g. USD.JPY) and non-USD instruments with their
    money/PnL fields in the instrument's own currency. Scaling each by
    fxRateToBase yields a single-currency (USD) report. Per-unit prices
    (tradePrice) are converted for non-FX instruments so a non-USD stock shows a
    USD price; for FX (assetCategory CASH) the 'price' is an exchange rate, not a
    money amount, so it is left as quoted.

    Commission is converted by ITS OWN currency (ibCommissionCurrency), not the
    trade's: IBKR bills forex commission in USD even for a JPY-quoted pair, so
    scaling it by the JPY rate would wrongly shrink a real ~$2 fee to ~$0.01.
    fx_rates (see build_fx_rate_map) supplies the rate when the commission
    currency differs from the trade currency."""
    fx_rates = fx_rates or {}
    for r in trade_rows:
        trade_ccy = (r.get("currency") or BASE_CURRENCY).upper()
        # Everything is reported in USD. Use ONE USD rate per currency (the most
        # recent, from fx_rates) for ALL of that currency's trades — not each
        # leg's own historical rate. Otherwise an FX pair whose buy and sell
        # settled at different rates (USD.JPY: 0.0061507 vs 0.0062066) picks up a
        # translation gain that flips a real loss into a green (profit) figure.
        # With one rate, PnL correctly follows the buy-vs-sell price.
        if trade_ccy == BASE_CURRENCY:
            rate = 1.0
        else:
            rate = fx_rates.get(trade_ccy) or _fx_rate_to_base(r)

        # ── commission: convert by the commission's own currency ──────────────
        comm_ccy = (r.get("ibCommissionCurrency")
                    or (BASE_CURRENCY if (r.get("assetCategory") or "").upper() == "CASH"
                        else trade_ccy)).upper()
        if comm_ccy != BASE_CURRENCY:
            comm_rate = rate if comm_ccy == trade_ccy else fx_rates.get(comm_ccy)
            if comm_rate and comm_rate > 0:
                for field in _COMMISSION_FIELDS:
                    raw = r.get(field)
                    if raw in (None, ""):
                        continue
                    try:
                        r[field] = float(raw) * comm_rate
                    except (ValueError, TypeError):
                        continue
        r["ibCommissionCurrency"] = BASE_CURRENCY

        if rate == 1.0:
            r["currency"] = BASE_CURRENCY
            continue
        for field in _MONEY_FIELDS:
            raw = r.get(field)
            if raw in (None, ""):
                continue
            try:
                r[field] = float(raw) * rate
            except (ValueError, TypeError):
                continue
        # tradePrice: a genuine per-share price for stocks/options (convert to
        # USD), but an FX quote for CASH pairs (leave alone).
        if (r.get("assetCategory") or "").upper() != "CASH":
            tp = r.get("tradePrice")
            if tp not in (None, ""):
                try:
                    r["tradePrice"] = float(tp) * rate
                except (ValueError, TypeError):
                    pass
        # Now expressed in base currency — reflect that so nothing downstream
        # mistakes the row for its original currency.
        r["currency"] = BASE_CURRENCY
    return trade_rows


# Live-position money fields (from reqPositions / updatePortfolio) that must read
# in USD on the Open Position sheet and the Dashboard's unrealized total.
_POSITION_MONEY_FIELDS = ("marketValue", "unrealizedPNL", "realizedPNL")


def normalize_positions_to_usd(live_positions, fx_rates=None):
    """Convert each live position's money fields (market value, unrealized &
    realized PnL) and its average cost into USD in place, so the Open Position
    sheet and the Dashboard never show a non-USD (e.g. JPY) figure for an FX pair.

    IBKR's live portfolio feed reports these in the position's own quote currency
    for FX pairs (JPY for USD.JPY); scaling by that currency's fxRateToBase (see
    build_fx_rate_map) expresses them in USD. USD positions — and any currency
    with no known rate — pass through unchanged. averageCost for an FX pair is the
    quote rate (JPY per USD), so converting it yields the ~1.0 USD-normalized avg
    the rest of the report already uses (matching normalize_trades_to_usd)."""
    if not live_positions:
        return live_positions
    fx_rates = fx_rates or {}
    for p in live_positions:
        ccy = (p.get("currency") or BASE_CURRENCY).upper()
        if ccy == BASE_CURRENCY:
            continue
        rate = fx_rates.get(ccy)
        if not rate or rate <= 0:
            continue
        for field in (*_POSITION_MONEY_FIELDS, "averageCost", "avgCost"):
            v = p.get(field)
            if _ibkr_has(v):
                p[field] = float(v) * rate
    return live_positions


def attach_fifo_realized(trade_rows):
    """Store per-fill realized PnL (USD) on each row as '_realized_usd', derived
    by FIFO-matching each symbol's fills chronologically.

    IBKR's own fifoPnlRealized is 0 / unreliable for FX, and the Trade Summary
    derives PnL as (sell_value - buy_value) in USD, so we mirror that here at the
    fill level: a closing fill carries the realized PnL of the lots it closes, and
    for a squared-off symbol the fills sum exactly to the Trade Summary PnL. Must
    run AFTER normalize_trades_to_usd so tradeMoney is already in USD."""
    by_symbol = defaultdict(list)
    for r in trade_rows:
        sym = r.get("symbol") or r.get("description") or "UNKNOWN"
        by_symbol[sym].append(r)

    for _sym, rs in by_symbol.items():
        chron = sorted(rs, key=lambda r: (parse_trade_date(r) or dt.date.min,
                                          r.get("dateTime") or r.get("tradeDate") or ""))
        lots = deque()                          # each: [signed_qty, usd_unit_price]
        for r in chron:
            qty = abs(_flt(r.get("quantity")))
            if qty <= 1e-9:
                r["_realized_usd"] = 0.0
                continue
            is_buy = (r.get("buySell") or "").upper() in ("BUY", "BOT", "B")
            sign   = 1 if is_buy else -1
            money  = abs(_flt(r.get("tradeMoney"))) or (qty * abs(_flt(r.get("tradePrice"))))
            unit   = money / qty
            remaining, realized = qty, 0.0
            # Close opposite-sign lots FIFO; realized = matched * (sell - buy)/unit.
            while remaining > 1e-9 and lots and (lots[0][0] > 0) != is_buy:
                lot     = lots[0]
                lot_qty = abs(lot[0])
                matched = min(remaining, lot_qty)
                realized += (matched * (unit - lot[1]) if lot[0] > 0
                             else matched * (lot[1] - unit))
                if lot_qty - matched <= 1e-9:
                    lots.popleft()
                else:
                    lot[0] = (lot_qty - matched) * (1 if lot[0] > 0 else -1)
                remaining -= matched
            if remaining > 1e-9:                 # leftover opens a new lot
                lots.append([sign * remaining, unit])
            r["_realized_usd"] = round(realized, 2)
    return trade_rows


def parse_open_positions(xml_text):
    """Authoritative open positions from the Flex statement's <OpenPosition>
    section.

    Netting the trade history alone (aggregate) misses any position whose opening
    trades fall outside the statement's date range, so we read IBKR's own open-
    position snapshot instead. Returns [] when the query has no OpenPosition
    section configured, in which case the caller falls back to trade netting."""
    root = ET.fromstring(xml_text)
    return [dict(op.attrib) for op in root.findall(".//OpenPosition")]


def _flex_open_to_position(op, fx_rates=None):
    """Map a Flex <OpenPosition> record onto the same position-dict shape the live
    reqPositions/updatePortfolio feed produces, so _build_live_open_positions can
    consume either source. Monetary fields are converted to the base currency
    (USD); per-unit prices are converted for non-FX instruments only (see
    normalize_trades_to_usd for the same rule).

    The conversion rate prefers the shared fx_rates map (build_fx_rate_map), so a
    record missing its own fxRateToBase still converts instead of leaking a native
    (JPY) figure; it falls back to the record's own rate."""
    ccy     = (op.get("currency") or BASE_CURRENCY).upper()
    rate    = ((fx_rates or {}).get(ccy) if ccy != BASE_CURRENCY else 1.0) \
              or _fx_rate_to_base(op)
    is_cash = (op.get("assetCategory") or "").upper() == "CASH"

    def _money(key):
        v = op.get(key)
        if v in (None, ""):
            return None
        try:
            return float(v) * rate
        except (ValueError, TypeError):
            return None

    def _price(key):
        v = op.get(key)
        if v in (None, ""):
            return None
        try:
            f = float(v)
        except (ValueError, TypeError):
            return None
        return f * rate if not is_cash else f      # FX 'price' is a quote, not money

    try:
        qty = float(op.get("position") or 0)
    except (ValueError, TypeError):
        qty = 0.0
    try:
        conid = int(float(op.get("conid"))) if op.get("conid") else None
    except (ValueError, TypeError):
        conid = None

    return {
        "conId":         conid,
        "symbol":        op.get("symbol") or "",
        "secType":       op.get("assetCategory") or "",
        # Keep the position's own (quote) currency: _build_live_open_positions
        # uses it solely to build the FX pair for matching, exactly as the live
        # reqPositions feed reports it (e.g. USD.JPY -> symbol USD, currency JPY).
        "currency":      op.get("currency") or "",
        "name":          op.get("description") or op.get("symbol") or "",
        "exchange":      op.get("listingExchange") or op.get("exchange") or "",
        "position":      qty,
        "averageCost":   _price("costBasisPrice") or _price("openPrice"),
        "marketValue":   _money("positionValue"),
        "unrealizedPNL": _money("fifoPnlUnrealized"),
    }


def parse_account_id(xml_text):
    root = ET.fromstring(xml_text)
    stmt = root.find(".//FlexStatement")
    if stmt is not None:
        return stmt.get("accountId", "N/A")
    return "N/A"


def parse_trade_date(row):
    """Extract a date from common IBKR datetime attributes."""
    raw = row.get("dateTime") or row.get("tradeDate") or row.get("reportDate") or ""
    raw = raw.strip()
    if not raw:
        return None
    raw = raw.replace(",", " ").replace(";", " ")
    datepart = raw.split()[0]
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(datepart, fmt).date()
        except ValueError:
            continue
    return None


def parse_trade_datetime(row):
    """Extract a full datetime string from IBKR attributes, formatted for display.

    IBKR Flex sends dateTime as 'DD/MM/YYYY;HHMMSS' (e.g. '20/05/2026;080617'),
    but other configs use 'YYYYMMDD;HHMMSS' or ISO. We normalise separators
    then parse the date and time parts independently.
    """
    raw = (row.get("dateTime") or row.get("tradeDate") or "").strip()
    if not raw:
        return ""
    norm  = raw.replace(";", " ").replace(",", " ")
    parts = norm.split()
    datepart = parts[0]
    timepart = parts[1] if len(parts) > 1 else ""

    d = None
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            d = dt.datetime.strptime(datepart, fmt)
            break
        except ValueError:
            continue
    if d is None:
        return raw

    if timepart:
        for tfmt in ("%H%M%S", "%H:%M:%S"):
            try:
                t = dt.datetime.strptime(timepart, tfmt).time()
                return dt.datetime.combine(d.date(), t).strftime("%d-%b-%Y %H:%M:%S")
            except ValueError:
                continue
    return d.strftime("%d-%b-%Y")


# ----------------------------------------------------------------------
# STEP 4: aggregate raw trades into per-contract summary rows
# ----------------------------------------------------------------------
HEADERS = [
    "First Trade Date (UTC)", "First Trade Date (GST)",
    "Last Trade Date (UTC)",  "Last Trade Date (GST)",
    "Contract", "Name", "Buys", "Sells", "Net",
    "Avg (bought)", "Avg (sold)",
    "Total (bought)", "Total (sold)",
    "Exchange List",
    "Commission",
    "PnL", "Unrealized PnL",
]

# Trade Summary columns shaded green (positive) / red (negative).
# Commission is intentionally excluded — it is shown with no colour.
COLOR_COLUMNS_SUMMARY = ["PnL", "Unrealized PnL"]

# Money columns formatted as 1,234.00 on the aggregate sheets.
AMOUNT_COLUMNS_SUMMARY = ["Total (bought)", "Total (sold)", "Commission",
                          "PnL", "Unrealized PnL"]

# Average price columns need finer precision (e.g. 1.14205), so they get a
# 5-decimal format instead of the 2-decimal money format above.
PRICE_COLUMNS_SUMMARY = ["Avg (bought)", "Avg (sold)"]


def _flt(val):
    try:
        return float(val) if val not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def aggregate(trade_rows):
    data = defaultdict(lambda: {
        "buy_qty": 0.0, "sell_qty": 0.0,
        "buy_value": 0.0, "sell_value": 0.0,
        "commission": 0.0, "pnl": 0.0,
        "exchanges": set(),
        "dates": [],
        "name": "",
    })

    for r in trade_rows:
        symbol     = r.get("symbol") or r.get("description") or "UNKNOWN"
        name       = r.get("description") or r.get("underlyingSymbol") or ""
        side       = r.get("buySell", "").upper().strip()
        qty        = abs(_flt(r.get("quantity", 0)))
        price      = abs(_flt(r.get("tradePrice", 0)))
        money      = abs(_flt(r.get("tradeMoney", 0))) or qty * price
        # IBKR Flex uses 'ibCommission'; the TWS API uses 'commission'.
        commission = _flt(r.get("ibCommission") or r.get("commission") or 0)
        pnl        = _flt(r.get("fifoPnlRealized", 0))
        exchange   = r.get("exchange", "")

        c = data[symbol]
        if side in ("BUY", "B"):
            c["buy_qty"]   += qty
            c["buy_value"] += money
        elif side in ("SELL", "S"):
            c["sell_qty"]   += qty
            c["sell_value"] += money
        c["commission"] += commission
        c["pnl"]        += pnl
        if exchange:
            c["exchanges"].add(exchange)
        if name and not c["name"]:
            c["name"] = name
        d = parse_trade_date(r)
        if d:
            c["dates"].append((d, parse_trade_datetime(r)))

    result = []
    for symbol, c in data.items():
        avg_b     = c["buy_value"]  / c["buy_qty"]  if c["buy_qty"]  else 0.0
        avg_s     = c["sell_value"] / c["sell_qty"] if c["sell_qty"] else 0.0
        net_qty   = c["buy_qty"] - c["sell_qty"]
        # Realized P&L mirrors IBKR's "Net Total" = Total (sold) - Total (bought).
        # IBKR's fifoPnlRealized/mtmPnl fields are unreliable for FX trades — they
        # arrive as 0 (realized) or as a mark-to-market figure with the wrong sign
        # (mtmPnl), so derive realized P&L from the traded values instead.
        realized   = c["sell_value"] - c["buy_value"]
        # The Flex trade history carries NO genuine unrealized figure — its
        # mtmPnl is mark-to-market on the (mostly closed) fills, not the open
        # position's unrealized PnL, and for FX it leaks a large phantom value
        # onto a tiny fractional residual (e.g. a -0.75 leftover on a 43,750
        # round-trip). Unrealized only exists for a live open position, so it is
        # sourced solely from IBKR's live portfolio feed (see the Open Position
        # sheet and _live_unrealized_total); here it is always 0.
        unrealized = 0.0
        datetimes = sorted(c["dates"], key=lambda x: x[0])
        result.append({
            "Contract":             symbol,
            "Name":                 c["name"],
            "Buys":                 round(c["buy_qty"], 4),
            "Sells":                round(c["sell_qty"], 4),
            "Net":                  round(net_qty, 4),
            "Avg (bought)":         round(avg_b, 6),
            "Avg (sold)":           round(avg_s, 6),
            "Total (bought)":       round(c["buy_value"], 2),
            "Total (sold)":         round(c["sell_value"], 2),
            "Exchange List":        ", ".join(sorted(c["exchanges"])),
            "Commission":           round(c["commission"], 2),
            "PnL":                  round(realized, 2),
            "Unrealized PnL":       unrealized,
            "First Trade Date (UTC)": datetimes[0][1]  if datetimes else "",
            "First Trade Date (GST)": _to_gst(datetimes[0][1])  if datetimes else "",
            "Last Trade Date (UTC)":  datetimes[-1][1] if datetimes else "",
            "Last Trade Date (GST)":  _to_gst(datetimes[-1][1]) if datetimes else "",
        })
    result.sort(key=lambda x: x["Contract"])
    return result


# ----------------------------------------------------------------------
# STEP 5: write Excel
# ----------------------------------------------------------------------

# Style constants — clean corporate palette: navy headers, neutral text
_NAVY        = PatternFill("solid", fgColor="1F3864")   # dark navy (title + header)
_TITLE_FONT  = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
_HDR_FONT    = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
_DATA_FONT   = Font(size=10, color="262626", name="Calibri")   # near-black body text
_LABEL_FONT  = Font(bold=True, size=10, color="1F3864", name="Calibri")
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center")
_RIGHT       = Alignment(horizontal="right",  vertical="center")

# PnL cell highlights: green for profit, faint red for loss.
_PNL_POS_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green (positive)
_PNL_NEG_FILL = PatternFill("solid", fgColor="FFC7CE")   # faint red (negative)

# Thin grey gridline border applied around every cell in a table.
_GRID_SIDE   = Side(style="thin", color="BFBFBF")
_CELL_BORDER = Border(left=_GRID_SIDE, right=_GRID_SIDE, top=_GRID_SIDE, bottom=_GRID_SIDE)

# Uniform amount/money display across every sheet, e.g. 1,234.00 (and -1,234.00).
_AMOUNT_FMT = "#,##0.00"
# Average price display with 5 decimals, e.g. 1.14205.
_PRICE_FMT  = "#,##0.00000"


def _apply_amount_format(ws, col, start_row=3, fmt=_AMOUNT_FMT):
    """Apply a number format (default 1,234.00) to numeric cells in a column."""
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = fmt

# Green (Excel selection green) border drawn around each sheet's title banner.
_TITLE_SIDE   = Side(style="medium", color="00B050")
_TITLE_BORDER = Border(left=_TITLE_SIDE, right=_TITLE_SIDE, top=_TITLE_SIDE, bottom=_TITLE_SIDE)


def _apply_cell_borders(ws, n_cols, start_row=2):
    """Draw a thin border around every cell from start_row..max_row, cols 1..n_cols."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.border = _CELL_BORDER


def _color_pnl_cell(cell):
    """Shade a cell green if its numeric value is positive, faint red if negative."""
    try:
        v = float(str(cell.value).replace(",", ""))
    except (ValueError, TypeError, AttributeError):
        return
    if v > 0:
        cell.fill = _PNL_POS_FILL
    elif v < 0:
        cell.fill = _PNL_NEG_FILL

# Sheet display titles
_SHEET_TITLES = {
    "Dashboard":       "ACCOUNT DASHBOARD",
    "Pending Order":   "PENDING ORDERS",
    "Trades_Status":   "TRADES STATUS",
    "All Trades":      "ALL TRADES",
    "Open Position":   "OPEN POSITIONS",
    "Trade Summary":    "TRADE SUMMARY",
    "Strategy Details": "STRATEGY DETAILS",
    "Bugs":             "BUGS",
    "Shubham_Activity": "SHUBHAM ACTIVITY",
    "Ajay_Activity":    "AJAY ACTIVITY",
    "Pending_Task":     "PENDING TASKS",
}

# Manually-filled sheets — only the title/headers (and dropdowns) are generated.
BUGS_HEADERS = ["SR NO", "BUG_DETAILS", "Date Identified", "Severity",
                "Current_Status", "Date of Resolution"]
BUGS_WIDTHS  = [8, 70, 16, 12, 16, 18]

# Dropdown choices for the Bugs sheet severity/status columns.
BUGS_SEVERITY_CHOICES = "Critical,High,Medium,Low"
BUGS_STATUS_CHOICES   = "Fixed,Not Fixed"

STRATEGY_HEADERS = ["SR NO", "STRATEGY", "DETAILS", "STATUS"]
STRATEGY_WIDTHS  = [8, 28, 55, 14]

ACTIVITY_HEADERS = ["SR NO", "TASK", "NO. OF HOURS"]
ACTIVITY_WIDTHS  = [8, 60, 16]

# Pending Task sheet (manual entry) — same layout the user maintains by hand.
PENDING_TASK_HEADERS = ["SR NO", "Task", "COMPLETION DATE", "Severity",
                        "ETA (Expected Timeline)"]
PENDING_TASK_WIDTHS  = [11, 32, 22, 18, 24]

# Dropdown choices for the Pending Task severity column.
PENDING_TASK_SEVERITY_CHOICES = "Very High,High,Medium,Low"

# Manually-maintained sheets carry their data forward each run, copied verbatim
# from the previous report. New daily entries are typed into the empty rows
# Excel provides beneath the carried data.


def _write_title_row(ws, text, n_cols):
    """Row 1: merged dark-navy title spanning all columns, framed by a green border."""
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{end_col}1")
    cell = ws["A1"]
    cell.value     = text
    cell.font      = _TITLE_FONT
    cell.fill      = _NAVY
    cell.alignment = _CENTER
    ws.row_dimensions[1].height = 24
    # Green frame around the whole merged title banner. For a merged cell the
    # border must be set on every underlying cell so all four outer edges show.
    for col in range(1, n_cols + 1):
        ws.cell(row=1, column=col).border = _TITLE_BORDER


def _write_header_row(ws, headers, row=2):
    """Row 2: per-column headers in dark navy."""
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = _HDR_FONT
        cell.fill      = _NAVY
        cell.alignment = _CENTER
    ws.row_dimensions[row].height = 30


def _style_data_rows(ws, start_row, n_cols):
    """Apply data font and alignment to all data rows."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.font      = _DATA_FONT
            cell.alignment = _CENTER


def _apply_autofilter(ws, n_cols, header_row=2):
    """Enable Excel column filters on the header row across all columns."""
    last_col = get_column_letter(n_cols)
    last_row = max(ws.max_row, header_row)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


def _copy_reference_sheet(wb, src_ws, new_title):
    """Copy a worksheet VERBATIM (values + formatting) into wb under new_title.

    Reproduces cell values, fonts/fills/borders/alignment/number formats,
    merged ranges, column widths, row heights, freeze panes and gridline
    visibility — so the copied sheet looks exactly like the source."""
    from copy import copy

    ws = wb.create_sheet(title=new_title)
    for row in src_ws.iter_rows():
        for cell in row:
            nc = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font          = copy(cell.font)
                nc.border        = copy(cell.border)
                nc.fill          = copy(cell.fill)
                nc.alignment     = copy(cell.alignment)
                nc.protection    = copy(cell.protection)
                nc.number_format = cell.number_format
            if cell.hyperlink:
                nc.hyperlink = copy(cell.hyperlink)

    for rng in src_ws.merged_cells.ranges:
        ws.merge_cells(str(rng))

    # Column widths. A single ColumnDimension can span a RANGE of columns
    # (e.g. min=1,max=2 -> both A and B share one width), and openpyxl keys it
    # only by the first letter. Apply the width/hidden flag to EVERY column in
    # the range, or the trailing columns silently fall back to the default
    # width — which made copied text wrap and rows grow tall.
    for key, dim in src_ws.column_dimensions.items():
        lo = dim.min or column_index_from_string(key)
        hi = dim.max or lo
        for idx in range(lo, hi + 1):
            letter = get_column_letter(idx)
            if dim.width is not None:
                ws.column_dimensions[letter].width = dim.width
            ws.column_dimensions[letter].hidden = dim.hidden

    for key, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            ws.row_dimensions[key].height = dim.height
        ws.row_dimensions[key].hidden = dim.hidden

    # Sheet-level default sizing, so rows/columns without an explicit dimension
    # render at the same size as the source (not openpyxl's own defaults).
    sf = src_ws.sheet_format
    if sf.defaultRowHeight is not None:
        ws.sheet_format.defaultRowHeight = sf.defaultRowHeight
    if sf.defaultColWidth is not None:
        ws.sheet_format.defaultColWidth = sf.defaultColWidth
    if sf.baseColWidth is not None:
        ws.sheet_format.baseColWidth = sf.baseColWidth

    ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    ws.freeze_panes = src_ws.freeze_panes
    return ws


def _load_reference_sheet_pairs():
    """Open OLD_REFERENCE.xlsx and return (src_wb, [(src_ws, report_title), ...]).

    src_wb is returned so it stays open until the sheets have been copied; the
    caller must close it. Returns (None, []) when the file is missing or locked
    (e.g. open in Excel) so a run never fails just because the reference is
    unavailable — the report is simply written without the old-account sheets."""
    from openpyxl import load_workbook

    if not os.path.exists(OLD_REFERENCE_FILE):
        print(f"  [Old AC] {OLD_REFERENCE_FILE} not found - reference sheets skipped.",
              file=sys.stderr)
        return None, []
    try:
        src = load_workbook(OLD_REFERENCE_FILE)
    except Exception as e:                                  # noqa: BLE001
        print(f"  [Old AC] could not open {OLD_REFERENCE_FILE}: {e} - "
              f"reference sheets skipped.", file=sys.stderr)
        return None, []

    by_name = {ws.title.strip().lower(): ws for ws in src.worksheets}
    pairs = []
    for report_title, candidates in OLD_REFERENCE_SHEETS:
        ws_src = next((by_name[c.strip().lower()] for c in candidates
                       if c.strip().lower() in by_name), None)
        if ws_src is None:
            print(f"  [Old AC] no source tab {candidates} for '{report_title}' "
                  f"in OLD_REFERENCE.xlsx - skipped.", file=sys.stderr)
            continue
        pairs.append((ws_src, report_title))
    return src, pairs


def write_excel(rows, pending_rows, trade_rows, account_id, account_data,
                order_lookup=None, live_positions=None, flex_positions=None,
                completed_orders=None, fx_rates=None):
    today     = dt.date.today()
    yesterday = today - dt.timedelta(days=1)
    cutoff_7  = today - dt.timedelta(days=7)

    # ── No date cutoff: include every trade IBKR returns (all BUY/SELL fills) ──
    print(f"  [Filter] {len(rows)} trade(s) pulled from IBKR (no date cutoff).")

    # Today's set mirrors the Today Orders sheet: today's fills, or yesterday's
    # as a fallback when IBKR's Flex batch hasn't published today's yet.
    rows_today, rows_7 = [], []
    for row in rows:
        d = parse_trade_date(row)
        if d is None:
            continue
        if d in (today, yesterday):
            rows_today.append(row)
        if d >= cutoff_7:
            rows_7.append(row)

    agg_today = aggregate(rows_today)
    agg_7     = aggregate(rows_7)
    agg_all   = aggregate(rows)             # all trades returned by IBKR

    # Carry forward what the user maintains by hand in the previous report:
    # trigger/limit prices (All Trades) and every manual sheet's rows.
    manual_prices, carried = load_previous_report()

    # All Trades: every individual trade IBKR returns (ledger trigger/limit,
    # with manually-entered prices from the previous report taking precedence).
    all_trade_rows = build_individual_trade_rows(rows, order_lookup=order_lookup,
                                                 manual_prices=manual_prices)
    print(f"  [All Trades] {len(all_trade_rows)} individual trade(s).")

    # Static "old account" reference sheets (from OLD_REFERENCE.xlsx). Loaded up
    # front so the Index can list whichever ones are actually available, and so
    # each can be slotted into its fixed position in the sheet sequence below.
    ref_wb, ref_pairs = _load_reference_sheet_pairs()
    ref_by_title = {title: ws for ws, title in ref_pairs}

    def _copy_ref(title):
        """Copy the named old-AC reference sheet into wb if it was available."""
        src = ref_by_title.get(title)
        if src is not None:
            _copy_reference_sheet(wb, src, title)

    # Final sheet sequence — all four old-AC reference sheets are grouped first
    # (right after the Index), landing at fixed positions: Dashboard old AC (2),
    # Open Position old AC (3), Trade Summary old AC (4), All Trades old AC (5).
    # The live sheets follow, then the manual sheets. The Index lists the sheets
    # in this exact order, so its SI numbers match the tab order.
    index_order = [
        ("Dashboard old AC",     "ref"),
        ("Open Position old AC", "ref"),
        ("Trade Summary old AC", "ref"),
        ("All Trades old AC",    "ref"),
        ("Dashboard",            "live"),
        ("Pending Order",        "live"),
        ("Trades_Status",        "live"),
        ("Open Position",        "live"),
        ("All Trades",           "live"),
        ("Trade Summary",        "live"),
        ("Strategy Details",     "live"),
        ("Bugs",                 "live"),
        ("Shubham_Activity",     "live"),
        ("Ajay_Activity",        "live"),
        ("Pending_Task",         "live"),
    ]
    # Only list reference sheets that actually loaded (e.g. OLD_REFERENCE.xlsx
    # missing/locked) so the Index never shows a dead link.
    index_titles = [t for t, kind in index_order
                    if kind == "live" or t in ref_by_title]

    wb = Workbook()
    _fill_index_sheet(wb.active, index_titles, account_id)

    # Old-AC reference sheets first (grouped right after the Index).
    _copy_ref("Dashboard old AC")
    _copy_ref("Open Position old AC")
    _copy_ref("Trade Summary old AC")
    _copy_ref("All Trades old AC")

    # Live sheets, then the manual sheets.
    _fill_dashboard_sheet(wb.create_sheet(), account_id, account_data, pending_rows,
                          agg_7, agg_today, agg_all, len(all_trade_rows),
                          live_positions=live_positions)
    _fill_pending_sheet(wb.create_sheet(),            pending_rows)
    trades_status_rows = build_trades_status_rows(pending_rows, completed_orders,
                                                  all_trade_rows, fx_rates)
    _fill_trades_status_sheet(wb.create_sheet(),      trades_status_rows)
    print(f"  [Trades Status] {len(trades_status_rows)} row(s) "
          f"(pending + cancelled + filled).")
    _fill_running_positions_sheet(wb.create_sheet(),  rows, live_positions, flex_positions,
                                  fx_rates)
    _fill_trade_list_sheet(wb.create_sheet(), "All Trades", all_trade_rows)
    _fill_sheet(wb.create_sheet(), "Trade Summary",   agg_all)
    _fill_strategy_sheet(wb.create_sheet(), carried.get("Strategy Details"))
    _fill_bugs_sheet(wb.create_sheet(), carried.get("Bugs"))
    _fill_activity_sheet(wb.create_sheet(), "Shubham_Activity", carried.get("Shubham_Activity"))
    _fill_activity_sheet(wb.create_sheet(), "Ajay_Activity", carried.get("Ajay_Activity"))
    _fill_pending_task_sheet(wb.create_sheet(), carried.get("Pending_Task"))

    if ref_wb is not None:
        ref_wb.close()
        n_ref = sum(1 for t, kind in index_order if kind == "ref" and t in ref_by_title)
        print(f"  [Old AC] copied {n_ref} reference sheet(s) at positions 2/3/4/5.")

    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}: {len(pending_rows)} pending, "
          f"{len(all_trade_rows)} all trades, "
          f"{len(agg_today)} contracts (today), "
          f"{len(agg_7)} contracts (7d), {len(agg_all)} contracts (summary)")


# ── Index sheet (links to every other sheet) ──────────────────────────
def _ordinal(n):
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _fill_index_sheet(ws, sheet_titles, account_id="N/A"):
    ws.title = "Index"

    title_fill = PatternFill("solid", fgColor="DDEBF7")   # light blue
    head_fill  = PatternFill("solid", fgColor="DDD9C4")   # beige
    data_fill  = PatternFill("solid", fgColor="DDEBF7")   # light blue
    thin       = Side(style="thin", color="1F3864")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font  = Font(name="Calibri", size=11, bold=True, color="0563C1", underline="single")
    head_font  = Font(name="Calibri", size=11, bold=True, color="1F3864")
    title_font = Font(name="Calibri", size=12, bold=True, color="1F3864")
    label_font = Font(name="Calibri", size=11, bold=True, color="1F3864")
    value_font = Font(name="Calibri", size=11, color="1F3864")

    today = dt.date.today()
    report_title = f"IBKR - MIS Report as on {_ordinal(today.day)} {today:%B %Y}"

    # Row 3: report title (merged B:C)
    ws.merge_cells("B3:C3")
    t = ws["B3"]
    t.value, t.font, t.fill = report_title, title_font, title_fill
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Row 4: "Index" (merged B:C)
    ws.merge_cells("B4:C4")
    ix = ws["B4"]
    ix.value, ix.font, ix.fill = "Index", head_font, title_fill
    ix.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5: column headers
    for col, label in ((2, "SI"), (3, "Description")):
        c = ws.cell(row=5, column=col, value=label)
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Rows 6+: one linked row per sheet
    for i, title in enumerate(sheet_titles, start=1):
        r = 5 + i
        sl = ws.cell(row=r, column=2, value=i)
        sl.fill = data_fill
        sl.alignment = Alignment(horizontal="center", vertical="center")
        sl.font = Font(name="Calibri", size=11, color="1F3864")

        link = ws.cell(row=r, column=3, value=title)
        link.hyperlink  = f"#'{title}'!A1"
        link.font       = link_font
        link.fill       = data_fill
        link.alignment  = Alignment(horizontal="center", vertical="center")

    # Borders around the whole table (rows 3..last, cols B:C)
    last_row = 5 + len(sheet_titles)
    for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=2, max_col=3):
        for cell in row:
            cell.border = border

    # Green frame around the report title banner (B3:C3), matching other sheets.
    ws["B3"].border = _TITLE_BORDER
    ws["C3"].border = _TITLE_BORDER

    # ── Report Details box (cols E:F), alongside the index table ──────
    details = [
        ("Account ID",         account_id),
        ("Report Date",        today.strftime("%d-%b-%Y")),
    ]
    ws.merge_cells("E4:F4")
    dt_title = ws["E4"]
    dt_title.value, dt_title.font, dt_title.fill = "Report Details", head_font, title_fill
    dt_title.alignment = Alignment(horizontal="center", vertical="center")
    for i, (label, value) in enumerate(details, start=1):
        r = 4 + i
        lc = ws.cell(row=r, column=5, value=label)
        lc.font, lc.fill = label_font, data_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=6, value=value)
        vc.font, vc.fill = value_font, data_fill
        vc.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=4, max_row=4 + len(details), min_col=5, max_col=6):
        for cell in row:
            cell.border = border

    # ── Login Details box (cols E:F), below the Report Details box ─────
    # Report Details fills rows 4..(4+len(details)); leave one blank row, then
    # start the credentials box beneath it.
    cred_title_row = 4 + len(details) + 2          # one-row gap after Report Details
    credentials = [
        ("Username", "algo-ggt11"),
        ("Password", "algoggt12))"),
    ]
    ws.merge_cells(start_row=cred_title_row, start_column=5,
                   end_row=cred_title_row, end_column=6)
    ct = ws.cell(row=cred_title_row, column=5)
    ct.value, ct.font, ct.fill = "Login Details", head_font, title_fill
    ct.alignment = Alignment(horizontal="center", vertical="center")
    for i, (label, value) in enumerate(credentials, start=1):
        r = cred_title_row + i
        lc = ws.cell(row=r, column=5, value=label)
        lc.font, lc.fill = label_font, data_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=6, value=value)
        vc.font, vc.fill = value_font, data_fill
        vc.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=cred_title_row,
                            max_row=cred_title_row + len(credentials),
                            min_col=5, max_col=6):
        for cell in row:
            cell.border = border

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.sheet_view.showGridLines = False


# ── Dashboard sheet ────────────────────────────────────────────────────
def _live_unrealized_total(live_positions):
    """Total unrealized PnL, sourced solely from IBKR's live portfolio feed —
    the only authoritative source (the Flex trade history has none). Sums
    unrealizedPNL over the actually-open live positions and returns 0.0 when
    TWS reported the account flat, or when it was unavailable (live_positions
    is None). Matches what the Open Position sheet shows per contract."""
    total = 0.0
    for p in (live_positions or []):
        if abs(p.get("position") or 0) < 1e-9:      # flat → nothing unrealized
            continue
        upnl = p.get("unrealizedPNL")
        if _ibkr_has(upnl):
            total += float(upnl)
    return round(total, 2)


def _fill_dashboard_sheet(ws, account_id, account_data, pending_rows,
                          agg_7, agg_today, agg_all, num_all_trades,
                          live_positions=None):
    ws.title  = "Dashboard"
    now_ist   = dt.datetime.now(_IST)
    n_cols    = 2   # Parameter | Value

    _write_title_row(ws, _SHEET_TITLES["Dashboard"], n_cols)
    _write_header_row(ws, ["Parameter", "Value"], row=2)

    # ── Report info ───────────────────────────────────────────────────
    info = [
        ("Account ID",         account_id,                     ""),
        ("Report Date",        now_ist.strftime("%d-%b-%Y"),    ""),
    ]

    # ── All-trades summary (everything IBKR returns) ──────────────────
    at_buys   = sum(r["Total (bought)"]       for r in agg_all)
    at_sells  = sum(r["Total (sold)"]         for r in agg_all)
    at_comm   = sum(r["Commission"]           for r in agg_all)
    at_pnl    = sum(r["PnL"]                  for r in agg_all)
    # Unrealized PnL comes straight from IBKR's live feed, not the trade history.
    at_unreal = _live_unrealized_total(live_positions)

    summary_all = [
        ("Total Trades (fills)", num_all_trades,            ""),
        ("Contracts Traded",     len(agg_all),              ""),
        ("Total Bought Value",   f"{at_buys:,.2f}",          ""),
        ("Total Sold Value",     f"{at_sells:,.2f}",         ""),
        ("Commission",           f"{at_comm:,.2f}",          ""),
        ("Realized PnL",         f"{at_pnl:,.2f}",           ""),
        ("Unrealized PnL",       f"{at_unreal:,.2f}",        ""),
        ("Max Exposure",         f"{MAX_EXPOSURE:,.2f}",     ""),
        ("Daily Max Loss",       f"{DAILY_MAX_LOSS:,.2f}",   ""),
        ("Loss due to bugs",     f"{LOSS_DUE_TO_BUGS:,.2f}", ""),
    ]

    # Dashboard value cells shaded green (positive) / red (negative).
    # Commission is intentionally excluded — it is shown with no colour.
    color_labels = {"Loss due to bugs", "Unrealized PnL", "Realized PnL"}

    # ── Assemble all sections with dividers ───────────────────────────
    sections = [
        ("REPORT INFO",                          info),
        ("ALL TRADES SUMMARY",                   summary_all),
    ]

    r_idx = 3
    for section_title, section_rows in sections:
        # Blank gap before each section (except first)
        if r_idx > 3:
            ws.row_dimensions[r_idx].height = 6
            r_idx += 1

        # Section sub-heading row
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r_idx, column=col,
                           value=section_title if col == 1 else "")
            cell.font      = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            cell.fill      = _NAVY
            cell.alignment = _LEFT
        ws.row_dimensions[r_idx].height = 18
        r_idx += 1

        # Data rows (Currency column dropped — values are USD as shown in TWS)
        for label, value, _currency in section_rows:
            ca = ws.cell(row=r_idx, column=1, value=label)
            cb = ws.cell(row=r_idx, column=2, value=value)
            ca.font = _LABEL_FONT
            cb.font = _DATA_FONT
            cb.alignment = _RIGHT
            if label in color_labels:
                _color_pnl_cell(cb)
            ws.row_dimensions[r_idx].height = 16
            r_idx += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A3"


# ── Per-trade sheets (All Trades) ─────────────────────────────────────
def _fill_trade_list_sheet(ws, sheet_key, rows):
    ws.title  = sheet_key
    n_cols    = len(TODAY_HEADERS)

    _write_title_row(ws, _SHEET_TITLES[sheet_key], n_cols)
    _write_header_row(ws, TODAY_HEADERS, row=2)

    for row in rows:
        ws.append(row)

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    # Highlight today's trades in yellow (Date & Time is column 1).
    today_str = dt.date.today().strftime("%d-%b-%Y")
    yellow    = PatternFill("solid", fgColor="FFFF00")
    for r in range(3, ws.max_row + 1):
        dt_val = ws.cell(row=r, column=1).value
        if dt_val and str(dt_val).startswith(today_str):
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = yellow

    # Amount columns: uniform 1,234.00 number format.
    for h in ("Commission", "Realized PnL"):
        _apply_amount_format(ws, TODAY_HEADERS.index(h) + 1)

    # Shade the Commission / Realized PnL cells green (positive) / red (negative).
    # Applied after the yellow pass so the colour wins on those cells.
    for h in COLOR_COLUMNS_TRADES:
        col = TODAY_HEADERS.index(h) + 1
        for r in range(3, ws.max_row + 1):
            _color_pnl_cell(ws.cell(row=r, column=col))

    for col, width in enumerate(TODAY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _apply_cell_borders(ws, n_cols, start_row=2)
    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


# ── Pending Order sheet ────────────────────────────────────────────────
def _fill_pending_sheet(ws, rows):
    ws.title  = "Pending Order"
    n_cols    = len(PENDING_HEADERS)

    _write_title_row(ws, _SHEET_TITLES["Pending Order"], n_cols)
    _write_header_row(ws, PENDING_HEADERS, row=2)

    for row in rows:
        ws.append(row)

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    for col, width in enumerate(PENDING_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


# ── Trades Status sheet ────────────────────────────────────────────────
def _fill_trades_status_sheet(ws, rows):
    ws.title = "Trades_Status"
    n_cols   = len(TRADES_STATUS_HEADERS)

    _write_title_row(ws, _SHEET_TITLES["Trades_Status"], n_cols)
    _write_header_row(ws, TRADES_STATUS_HEADERS, row=2)

    for row in rows:
        ws.append(row)

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    for col, width in enumerate(TRADES_STATUS_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


# ── Open Position sheet ───────────────────────────────────────────────
def _is_flat_residual(net_qty, gross_traded=0.0):
    """True when a net position is effectively flat and must NOT be shown as an
    open position. Besides an exact zero, an FX round-trip can leave a negligible
    fractional residual — e.g. -0.75 units left on a 43,750-unit EUR.USD
    round-trip — that IBKR/Flex still report as a standing position. Such 'dust'
    (a net below 0.01% of the volume traded in the contract) is treated as
    closed. When there is no traded volume to compare against (position opened
    outside the Flex window), only an exact zero counts as flat."""
    q = abs(net_qty or 0)
    if q < 1e-9:
        return True
    return gross_traded > 0 and q / gross_traded < 1e-4


def _build_live_open_positions(live_positions, trade_rows):
    """Open Position rows in the standard HEADERS format, driven by the LIVE
    IBKR positions (reqPositions) so the sheet shows the broker's actual
    holdings — not a netting of the (possibly incomplete) Flex trade history.

    The Net column always carries IBKR's live signed quantity, and the Unrealized
    PnL / Realized PnL / position value come straight from IBKR's live portfolio
    feed (updatePortfolio) when available. The remaining descriptive columns
    (trade dates, buys/sells, bought-vs-sold split, commission) are merged from
    the Flex trade aggregation when the contract appears there, since IBKR reports
    only a single net figure per position and cannot split them; for a position
    opened outside the Flex window those columns are left blank and the contract's
    exchange falls back to the one IBKR reports on the position.

    Matching the live position to the trade aggregation is the subtle part: for
    FX, reqPositions reports contract.symbol as the bare base currency (e.g.
    "USD") while the Flex trade history keys the same position by the PAIR
    ("USD.JPY", which equals the position's localSymbol). So we try several
    candidate identifiers — localSymbol, the constructed base.quote pair, and
    the bare symbol — and use the first that exists in the aggregation.
    """
    agg_by_symbol = {r["Contract"]: r for r in aggregate(trade_rows)}
    rows = []
    for p in live_positions:
        qty = p.get("position") or 0
        if abs(qty) < 1e-9:                 # flat → not an open position
            continue
        symbol = p.get("symbol") or ""
        name   = p.get("name") or ""        # localSymbol, e.g. "USD.JPY" for FX
        # For FX (CASH) the tradeable instrument is the base.quote pair, which is
        # how it appears in the Flex history — build it so it can be matched.
        pair = (f"{symbol}.{p.get('currency')}"
                if p.get("secType") == "CASH" and p.get("currency") else "")
        # Try the candidate identifiers against the trade aggregation, most
        # specific first, so the descriptive columns merge in for FX and stocks.
        agg_row, match_key = {}, symbol
        for cand in (name, pair, symbol):
            if cand and cand in agg_by_symbol:
                agg_row, match_key = agg_by_symbol[cand], cand
                break
        base = dict(agg_row)
        base["Contract"] = match_key or symbol
        if not base.get("Name"):
            base["Name"] = name or pair or symbol
        base["Net"] = round(qty, 4)         # live broker quantity wins
        # A negligible fractional leftover from a round-trip (FX "dust") is not a
        # real open position — drop it so a closed contract stops showing here.
        gross = (base.get("Buys") or 0) + (base.get("Sells") or 0)
        if _is_flat_residual(qty, gross):
            continue
        if not base.get("Exchange List"):
            base["Exchange List"] = p.get("exchange", "")

        # ── IBKR live portfolio figures (authoritative per-position) ──────────
        # Unrealized PnL has no per-fill equivalent in the Flex trade history, so
        # IBKR is the source of truth here. Realized PnL likewise comes from IBKR
        # when it reports one, otherwise the FIFO figure from the trade history is
        # kept. Avg cost / position value are filled from IBKR only when the Flex
        # aggregation has no value (e.g. a position opened before the Flex window),
        # because IBKR reports a single net figure and can't split bought vs sold.
        upnl = p.get("unrealizedPNL")
        if _ibkr_has(upnl):
            base["Unrealized PnL"] = round(float(upnl), 2)

        rpnl = p.get("realizedPNL")
        if _ibkr_has(rpnl) and not base.get("PnL"):
            base["PnL"] = round(float(rpnl), 2)

        avg = p.get("averageCost")
        if _ibkr_has(avg) and avg:
            avg = float(avg)
            mkt_val = p.get("marketValue")
            total   = (round(float(mkt_val), 2) if _ibkr_has(mkt_val)
                       else round(avg * abs(qty), 2))
            if qty > 0 and not base.get("Avg (bought)"):
                base["Avg (bought)"] = round(avg, 6)
                if not base.get("Total (bought)"):
                    base["Total (bought)"] = total
            elif qty < 0 and not base.get("Avg (sold)"):
                base["Avg (sold)"] = round(avg, 6)
                if not base.get("Total (sold)"):
                    base["Total (sold)"] = total
        rows.append(base)
    rows.sort(key=lambda x: x["Contract"])
    return rows


def _fill_running_positions_sheet(ws, trade_rows, live_positions=None,
                                  flex_positions=None, fx_rates=None):
    """Open Position sheet, sourced in priority order:

      1. Live IBKR positions (reqPositions) — real-time broker truth.
      2. Flex <OpenPosition> snapshot — authoritative when TWS is unreachable OR
         returns nothing (a slow/empty live feed no longer blanks the sheet).
      3. Netting the Flex trade history — last resort; only correct when the
         statement's date range covers each position's opening trades.

    fx_rates converts the Flex-snapshot figures to USD (live positions are already
    normalized upstream) so the sheet never mixes currencies.
    """
    flex_open = [p for p in (_flex_open_to_position(op, fx_rates) for op in (flex_positions or []))
                 if abs(p.get("position") or 0) > 1e-9]

    if live_positions:                              # connected AND holding positions
        running = _build_live_open_positions(live_positions, trade_rows)
        src = "live IBKR positions"
    elif flex_open:                                 # TWS down or returned none
        running = _build_live_open_positions(flex_open, trade_rows)
        why = "TWS unavailable" if live_positions is None else "TWS returned no positions"
        src = f"Flex open positions ({why})"
    else:
        agg     = aggregate(trade_rows)
        running = [r for r in agg
                   if not _is_flat_residual(r["Net"],
                                            (r.get("Buys") or 0) + (r.get("Sells") or 0))]
        src = "trade-netted (no live or Flex positions)"
    print(f"  [Open Position] {len(running)} open position(s) from {src}.")
    _fill_sheet(ws, "Open Position", running)


# ── Generic trade aggregation sheet ──────────────────────────────────
def _fill_sheet(ws, title, agg_rows):
    ws.title = title
    n_cols   = len(HEADERS)

    _write_title_row(ws, _SHEET_TITLES.get(title, title.upper()), n_cols)
    _write_header_row(ws, HEADERS, row=2)

    for row in agg_rows:
        ws.append([row.get(h, "") for h in HEADERS])

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    # Amount columns: uniform 1,234.00 number format.
    for h in AMOUNT_COLUMNS_SUMMARY:
        if h in HEADERS:
            _apply_amount_format(ws, HEADERS.index(h) + 1)

    # Average price columns: 5-decimal format, e.g. 1.14205.
    for h in PRICE_COLUMNS_SUMMARY:
        if h in HEADERS:
            _apply_amount_format(ws, HEADERS.index(h) + 1, fmt=_PRICE_FMT)

    # Shade money columns green (positive) / faint red (negative).
    for h in COLOR_COLUMNS_SUMMARY:
        col = HEADERS.index(h) + 1
        for r in range(3, ws.max_row + 1):
            _color_pnl_cell(ws.cell(row=r, column=col))

    # Auto-fit column widths based on header + data content
    for i, h in enumerate(HEADERS, start=1):
        values = [str(r.get(h, "")) for r in agg_rows]
        width  = max(len(h), *(len(v) for v in values)) if values else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 3, 13), 32)

    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


# ── Manual-sheet helpers ──────────────────────────────────────────────
def _col_by_header(ws, *names, header_row=2):
    """Column letter of the first header (row 2) matching any of `names`
    (case-insensitive). Returns None if absent. Lets dropdowns target a column
    by name, so they stay on the right column even if the user reorders or
    inserts columns in a carried-forward sheet."""
    wanted = {n.strip().lower() for n in names}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip().lower() in wanted:
            return get_column_letter(c)
    return None


def _drop_carried_column(carried, *col_names):
    """Remove the named column(s) from a carried (headers, rows, layout) triple.

    Used to retire a column from a manual sheet that is otherwise reproduced
    verbatim from the previous report — without this, the old column would keep
    coming back every run via the carry-forward. Matching is case-insensitive
    and ignores surrounding whitespace. Carried column widths are re-indexed to
    the surviving columns. Returns a new triple (or the original `carried`
    unchanged when there is nothing to drop)."""
    if not carried:
        return carried
    headers, rows, layout = carried
    drop = {n.strip().lower() for n in col_names}
    keep = [i for i, h in enumerate(headers)
            if str(h).strip().lower() not in drop]
    if len(keep) == len(headers):
        return carried                      # nothing matched — leave as-is
    new_headers = [headers[i] for i in keep]
    new_rows    = [[row[i] for i in keep if i < len(row)] for row in rows]
    # Re-index carried column widths (1-based) onto the surviving columns.
    old_widths  = (layout or {}).get("col_widths", {})
    new_widths  = {new_idx: old_widths[old_i + 1]
                   for new_idx, old_i in enumerate(keep, start=1)
                   if (old_i + 1) in old_widths}
    new_layout  = dict(layout or {})
    new_layout["col_widths"] = new_widths
    return new_headers, new_rows, new_layout


def _add_list_dropdown(ws, col_letter, choices):
    """Attach a list-validation dropdown to col_letter, rows 3..(max or 1000)."""
    if not col_letter:
        return
    dv = DataValidation(type="list", formula1=f'"{choices}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}3:{col_letter}{max(ws.max_row, 1000)}")


# ── Manual-sheet helper: verbatim copy (headers + data, no blank rows) ──
def _write_verbatim_manual_sheet(ws, sheet_key, default_headers, widths, carried):
    """Reproduce a manual sheet exactly as it was left in the previous report:
    the carried header row and every carried data row, with NO blank entry rows.
    New entries are typed straight into the empty rows below the data.

    `carried` is the (headers, rows, layout) triple from _read_manual_sheet. On
    the very first run (nothing carried) we fall back to the built-in
    default_headers and the built-in `widths`. Returns the column count used."""
    headers, rows, layout = carried if carried else (None, None, None)
    headers = list(headers) if headers else list(default_headers)
    layout  = layout or {}
    n_cols  = len(headers)

    ws.title = sheet_key
    _write_title_row(ws, _SHEET_TITLES[sheet_key], n_cols)
    _write_header_row(ws, headers, row=2)

    # Each carried cell is a (value, style) pair, where `style` is either the
    # full style dict captured by _capture_cell_style (number format + borders +
    # fill/highlight + font + alignment) or, for legacy callers, just a number
    # format string. Reapply whatever is present so the cell looks exactly as it
    # did in the prior report — manual highlights included.
    row_heights = layout.get("row_heights", [])
    r = 3
    for ridx, row in enumerate(rows or []):
        for c in range(1, n_cols + 1):
            value, style = row[c - 1] if c - 1 < len(row) else ("", None)
            cell = ws.cell(row=r, column=c, value=value)
            if isinstance(style, dict):
                # The captured style objects are already detached copies, each
                # used for a single cell — assign directly, no re-copy needed.
                if style.get("number_format"):
                    cell.number_format = style["number_format"]
                if style.get("font"):      cell.font      = style["font"]
                if style.get("fill"):      cell.fill      = style["fill"]
                if style.get("border"):    cell.border    = style["border"]
                if style.get("alignment"): cell.alignment = style["alignment"]
            elif style and style != "General":
                cell.number_format = style
        # Carry the row's height forward when the user set a custom one.
        if ridx < len(row_heights) and row_heights[ridx] is not None:
            ws.row_dimensions[r].height = row_heights[ridx]
        r += 1

    # First run (no carried styles): fall back to the uniform data styling.
    if not rows:
        _style_data_rows(ws, start_row=3, n_cols=n_cols)

    # Column widths: use the widths carried from the previous report (preserving
    # any the user adjusted, with ranges already expanded), falling back to the
    # built-in default for any column the previous report didn't size.
    carried_widths = layout.get("col_widths", {})
    for col in range(1, n_cols + 1):
        w = carried_widths.get(col)
        if w is None and col - 1 < len(widths):
            w = widths[col - 1]
        if w is not None:
            ws.column_dimensions[get_column_letter(col)].width = w

    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"
    return n_cols


# ── Strategy Details sheet (manual entry) ─────────────────────────────
def _fill_strategy_sheet(ws, carried=None):
    """SR NO / STRATEGY / DETAILS / STATUS reproduced verbatim from the previous
    report — the header row and every data row exactly as left, with no blank
    rows. The STATUS column keeps an Active/Inactive dropdown."""
    _write_verbatim_manual_sheet(ws, "Strategy Details", STRATEGY_HEADERS,
                                 STRATEGY_WIDTHS, carried)
    _add_list_dropdown(ws, _col_by_header(ws, "STATUS"), "Active,Inactive")


# ── Activity sheets (manual per-person task log) ──────────────────────
def _fill_activity_sheet(ws, sheet_key, carried=None):
    """SR NO / TASK / NO. OF HOURS reproduced verbatim from the previous report
    — the header row and every data row exactly as left, with no blank rows."""
    _write_verbatim_manual_sheet(ws, sheet_key, ACTIVITY_HEADERS,
                                 ACTIVITY_WIDTHS, carried)


# ── Pending Task sheet (manual priority/task tracker) ─────────────────
def _fill_pending_task_sheet(ws, carried=None):
    """SR NO / Task / COMPLETION DATE / Severity / ETA reproduced verbatim from
    the previous report (no blank rows). The Severity column keeps a
    Very High/High/Medium/Low dropdown."""
    _write_verbatim_manual_sheet(ws, "Pending_Task", PENDING_TASK_HEADERS,
                                 PENDING_TASK_WIDTHS, carried)
    _add_list_dropdown(ws, _col_by_header(ws, "Severity"),
                       PENDING_TASK_SEVERITY_CHOICES)


# ── Bugs sheet (manual daily entry) ───────────────────────────────────
def _fill_bugs_sheet(ws, carried=None):
    """Bug log reproduced verbatim from the previous report — the header row and
    every bug row exactly as left, with no blank rows.

    The Severity column keeps a Critical/High/Medium/Low dropdown and the
    Current_Status column a Fixed/Not Fixed dropdown (located by header name so
    they stay correct even if columns shift).

    The legacy "Resolution" column is dropped from any carried-forward data so
    it doesn't keep reappearing from older reports.
    """
    carried = _drop_carried_column(carried, "Resolution")
    _write_verbatim_manual_sheet(ws, "Bugs", BUGS_HEADERS, BUGS_WIDTHS, carried)

    # Bug details column reads better left-aligned than centred.
    details_col = _col_by_header(ws, "BUG_DETAILS")
    if details_col:
        ci = column_index_from_string(details_col)
        for r in range(3, ws.max_row + 1):
            ws.cell(row=r, column=ci).alignment = _LEFT

    _add_list_dropdown(ws, _col_by_header(ws, "Severity"), BUGS_SEVERITY_CHOICES)
    _add_list_dropdown(ws, _col_by_header(ws, "Current_Status"), BUGS_STATUS_CHOICES)


def main():
    print("Connecting to TWS - fetching open/completed orders, executions, positions, and account summary...")
    (orders_by_symbol, account_data, executions,
     completed_orders, live_positions) = fetch_tws_data()
    trade_rows   = build_trade_rows(executions)
    print(f"  {len(orders_by_symbol)} pending symbol group(s) | "
          f"{len(trade_rows)} today's executions | "
          f"{len(completed_orders)} completed orders | "
          f"{0 if live_positions is None else len(live_positions)} live positions | "
          f"{len(account_data)} account tags collected.")

    # Persist trigger/limit from both open AND completed orders, then build the
    # lookup used to back-fill those prices onto matching trades.
    ledger = load_order_ledger()
    update_order_ledger(ledger, orders_by_symbol, completed_orders)
    save_order_ledger(ledger)
    order_lookup = build_order_lookup(ledger)
    print(f"  [Order Ledger] {len(ledger)} order(s) tracked for trigger/limit history "
          f"({LEDGER_FILE}).")

    # Persist EVERY live TWS fill across runs so no day's orders are ever lost —
    # today's fills are saved now and every prior day's remain available even
    # when reqExecutions (current-day only) or the Flex batch no longer has them.
    trades_ledger = load_trades_ledger()
    update_trades_ledger(trades_ledger, executions)
    save_trades_ledger(trades_ledger)
    print(f"  [Trades Ledger] {len(trades_ledger)} lifetime fill(s) retained "
          f"({TRADES_LEDGER_FILE}).")

    ref        = send_request()
    xml_text   = get_statement(ref)
    rows       = parse_trades(xml_text)
    # Authoritative open positions from the Flex statement — parsed up front so
    # their fxRateToBase feeds the FX-rate map below (and used later for the Open
    # Position sheet when the live TWS feed is unavailable or returns nothing).
    open_positions = parse_open_positions(xml_text)
    # Harvest FX rates from the Flex trades and open positions (which carry
    # fxRateToBase) BEFORE normalization overwrites their currency, so live fills,
    # pending orders and positions can all be converted to USD using the same
    # rates. Falls back to the rate implied by a USD-based pair's own price.
    fx_rates = build_fx_rate_map(rows, open_positions)
    # Every live position's money fields to USD, so the Open Position sheet and
    # the Dashboard's unrealized total never show a non-USD (JPY) figure.
    normalize_positions_to_usd(live_positions, fx_rates)
    # Build pending-order rows now that FX rates are known, so their Total Amount
    # (and the Trades_Status sheet's) reads in USD for FX pairs like USD.JPY.
    pending_rows = build_pending_rows(orders_by_symbol, fx_rates)
    print(f"  {len(pending_rows)} pending order rows.")
    # Splice in live TWS fills — this session's plus every prior day's from the
    # trades ledger. IBKR's Flex batch lags ~a day and reqExecutions only returns
    # the current day, so this is what guarantees every day's orders are present.
    merge_todays_executions(rows, trades_ledger, fx_rates)
    # Express every trade in the base currency (USD) so the report never mixes
    # currencies, even when the account holds FX pairs or non-USD instruments.
    normalize_trades_to_usd(rows, fx_rates)
    # Derive per-fill realized PnL (USD) by FIFO so All Trades shows a real figure
    # on each squared-off fill that sums to the Trade Summary PnL.
    attach_fifo_realized(rows)
    account_id = parse_account_id(xml_text)
    print(f"Account: {account_id} | Parsed {len(rows)} trade records "
          f"(monetary values normalized to {BASE_CURRENCY}) | "
          f"{len(open_positions)} Flex open position(s).")
    if not rows:
        print("No trades found. Check the query's date range and section config.")
    write_excel(rows, pending_rows, trade_rows, account_id, account_data,
                order_lookup, live_positions, open_positions, completed_orders,
                fx_rates)


if __name__ == "__main__":
    main()
