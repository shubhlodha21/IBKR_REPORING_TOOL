"""
Merge a live MIS report and its matching paper MIS report into one workbook.

For a given date the Reports folder contains two files:
    MIS_<date>.xlsx          -> live account report  (Index + 10 sheets)
    MIS_paper_<date>.xlsx    -> paper account report (paper_Index + 5 sheets)

This script produces:
    MIS_merged_<date>.xlsx

Merge rules
-----------
* The merged Index is the LIVE Index, with the paper report's description rows
  appended below it, renumbered to continue the live list (11, 12, ...).
* The paper account's "Report Details" and "Login Details (Paper Account)"
  boxes are copied below the live boxes so paper credentials are preserved.
* Sheet order: Index, then every live data sheet, then every paper data sheet.
  The standalone paper_Index sheet is dropped (its list is merged into Index).

Usage
-----
    python merge_mis.py                # merge every paired date found in Reports/
    python merge_mis.py 10Jun2026      # merge just that date
"""

import copy
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

REPORTS_DIR = Path(__file__).resolve().parent / "Reports"

# --- Index layout constants (1-based rows/cols, matching the source files) ---
SI_COL = "B"             # column holding the serial number
DESC_COL = "C"           # column holding the description
LIVE_LAST_ITEM_ROW = 15  # row of live item #10 (Pending_Task)
LIVE_FIRST_ITEM_ROW = 6  # row of live item #1 (Dashboard)
PAPER_FIRST_ITEM_ROW = 6 # row of paper item #1 in paper_Index
PAPER_LAST_ITEM_ROW = 10 # row of paper item #5 in paper_Index
BOX_GAP = 1              # blank rows between the live and paper detail boxes


def copy_cell_style(src, dst):
    """Copy value + full formatting from one cell to another (cross-workbook safe)."""
    dst.value = src.value
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = copy.copy(src.number_format)
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def copy_sheet(src_ws, dst_wb, title):
    """Deep-copy a worksheet (cells, styles, merges, dimensions) into another workbook."""
    dst = dst_wb.create_sheet(title=title)

    for row in src_ws.iter_rows():
        for cell in row:
            copy_cell_style(cell, dst[cell.coordinate])

    for merged in src_ws.merged_cells.ranges:
        dst.merge_cells(str(merged))

    for col, dim in src_ws.column_dimensions.items():
        d = dst.column_dimensions[col]
        d.width = dim.width
        d.hidden = dim.hidden
        d.outlineLevel = dim.outlineLevel

    for idx, dim in src_ws.row_dimensions.items():
        d = dst.row_dimensions[idx]
        d.height = dim.height
        d.hidden = dim.hidden
        d.outlineLevel = dim.outlineLevel

    dst.freeze_panes = src_ws.freeze_panes
    dst.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
    dst.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    dst.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth
    return dst


def paper_label(paper_path):
    """Distinguishing token of a paper file name (empty for the default account).

    MIS_paper_<date>.xlsx      -> ""    (default / first paper account)
    MIS_paper2_<date>.xlsx     -> "2"
    MIS_paper_AJAY_<date>.xlsx -> "_AJAY"
    """
    rest = paper_path.stem[len("MIS_paper"):]  # e.g. "2_16Jun2026" / "_16Jun2026"
    return rest.rsplit("_", 1)[0]


def paper_sheet_prefix(paper_path):
    """Sheet-name prefix used to keep one paper account's sheets distinct.

    The default account keeps the original "paper_" prefix; extra accounts get
    "paper<label>_" so their identically-named sheets don't collide on merge.
    """
    return f"paper{paper_label(paper_path)}_"


def rename_paper_title(title, prefix):
    """Re-prefix a "paper_<name>" sheet/description title for a specific account."""
    if title.startswith("paper_") and prefix != "paper_":
        return prefix + title[len("paper_"):]
    return title


def append_paper_index_rows(index_ws, paper_index_ws, prefix, out_row, next_number):
    """Append one paper account's item rows below the current Index list.

    Continues the serial numbering from ``next_number`` at row ``out_row`` and
    returns the (next free row, next serial number) after the appended rows.
    """
    si_template = index_ws[f"{SI_COL}{LIVE_LAST_ITEM_ROW}"]
    desc_template = index_ws[f"{DESC_COL}{LIVE_LAST_ITEM_ROW}"]

    for src_row in range(PAPER_FIRST_ITEM_ROW, PAPER_LAST_ITEM_ROW + 1):
        desc = paper_index_ws[f"{DESC_COL}{src_row}"].value
        if desc is None:
            continue
        desc = rename_paper_title(str(desc), prefix)

        si_cell = index_ws[f"{SI_COL}{out_row}"]
        copy_cell_style(si_template, si_cell)
        si_cell.value = next_number

        desc_cell = index_ws[f"{DESC_COL}{out_row}"]
        copy_cell_style(desc_template, desc_cell)
        desc_cell.value = desc
        # Internal link to the matching sheet (description text == sheet name),
        # mirroring how the live Index rows store their links.
        desc_cell.hyperlink = Hyperlink(
            ref=desc_cell.coordinate, target=f"#'{desc}'!A1", display=str(desc),
        )

        # match the source row height if it was customised
        if src_row in paper_index_ws.row_dimensions:
            index_ws.row_dimensions[out_row].height = (
                paper_index_ws.row_dimensions[src_row].height
            )

        next_number += 1
        out_row += 1

    return out_row, next_number


def copy_box(src_ws, dst_ws, src_range, dst_top_row):
    """Copy a rectangular detail box (with its merges) to dst, anchored at dst_top_row.

    src_range is (min_row, min_col, max_row, max_col). Columns are preserved.
    Returns the row index immediately after the pasted box.
    """
    min_r, min_c, max_r, max_c = src_range
    row_shift = dst_top_row - min_r

    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            src = src_ws.cell(row=r, column=c)
            dst = dst_ws.cell(row=r + row_shift, column=c)
            copy_cell_style(src, dst)

    for merged in src_ws.merged_cells.ranges:
        if (merged.min_row >= min_r and merged.max_row <= max_r
                and merged.min_col >= min_c and merged.max_col <= max_c):
            new_min = get_column_letter(merged.min_col) + str(merged.min_row + row_shift)
            new_max = get_column_letter(merged.max_col) + str(merged.max_row + row_shift)
            dst_ws.merge_cells(f"{new_min}:{new_max}")

    return max_r + row_shift + 1


def append_paper_detail_boxes(index_ws, paper_index_ws, start_row):
    """Copy one paper account's Report Details + Login Details boxes.

    Anchors the "Report Details" box at ``start_row`` and returns the next free
    row below the pasted "Login Details" box.
    """
    # Paper "Report Details" box: E4:F7  (cols 5-6)
    next_row = copy_box(paper_index_ws, index_ws, (4, 5, 7, 6), start_row)
    # one blank row, then paper "Login Details (Paper Account)" box: E9:F11
    return copy_box(paper_index_ws, index_ws, (9, 5, 11, 6), next_row + 1)


def merge_pair(live_path: Path, paper_paths, out_path: Path):
    live_wb = openpyxl.load_workbook(live_path)
    index_ws = live_wb["Index"]

    # Detail boxes are stacked in cols E-F; the live Login box ends at row 11.
    box_row = 11 + 1 + BOX_GAP
    # Index item rows are stacked in cols B-C; live item #10 sits at LIVE_LAST_ITEM_ROW.
    out_row = LIVE_LAST_ITEM_ROW + 1
    next_number = index_ws[f"{SI_COL}{LIVE_LAST_ITEM_ROW}"].value + 1  # 11

    # Default ("paper_") account first, then any numbered/labelled extras.
    for paper_path in sorted(paper_paths, key=lambda p: (paper_label(p) != "", paper_label(p))):
        paper_wb = openpyxl.load_workbook(paper_path)
        paper_index_ws = paper_wb["paper_Index"]
        prefix = paper_sheet_prefix(paper_path)

        out_row, next_number = append_paper_index_rows(
            index_ws, paper_index_ws, prefix, out_row, next_number)
        box_row = append_paper_detail_boxes(index_ws, paper_index_ws, box_row) + BOX_GAP

        # Append every paper data sheet (everything except its Index), re-prefixed
        # so two accounts' identically-named sheets stay distinct.
        for ws in paper_wb.worksheets:
            if ws.title == "paper_Index":
                continue
            copy_sheet(ws, live_wb, rename_paper_title(ws.title, prefix))

    live_wb.save(out_path)
    print(f"  merged -> {out_path.name}")


def find_pairs(date_filter=None):
    """Yield (date, live_path, [paper_paths]) for each dated set in Reports/.

    Every ``MIS_paper*_<date>.xlsx`` file is collected, so multiple paper
    accounts sharing a date are merged together (one box-pair each).
    """
    by_date = {}
    for paper_path in REPORTS_DIR.glob("MIS_paper*.xlsx"):
        date = paper_path.stem.rsplit("_", 1)[-1]  # trailing token is the date
        if date_filter and date.lower() != date_filter.lower():
            continue
        by_date.setdefault(date, []).append(paper_path)

    for date in sorted(by_date):
        live_path = REPORTS_DIR / f"MIS_{date}.xlsx"
        if not live_path.exists():
            print(f"! No live file for paper reports dated {date}; skipping")
            continue
        yield date, live_path, by_date[date]


def main():
    date_filter = sys.argv[1] if len(sys.argv) > 1 else None
    found = False
    for date, live_path, paper_paths in find_pairs(date_filter):
        found = True
        papers = " + ".join(p.name for p in paper_paths)
        print(f"Merging {date}: {live_path.name} + {papers}")
        out_path = REPORTS_DIR / f"MIS_merged_{date}.xlsx"
        merge_pair(live_path, paper_paths, out_path)

    if not found:
        target = date_filter or "any date"
        print(f"No live/paper pair found in {REPORTS_DIR} for {target}.")


if __name__ == "__main__":
    main()
