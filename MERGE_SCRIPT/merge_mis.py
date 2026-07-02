"""
Merge a live MIS report and any number of paper MIS reports into one workbook.

For a given date the Reports folder contains one live file plus one paper file
per paper account:
    MIS_<date>.xlsx                -> live account report  (Index + 10 sheets)
    MIS_paper_<date>_<name>.xlsx   -> paper account report (paper_Index + 5 sheets)

The <name> token identifies the paper account (e.g. Shubham, Ajay). A bare
MIS_paper_<date>.xlsx (no name) is treated as the default paper account.

This script produces:
    MIS_merged_<date>.xlsx

Merge rules
-----------
* The merged Index is the LIVE Index, with ONE set of combined paper description
  rows appended below it (paper_Dashboard, paper_Pending Order, ...), renumbered
  to continue the live list (11, 12, ...).
* Each paper account's "Report Details" and "Login Details (Paper Account)"
  boxes are stacked below the live boxes so every account's credentials are kept.
* Sheet order: Index, then every live data sheet, then ONE combined paper sheet
  per report type. The standalone paper_Index sheets are dropped (merged into
  Index).
* The paper data sheets pool every paper account's rows into a single sheet:
    - paper_All Trades        -> rows sorted by Date & Time (UTC) descending,
                                 Sr No renumbered.
    - paper_Open Position,
      paper_Trade Summary     -> rows sorted by First Trade Date (UTC) descending
                                 (only one of the two date pairs can be sorted).
    - paper_Pending Order     -> groups appended account after account (no date
                                 column), Sr No renumbered.
    - paper_Dashboard         -> numeric parameters summed across paper accounts;
                                 text fields kept from the first account.
* A Daywize_Snapshot sheet (Index item 16) holds one row per trade date found
  in the pooled paper_All Trades data, newest first: Open Trades Count is the
  running cumulative fill count up to and including that date, No. of Trades
  taken Today is the fill count on that date alone; Bugs Found/Fixed default to
  0. Rows accumulate across days via a persistent Daywize_Snapshot.xlsx ledger
  beside this script (bug counts entered there are preserved).

Usage
-----
    python merge_mis.py                # merge every dated set found in Reports/
    python merge_mis.py 17Jun2026      # merge just that date
"""

import copy
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

REPORTS_DIR = Path(__file__).resolve().parent / "Reports"

# Persistent daily ledger: one accumulating row per report date, copied into
# every merged workbook as the "Daywize_Snapshot" sheet (newest day on top).
SNAPSHOT_FILE = Path(__file__).resolve().parent / "Daywize_Snapshot.xlsx"
SNAPSHOT_SHEET = "Daywize_Snapshot_paper"
SNAPSHOT_HEADERS = [
    "Date",
    "Open Trades Count",
    "No. of Trades taken Today",
    "No. of Bugs Found",
    "No. of Bugs Fixed",
]

# --- Index layout constants (1-based rows/cols, matching the source files) ---
SI_COL = "B"             # column holding the serial number
DESC_COL = "C"           # column holding the description
LIVE_LAST_ITEM_ROW = 19  # row of the last live item (Pending_Task, #14)
LIVE_FIRST_ITEM_ROW = 6  # row of live item #1 (Dashboard)
LIVE_LOGIN_LAST_ROW = 10 # row of the live "Password" cell (end of Login box)
PAPER_FIRST_ITEM_ROW = 6 # row of paper item #1 in paper_Index
PAPER_LAST_ITEM_ROW = 10 # row of paper item #5 in paper_Index
BOX_GAP = 1              # blank rows between the live and paper detail boxes

# --- Combined paper data-sheet layout ---
DATA_START_ROW = 3       # row 1 = title, row 2 = header, row 3+ = data

# How to pool each paper sheet type across accounts.
#   mode      : "stack" (pool rows + optional date sort), "group" (append order
#               groups, no date sort), "sum" (sum numeric parameter values).
#   sort_col  : 1-based column to sort by (descending), or None for no sort.
#   srno_col  : 1-based serial-number column to renumber 1..N, or None.
PAPER_SHEET_CONFIG = {
    "paper_Dashboard":     ("sum",   None, None),
    "paper_Pending Order": ("group", None, 1),
    "paper_Open Position": ("stack", 1,    None),
    "paper_All Trades":    ("stack", 1,    3),
    "paper_Trade Summary": ("stack", 1,    None),
}

# Paper sheets dropped entirely from the merge (both the Index row and the data
# sheet are skipped). Empty: every paper sheet (incl. paper_Trade Summary) is
# kept so the merged Index carries all 16 items, each with its detail sheet.
EXCLUDED_PAPER_SHEETS = set()


def parse_dt(value):
    """Parse a 'DD-Mon-YYYY HH:MM:SS' (or date-only) timestamp.

    Blank/unparseable values return datetime.min so they sort to the bottom of
    a descending sort; a date with no time is treated as midnight that day.
    """
    if value is None:
        return datetime.min
    text = str(value).strip()
    for fmt in ("%d-%b-%Y %H:%M:%S", "%d-%b-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return datetime.min


def parse_number(value):
    """Return (number, kind) for a dashboard value, or (None, "str") if not numeric.

    kind is "int" for whole numbers, "float" for comma/decimal values. Thousands
    separators (e.g. "626,717,307.52") are stripped before parsing.
    """
    if isinstance(value, bool):
        return None, "str"
    if isinstance(value, int):
        return value, "int"
    if isinstance(value, float):
        return value, "float"
    if isinstance(value, str):
        s = value.replace(",", "").strip()
        try:
            return int(s), "int"
        except ValueError:
            pass
        try:
            return float(s), "float"
        except ValueError:
            pass
    return None, "str"


def copy_cell_style(src, dst):
    """Copy value + full formatting from one cell to another (cross-workbook safe)."""
    dst.value = src.value
    if src.has_style:
        dst.font = copy.copy(src.font)
        dst.border = copy.copy(src.border)
        dst.fill = copy.copy(src.fill)
        dst.number_format = copy.copy(src.number_format)
        dst.protection = copy.copy(src.protection)
        dst.alignment = copy.copy(src.alignment)


def copy_row_height(src_cell, dst_ws, dst_row):
    """Carry the source row's custom height onto dst_row (no-op if unset)."""
    src_ws = src_cell.parent
    dim = src_ws.row_dimensions.get(src_cell.row)
    if dim is not None and dim.height is not None:
        dst_ws.row_dimensions[dst_row].height = dim.height


def parse_paper_name(paper_path):
    """Split a paper file name into (date, account_name).

    MIS_paper_<date>_<name>.xlsx -> (<date>, <name>)
    MIS_paper_<date>.xlsx        -> (<date>, "")   (default paper account)

    e.g. MIS_paper_17Jun2026_Shubham.xlsx -> ("17Jun2026", "Shubham")
    """
    rest = paper_path.stem[len("MIS_paper_"):]  # "17Jun2026_Shubham" | "17Jun2026"
    date, _, name = rest.partition("_")
    return date, name


def paper_label(paper_path):
    """Title-cased account name, "" for the default (unnamed) paper account.

    Normalises filename casing so e.g. "ajay" and "Shubham" both read cleanly
    ("Ajay", "Shubham") in the merged Index and sheet tabs.
    """
    return parse_paper_name(paper_path)[1].title()


def rename_paper_title(title, prefix):
    """Re-prefix a "paper_<name>" sheet/description title for a specific account."""
    if title.startswith("paper_") and prefix != "paper_":
        return prefix + title[len("paper_"):]
    return title


def collect_index_titles(index_ws, paper_index_ws):
    """Build the ordered Index item titles for the merged workbook.

    Order: live "current" sheets first, then the combined paper sheets, then the
    Daywize snapshot, then the archived "... old AC" sheets pushed to the bottom.
    Every title equals its sheet name, so each Index row can link straight to it.
    """
    live_titles = [
        str(index_ws[f"{DESC_COL}{r}"].value)
        for r in range(LIVE_FIRST_ITEM_ROW, LIVE_LAST_ITEM_ROW + 1)
        if index_ws[f"{DESC_COL}{r}"].value is not None
    ]
    current = [t for t in live_titles if not t.endswith("old AC")]
    old_ac = [t for t in live_titles if t.endswith("old AC")]

    paper_titles = []
    for r in range(PAPER_FIRST_ITEM_ROW, PAPER_LAST_ITEM_ROW + 1):
        desc = paper_index_ws[f"{DESC_COL}{r}"].value
        if desc is None:
            continue
        title = rename_paper_title(str(desc), "paper_")
        if title in EXCLUDED_PAPER_SHEETS:
            continue
        paper_titles.append(title)

    return current + paper_titles + [SNAPSHOT_SHEET] + old_ac


def write_index_items(index_ws, titles):
    """(Re)write the Index item list (serial + linked description) top-to-bottom.

    Overwrites the live item rows in place from ``LIVE_FIRST_ITEM_ROW`` down so
    the merged Index carries every sheet in ``titles`` order, each numbered 1..N
    and hyperlinked to its sheet.
    """
    for i, title in enumerate(titles):
        add_index_row(index_ws, LIVE_FIRST_ITEM_ROW + i, i + 1, title)


def reorder_sheets(wb, titles):
    """Arrange the workbook tabs as Index, then ``titles``, matching the Index list.

    Any sheet not named in ``titles`` (there should be none) is kept, appended
    after the ordered ones so nothing is ever dropped.
    """
    order = ["Index"] + [t for t in titles if t in wb.sheetnames]
    order += [name for name in wb.sheetnames if name not in order]
    wb._sheets = [wb[name] for name in order]


def _is_srno_header(value):
    """True if a header cell already labels a serial-number column ("Sr No" etc.)."""
    return (value is not None
            and "".join(str(value).split()).lower().replace(".", "") == "srno")


def _last_data_row(ws, n_cols):
    """Row index of the last non-blank data row (>= header), or header row if none."""
    for r in range(ws.max_row, DATA_START_ROW - 1, -1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, n_cols + 1)):
            return r
    return DATA_START_ROW - 1


def add_srno_and_filter(ws):
    """Ensure column A is a "Sr No" serial column and the sheet has an autofilter.

    Sheets use the title(row1)/header(row2)/data(row3+) layout with a single
    banner merge on row 1. When column A is not already a Sr No column, a new one
    is inserted: every header/data cell shifts one column right (values + styles),
    the banner merge and column widths shift with them, and data rows are numbered
    1..N. Sheets that already lead with Sr No keep their numbering; either way an
    autofilter is applied over the header + data.
    """
    header_row = DATA_START_ROW - 1  # row 2

    if _is_srno_header(ws.cell(header_row, 1).value):
        if ws.auto_filter.ref is None:
            last = _last_data_row(ws, ws.max_column)
            if last >= DATA_START_ROW:
                ws.auto_filter.ref = (
                    f"A{header_row}:{get_column_letter(ws.max_column)}{last}")
        return

    last_col = ws.max_column
    last_row = ws.max_row

    # Drop the row-1 banner merge so it can be re-anchored after the shift.
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    # Shift every header/data cell (rows 2+) one column to the right.
    for col in range(last_col, 0, -1):
        for row in range(header_row, last_row + 1):
            src = ws.cell(row, col)
            dst = ws.cell(row, col + 1)
            dst.value = src.value
            dst._style = copy.copy(src._style)

    # Clear the vacated column A (rows 2+), styled like its new neighbour but
    # with a General number format so serial ints don't inherit a date/number
    # format from the adjacent column (e.g. the Activity sheets' Date column).
    for row in range(header_row, last_row + 1):
        cell = ws.cell(row, 1)
        cell.value = None
        cell._style = copy.copy(ws.cell(row, 2)._style)
        cell.number_format = "General"

    ws.cell(header_row, 1).value = "Sr No"

    srno = 0
    for row in range(DATA_START_ROW, last_row + 1):
        if any(ws.cell(row, c).value not in (None, "") for c in range(2, last_col + 2)):
            srno += 1
            ws.cell(row, 1).value = srno

    # Shift column widths / hidden flags one column right; narrow the Sr No column.
    widths = {}
    for c in range(1, last_col + 1):
        letter = get_column_letter(c)
        if letter in ws.column_dimensions:
            dim = ws.column_dimensions[letter]
            widths[c] = (dim.width, dim.hidden)
    for c in range(last_col, 0, -1):
        if c in widths:
            width, hidden = widths[c]
            tgt = ws.column_dimensions[get_column_letter(c + 1)]
            tgt.width, tgt.hidden = width, hidden
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["A"].hidden = False

    # Re-anchor the banner across the full width, including the new column.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col + 1)

    last = _last_data_row(ws, last_col + 1)
    if last >= DATA_START_ROW:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col + 1)}{last}"


def copy_box(src_ws, dst_ws, src_range, dst_top_row):
    """Copy a rectangular detail box (with its merges) to dst, anchored at dst_top_row.

    src_range is (min_row, min_col, max_row, max_col). Columns are preserved.
    Returns the row index immediately after the pasted box.
    """
    min_r, min_c, max_r, max_c = src_range
    row_shift = dst_top_row - min_r

    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            src = src_ws.cell(row=r, column=c)
            dst = dst_ws.cell(row=r + row_shift, column=c)
            copy_cell_style(src, dst)

    for merged in src_ws.merged_cells.ranges:
        if (merged.min_row >= min_r and merged.max_row <= max_r
                and merged.min_col >= min_c and merged.max_col <= max_c):
            new_min = get_column_letter(merged.min_col) + str(merged.min_row + row_shift)
            new_max = get_column_letter(merged.max_col) + str(merged.max_row + row_shift)
            dst_ws.merge_cells(f"{new_min}:{new_max}")

    return max_r + row_shift + 1


def append_paper_detail_boxes(index_ws, paper_index_ws, start_row, include_report=True):
    """Copy one paper account's Report Details + Login Details boxes.

    Anchors the "Report Details" box at ``start_row`` and returns the next free
    row below the pasted "Login Details" box. When ``include_report`` is False
    the Report Details box is skipped (only the Login box is copied), used to
    drop the duplicate Report Details box of the 2nd+ paper accounts.
    """
    row = start_row
    if include_report:
        # Paper "Report Details" box without the "Data From (cutoff)" row, so
        # only E4:F6 (Report Details / Account ID / Report Date) is copied.
        row = copy_box(paper_index_ws, index_ws, (4, 5, 6, 6), start_row) + BOX_GAP
    # paper "Login Details (Paper Account)" box: E9:F11
    return copy_box(paper_index_ws, index_ws, (9, 5, 11, 6), row)


def copy_header_rows(src_ws, dst_ws, n_rows, n_cols):
    """Copy the top ``n_rows`` (title + header) plus sheet-level formatting.

    Copies cell styles, any merges contained in those rows, column widths and
    the header row heights, so a freshly built sheet keeps the source look.
    """
    for r in range(1, n_rows + 1):
        for c in range(1, n_cols + 1):
            copy_cell_style(src_ws.cell(r, c), dst_ws.cell(r, c))

    for merged in src_ws.merged_cells.ranges:
        if merged.max_row <= n_rows:
            dst_ws.merge_cells(str(merged))

    for col, dim in src_ws.column_dimensions.items():
        d = dst_ws.column_dimensions[col]
        d.width = dim.width
        d.hidden = dim.hidden

    for r in range(1, n_rows + 1):
        if r in src_ws.row_dimensions:
            dst_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height

    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_properties.tabColor = src_ws.sheet_properties.tabColor
    dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines


def add_filter(ws, n_cols, last_row):
    """Enable an Excel AutoFilter over the header row + all data rows.

    No-op when the sheet has no data rows (last_row < header row).
    """
    header_row = DATA_START_ROW - 1
    if last_row < DATA_START_ROW:
        return
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"


def iter_data_rows(ws, n_cols):
    """Yield the source cells (cols 1..n_cols) of each non-blank data row (row 3+)."""
    for r in range(DATA_START_ROW, ws.max_row + 1):
        cells = [ws.cell(r, c) for c in range(1, n_cols + 1)]
        if all(cell.value is None for cell in cells):
            continue
        yield cells


def build_stacked_sheet(dst_wb, title, source_sheets, sort_col, srno_col):
    """Pool every account's data rows into one sheet, optionally date-sorted.

    sort_col (1-based) sorts rows descending by a 'DD-Mon-YYYY HH:MM:SS' column;
    srno_col (1-based), if given, is renumbered 1..N after sorting.
    """
    template = source_sheets[0]
    n_cols = template.max_column
    dst = dst_wb.create_sheet(title=title)
    copy_header_rows(template, dst, DATA_START_ROW - 1, n_cols)

    rows = [cells for ws in source_sheets for cells in iter_data_rows(ws, n_cols)]
    if sort_col:
        rows.sort(key=lambda cells: parse_dt(cells[sort_col - 1].value), reverse=True)

    for i, cells in enumerate(rows):
        out_row = DATA_START_ROW + i
        for c, src in enumerate(cells, start=1):
            copy_cell_style(src, dst.cell(out_row, c))
        copy_row_height(cells[0], dst, out_row)
        if srno_col:
            dst.cell(out_row, srno_col).value = i + 1

    add_filter(dst, n_cols, DATA_START_ROW + len(rows) - 1)
    return dst


def build_grouped_sheet(dst_wb, title, source_sheets, srno_col):
    """Append each account's rows (no date column), keeping multi-row order groups.

    A group leader is a row whose ``srno_col`` holds a value; its continuation
    rows leave that column blank. Leaders are renumbered 1..N across accounts.
    """
    template = source_sheets[0]
    n_cols = template.max_column
    dst = dst_wb.create_sheet(title=title)
    copy_header_rows(template, dst, DATA_START_ROW - 1, n_cols)

    out_row = DATA_START_ROW
    srno = 1
    for ws in source_sheets:
        for cells in iter_data_rows(ws, n_cols):
            is_leader = cells[srno_col - 1].value is not None
            for c, src in enumerate(cells, start=1):
                copy_cell_style(src, dst.cell(out_row, c))
            copy_row_height(cells[0], dst, out_row)
            if is_leader:
                dst.cell(out_row, srno_col).value = srno
                srno += 1
            out_row += 1

    add_filter(dst, n_cols, out_row - 1)
    return dst


def build_summed_sheet(dst_wb, title, source_sheets):
    """Build a Parameter/Value sheet whose numeric values are summed across accounts.

    Non-numeric values (Account ID, dates, "NA", section headers) are kept from
    the first account. Integer values stay integers; comma/decimal values are
    re-formatted with thousands separators and two decimals.
    """
    template = source_sheets[0]
    n_cols = template.max_column
    val_col = 2
    dst = dst_wb.create_sheet(title=title)
    copy_header_rows(template, dst, template.max_row, n_cols)

    for r in range(DATA_START_ROW, template.max_row + 1):
        total = 0
        kind = None
        for ws in source_sheets:
            num, k = parse_number(ws.cell(r, val_col).value)
            if num is None:
                continue
            total += num
            kind = "float" if "float" in (kind, k) else "int"
        if kind == "int":
            dst.cell(r, val_col).value = int(total)
        elif kind == "float":
            dst.cell(r, val_col).value = f"{total:,.2f}"
    return dst


def add_index_row(index_ws, out_row, number, title):
    """Append one Index item (serial + linked description) styled like live rows."""
    si_template = index_ws[f"{SI_COL}{LIVE_LAST_ITEM_ROW}"]
    desc_template = index_ws[f"{DESC_COL}{LIVE_LAST_ITEM_ROW}"]

    si_cell = index_ws[f"{SI_COL}{out_row}"]
    copy_cell_style(si_template, si_cell)
    si_cell.value = number

    desc_cell = index_ws[f"{DESC_COL}{out_row}"]
    copy_cell_style(desc_template, desc_cell)
    desc_cell.value = title
    # Native in-workbook link (location, no relationship) exactly as the live
    # Index stores its links, so clicking jumps straight to the sheet.
    desc_cell.hyperlink = Hyperlink(
        ref=desc_cell.coordinate, location=f"'{title}'!A1", display=title)


def paper_daily_metrics(paper_wbs):
    """Per trade-date fill counts pooled across all paper accounts.

    Reads every paper_All Trades sheet (Date & Time UTC in column 1), tallies
    fills per trade date, then returns an ordered dict
    {date_str: [open_cumulative, trades_today]} where date_str is 'DD-Mon-YYYY'.
    ``open_cumulative`` is the running total of all fills up to and including that
    date (oldest to newest); ``trades_today`` is the fill count on that date
    alone.
    """
    per_day = {}  # date -> fill count on that date
    for wb in paper_wbs:
        if "paper_All Trades" not in wb.sheetnames:
            continue
        ws = wb["paper_All Trades"]
        for cells in iter_data_rows(ws, ws.max_column):
            dt = parse_dt(cells[0].value)
            if dt == datetime.min:
                continue
            day = dt.date()
            per_day[day] = per_day.get(day, 0) + 1

    metrics = {}
    cumulative = 0
    for day in sorted(per_day):
        cumulative += per_day[day]
        metrics[day.strftime("%d-%b-%Y")] = [cumulative, per_day[day]]
    return metrics


def load_ledger():
    """Read the persistent snapshot ledger as {date_str: [open, trades, found, fixed]}."""
    if not SNAPSHOT_FILE.exists():
        return {}
    wb = openpyxl.load_workbook(SNAPSHOT_FILE)
    ws = wb.active
    ledger = {}
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None:
            continue
        # Excel may coerce the date string into a real date on manual edits.
        key = raw.strftime("%d-%b-%Y") if isinstance(raw, datetime) else str(raw)
        ledger[key] = [ws.cell(r, c).value for c in range(2, len(SNAPSHOT_HEADERS) + 1)]
    return ledger


def save_ledger(ledger):
    """Persist the ledger back to disk, newest date first."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SNAPSHOT_SHEET
    for c, head in enumerate(SNAPSHOT_HEADERS, start=1):
        ws.cell(1, c).value = head
    for i, key in enumerate(sorted(ledger, key=parse_dt, reverse=True)):
        ws.cell(2 + i, 1).value = key
        for c, val in enumerate(ledger[key], start=2):
            ws.cell(2 + i, c).value = val
    wb.save(SNAPSHOT_FILE)


def update_ledger(daily_metrics):
    """Upsert one row per trade date; keep any bug counts already entered."""
    ledger = load_ledger()
    for key, (open_cum, trades_today) in daily_metrics.items():
        prior = ledger.get(key)
        bugs_found = prior[2] if prior and prior[2] is not None else 0
        bugs_fixed = prior[3] if prior and prior[3] is not None else 0
        ledger[key] = [open_cum, trades_today, bugs_found, bugs_fixed]
    save_ledger(ledger)
    return ledger


def build_snapshot_sheet(dst_wb, ledger, template_ws):
    """Render the ledger into a Daywize_Snapshot sheet styled like the report."""
    n_cols = len(SNAPSHOT_HEADERS)
    dst = dst_wb.create_sheet(SNAPSHOT_SHEET)

    title_src = template_ws.cell(1, 1)   # banner row style
    header_src = template_ws.cell(2, 1)  # column-header style
    data_src = template_ws.cell(DATA_START_ROW, 1)

    for c in range(1, n_cols + 1):
        copy_cell_style(title_src, dst.cell(1, c))
    dst.cell(1, 1).value = "DAYWISE SNAPSHOT"
    dst.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    for c, head in enumerate(SNAPSHOT_HEADERS, start=1):
        copy_cell_style(header_src, dst.cell(2, c))
        dst.cell(2, c).value = head

    keys = sorted(ledger, key=parse_dt, reverse=True)
    for i, key in enumerate(keys):
        out_row = DATA_START_ROW + i
        for c, val in enumerate([key, *ledger[key]], start=1):
            copy_cell_style(data_src, dst.cell(out_row, c))
            dst.cell(out_row, c).value = val

    for c in range(1, n_cols + 1):
        dst.column_dimensions[get_column_letter(c)].width = 24
    dst.freeze_panes = "A3"
    add_filter(dst, n_cols, DATA_START_ROW + len(keys) - 1)
    return dst


def merge_pair(live_path: Path, paper_paths, out_path: Path):
    live_wb = openpyxl.load_workbook(live_path)
    index_ws = live_wb["Index"]

    # Default ("paper_") account first, then any numbered/labelled extras.
    ordered_paths = sorted(paper_paths, key=lambda p: (paper_label(p) != "", paper_label(p)))
    paper_wbs = [openpyxl.load_workbook(p) for p in ordered_paths]

    # Rebuild the whole Index item list in the fixed presentation order: live
    # current sheets, the combined paper sheets, the Daywize snapshot, then the
    # archived "old AC" sheets at the bottom. Each row links to its sheet.
    index_titles = collect_index_titles(index_ws, paper_wbs[0]["paper_Index"])
    write_index_items(index_ws, index_titles)

    # Detail boxes: stack every paper account's Report + Login boxes (cols E-F)
    # so all credentials are retained, below the live Login box.
    box_row = LIVE_LOGIN_LAST_ROW + 1 + BOX_GAP
    for i, wb in enumerate(paper_wbs):
        # Only the first paper account keeps its Report Details box; the rest
        # are duplicates (same Account ID/dates), so just stack their Login box.
        box_row = append_paper_detail_boxes(
            index_ws, wb["paper_Index"], box_row, include_report=(i == 0)) + BOX_GAP

    # One combined sheet per report type, pooling every paper account's data.
    for ws in paper_wbs[0].worksheets:
        title = ws.title
        if title == "paper_Index" or title in EXCLUDED_PAPER_SHEETS:
            continue
        sources = [wb[title] for wb in paper_wbs if title in wb.sheetnames]
        mode, sort_col, srno_col = PAPER_SHEET_CONFIG.get(title, ("stack", None, None))
        if mode == "sum":
            build_summed_sheet(live_wb, title, sources)
        elif mode == "group":
            build_grouped_sheet(live_wb, title, sources, srno_col)
        else:
            build_stacked_sheet(live_wb, title, sources, sort_col, srno_col)

    # Daily snapshot: tally per-trade-date fill counts from the pooled paper
    # All Trades data, update the persistent ledger, then render the full
    # (newest-first) history into the merged workbook.
    daily = paper_daily_metrics(paper_wbs)
    ledger = update_ledger(daily)
    build_snapshot_sheet(live_wb, ledger, live_wb["Open Position"])

    # Give every data sheet a Sr No first column + autofilter (Index excepted).
    for ws in live_wb.worksheets:
        if ws.title != "Index":
            add_srno_and_filter(ws)

    # Order the sheet tabs to match the Index list (Index first).
    reorder_sheets(live_wb, index_titles)

    live_wb.save(out_path)
    latest = max(daily, key=parse_dt, default=None)
    open_cum, trades_today = daily[latest] if latest else (0, 0)
    print(f"  merged -> {out_path.name}  "
          f"(dates={len(daily)}, cumulative={open_cum}, "
          f"latest={latest}, trades_today={trades_today})")


def find_pairs(date_filter=None):
    """Yield (date, live_path, [paper_paths]) for each dated set in Reports/.

    Every ``MIS_paper*_<date>.xlsx`` file is collected, so multiple paper
    accounts sharing a date are merged together (one box-pair each).
    """
    by_date = {}
    for paper_path in REPORTS_DIR.glob("MIS_paper*.xlsx"):
        date, _ = parse_paper_name(paper_path)  # leading token after MIS_paper_
        if date_filter and date.lower() != date_filter.lower():
            continue
        by_date.setdefault(date, []).append(paper_path)

    for date in sorted(by_date):
        live_path = REPORTS_DIR / f"MIS_{date}.xlsx"
        if not live_path.exists():
            print(f"! No live file for paper reports dated {date}; skipping")
            continue
        yield date, live_path, by_date[date]


def main():
    date_filter = sys.argv[1] if len(sys.argv) > 1 else None
    found = False
    for date, live_path, paper_paths in find_pairs(date_filter):
        found = True
        papers = " + ".join(p.name for p in paper_paths)
        print(f"Merging {date}: {live_path.name} + {papers}")
        out_path = REPORTS_DIR / f"MIS_merged_{date}.xlsx"
        merge_pair(live_path, paper_paths, out_path)

    if not found:
        target = date_filter or "any date"
        print(f"No live/paper pair found in {REPORTS_DIR} for {target}.")


if __name__ == "__main__":
    main()
