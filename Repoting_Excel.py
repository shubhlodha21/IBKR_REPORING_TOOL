"""
Fetch IBKR Flex Query trades and write an Excel file with four sheets:
  - Pending Order  (live from TWS)
  - PreviousDay
  - Past7Days
  - Past30Days

Usage:
    pip install requests openpyxl ibapi
    python Repoting_Excel.py

Notes:
  - The Flex Query (ID below) must be configured in IBKR to cover at least the
    last 30 days, otherwise the 30-day sheet will be incomplete.
  - Regenerate your token after use; it is sensitive.
  - TWS or IB Gateway must be running for the Pending Order sheet.
"""

import os
import sys
import time
import threading
import datetime as dt
import xml.etree.ElementTree as ET
from collections import defaultdict

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ibapi.client import EClient
from ibapi.wrapper import EWrapper

# ----------------------------------------------------------------------
# CONFIG  (set IBKR_TOKEN and IBKR_QUERY_ID as environment variables)
# ----------------------------------------------------------------------
TOKEN    = os.environ.get("IBKR_TOKEN", "")
QUERY_ID = os.environ.get("IBKR_QUERY_ID", "")

if not TOKEN or not QUERY_ID:
    sys.exit("Error: IBKR_TOKEN and IBKR_QUERY_ID must be set as environment variables.\n"
             "Copy .env.example to .env and fill in your credentials, then run via run.sh / run.ps1.")

_REPORTS_DIR = "reports"
os.makedirs(_REPORTS_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(_REPORTS_DIR, dt.date.today().strftime("Reports_%d%b%Y") + ".xlsx")

BASE     = "https://ndcdyn.interactivebrokers.com/AccountManagement/FlexWebService"
SEND_URL = f"{BASE}/SendRequest"
GET_URL  = f"{BASE}/GetStatement"
VERSION  = "3"

MAX_POLL_ATTEMPTS = 12   # GetStatement retries
POLL_WAIT_SECONDS = 5    # wait between retries while statement is generating

# TWS live connection (for Pending Orders)
HOST       = "127.0.0.1"
PORT       = 7496        # 7496 live TWS | 7497 paper | 4001/4002 Gateway
CLIENT_ID  = 998
ORDER_WAIT = 6           # seconds to wait for open-order callbacks

_UNSET = (0, 1.7976931348623157e308)  # IBKR "no value" sentinels


# ----------------------------------------------------------------------
# TWS: fetch live pending orders
# ----------------------------------------------------------------------
class IBOrderApp(EWrapper, EClient):
    def __init__(self):
        EClient.__init__(self, self)
        self.orders_by_symbol = defaultdict(list)
        self._done = threading.Event()
        self.connected_ok = False

    def error(self, reqId, errorCode, errorString, advancedOrderRejectJson=""):
        if errorCode in (502, 504, 1100, 1300):
            print(f"[TWS] {errorCode}: {errorString}", file=sys.stderr)
        elif errorCode not in (2104, 2106, 2107, 2119, 2158, 2100):
            print(f"[TWS] id={reqId} code={errorCode}: {errorString}", file=sys.stderr)

    def connectAck(self):
        self.connected_ok = True

    def nextValidId(self, orderId):
        self.connected_ok = True
        self.reqAllOpenOrders()

    def openOrder(self, orderId, contract, order, orderState):
        trigger = order.auxPrice if order.auxPrice not in _UNSET else None
        limit   = order.lmtPrice if order.lmtPrice not in _UNSET else None
        self.orders_by_symbol[contract.symbol].append({
            "orderId":   orderId,
            "action":    order.action,
            "orderType": order.orderType,
            "trigger":   trigger,
            "limit":     limit,
            "quantity":  int(order.totalQuantity),
            "parentId":  order.parentId,
        })

    def openOrderEnd(self):
        self._done.set()


def fetch_pending_orders():
    app = IBOrderApp()
    try:
        app.connect(HOST, PORT, CLIENT_ID)
    except Exception as e:
        print(f"[TWS] Could not connect ({e}) — Pending Order sheet will be empty.",
              file=sys.stderr)
        return {}
    threading.Thread(target=app.run, daemon=True).start()
    app._done.wait(timeout=ORDER_WAIT)
    time.sleep(0.3)
    app.disconnect()
    return app.orders_by_symbol


def _fmt_price(v):
    return f"{v:.2f}" if v is not None else ""


def _sl_pct(entry_trigger, sl_trigger):
    if entry_trigger and sl_trigger:
        return f"{abs(entry_trigger - sl_trigger) / entry_trigger * 100:.2f}%"
    return ""


def build_pending_rows(orders_by_symbol):
    rows = []
    sr = 0
    for symbol, legs in orders_by_symbol.items():
        sr += 1
        parents  = [o for o in legs if not o["parentId"]]
        children = [o for o in legs if o["parentId"]]
        entry = parents[0] if parents else legs[0]
        stop  = children[0] if children else (legs[1] if len(legs) > 1 else None)

        rows.append([
            sr, symbol, entry["action"], entry["orderType"],
            _fmt_price(entry["trigger"]), _fmt_price(entry["limit"]),
            _sl_pct(entry["trigger"], stop["trigger"] if stop else None),
            entry["quantity"],
        ])
        if stop:
            rows.append([
                "", "", stop["action"], stop["orderType"],
                _fmt_price(stop["trigger"]), _fmt_price(stop["limit"]),
                "", stop["quantity"],
            ])
    return rows


PENDING_HEADERS = ["Sr No", "Contract", "Action", "Type",
                   "Trigger", "Limit", "SL-Percentage", "Quantity"]
PENDING_WIDTHS  = [7, 12, 9, 10, 10, 10, 15, 10]


# ----------------------------------------------------------------------
# STEP 1: request the statement -> reference code
# ----------------------------------------------------------------------
def send_request():
    r = requests.get(SEND_URL, params={"t": TOKEN, "q": QUERY_ID, "v": VERSION}, timeout=30)
    r.raise_for_status()
    root = ET.fromstring(r.content)
    status = root.findtext("Status")
    if status != "Success":
        code = root.findtext("ErrorCode")
        msg  = root.findtext("ErrorMessage")
        sys.exit(f"SendRequest failed [{code}]: {msg}")
    ref = root.findtext("ReferenceCode")
    print(f"Reference code: {ref}")
    return ref


# ----------------------------------------------------------------------
# STEP 2: poll for the statement XML
# ----------------------------------------------------------------------
def get_statement(ref):
    for attempt in range(1, MAX_POLL_ATTEMPTS + 1):
        r = requests.get(GET_URL, params={"t": TOKEN, "q": ref, "v": VERSION}, timeout=60)
        r.raise_for_status()
        text = r.text
        # While generating, IBKR returns a small XML with a "Warn" status.
        if "<FlexQueryResponse" in text:
            print("Statement retrieved.")
            return text
        try:
            root = ET.fromstring(r.content)
            code = root.findtext("ErrorCode")
            msg  = root.findtext("ErrorMessage")
            if code and code not in ("1019",):  # 1019 = statement not yet ready
                sys.exit(f"GetStatement failed [{code}]: {msg}")
            print(f"  not ready yet (attempt {attempt}/{MAX_POLL_ATTEMPTS}), waiting...")
        except ET.ParseError:
            pass
        time.sleep(POLL_WAIT_SECONDS)
    sys.exit("Statement not ready after maximum retries.")


# ----------------------------------------------------------------------
# STEP 3: parse trades
# ----------------------------------------------------------------------
def parse_trades(xml_text):
    root   = ET.fromstring(xml_text)
    trades = root.findall(".//Trade")
    if not trades:
        # Some configs nest under Order; fall back gracefully.
        trades = root.findall(".//Order")
    rows = []
    for t in trades:
        rows.append(dict(t.attrib))
    return rows


def parse_trade_date(row):
    """Extract a date from common IBKR datetime attributes."""
    raw = row.get("dateTime") or row.get("tradeDate") or row.get("reportDate") or ""
    raw = raw.strip()
    if not raw:
        return None
    # Common formats: 'YYYYMMDD;HHMMSS', 'YYYYMMDD', 'YYYY-MM-DD, HH:MM:SS'
    raw = raw.replace(",", " ").replace(";", " ")
    datepart = raw.split()[0]
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(datepart, fmt).date()
        except ValueError:
            continue
    return None


# ----------------------------------------------------------------------
# STEP 4: aggregate raw trades into per-contract summary rows
# ----------------------------------------------------------------------
HEADERS = [
    "Contract", "Buys", "Sells", "Net",
    "Avg (bought)", "Avg (sold)",
    "Total (bought)", "Total (sold)",
    "Exchange List", "Net Total",
    "Commission", "Net Incl. Commission",
    "PnL", "Unrealized PnL",
]


def _flt(val):
    try:
        return float(val) if val not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def aggregate(trade_rows):
    data = defaultdict(lambda: {
        "buy_qty": 0.0, "sell_qty": 0.0,
        "buy_value": 0.0, "sell_value": 0.0,
        "commission": 0.0, "pnl": 0.0, "unrealized": 0.0,
        "exchanges": set(),
    })

    for r in trade_rows:
        symbol     = r.get("symbol") or r.get("description") or "UNKNOWN"
        side       = r.get("buySell", "").upper().strip()
        qty        = abs(_flt(r.get("quantity", 0)))
        price      = abs(_flt(r.get("tradePrice", 0)))
        # tradeMoney is signed in IBKR; use absolute value
        money      = abs(_flt(r.get("tradeMoney", 0))) or qty * price
        commission = _flt(r.get("commission", 0))
        pnl        = _flt(r.get("fifoPnlRealized", 0))
        unrealized = _flt(r.get("mtmPnl", 0)) or _flt(r.get("unrealizedPnl", 0))
        exchange   = r.get("exchange", "")

        c = data[symbol]
        if side in ("BUY", "B"):
            c["buy_qty"]   += qty
            c["buy_value"] += money
        elif side in ("SELL", "S"):
            c["sell_qty"]   += qty
            c["sell_value"] += money
        c["commission"] += commission
        c["pnl"]        += pnl
        c["unrealized"] += unrealized
        if exchange:
            c["exchanges"].add(exchange)

    result = []
    for symbol, c in data.items():
        avg_b     = c["buy_value"]  / c["buy_qty"]  if c["buy_qty"]  else 0.0
        avg_s     = c["sell_value"] / c["sell_qty"] if c["sell_qty"] else 0.0
        net_total = c["sell_value"] - c["buy_value"]
        result.append({
            "Contract":             symbol,
            "Buys":                 round(c["buy_qty"], 4),
            "Sells":                round(c["sell_qty"], 4),
            "Net":                  round(c["buy_qty"] - c["sell_qty"], 4),
            "Avg (bought)":         round(avg_b, 6),
            "Avg (sold)":           round(avg_s, 6),
            "Total (bought)":       round(c["buy_value"], 2),
            "Total (sold)":         round(c["sell_value"], 2),
            "Exchange List":        ", ".join(sorted(c["exchanges"])),
            "Net Total":            round(net_total, 2),
            "Commission":           round(c["commission"], 2),
            "Net Incl. Commission": round(net_total + c["commission"], 2),
            "PnL":                  round(c["pnl"], 2),
            "Unrealized PnL":       round(c["unrealized"], 2),
        })
    result.sort(key=lambda x: x["Contract"])
    return result


# ----------------------------------------------------------------------
# STEP 5: write Excel
# ----------------------------------------------------------------------
def write_excel(rows, pending_rows):
    today     = dt.date.today()
    yesterday = today - dt.timedelta(days=1)
    cutoff_7  = today - dt.timedelta(days=7)
    cutoff_30 = today - dt.timedelta(days=30)

    rows_prev, rows_7, rows_30 = [], [], []
    for row in rows:
        d = parse_trade_date(row)
        if d is None:
            continue
        if d == yesterday:
            rows_prev.append(row)
        if d >= cutoff_7:
            rows_7.append(row)
        if d >= cutoff_30:
            rows_30.append(row)

    agg_prev = aggregate(rows_prev)
    agg_7    = aggregate(rows_7)
    agg_30   = aggregate(rows_30)

    wb = Workbook()
    _fill_pending_sheet(wb.active,         pending_rows)
    _fill_sheet(wb.create_sheet(), "PreviousDay", agg_prev)
    _fill_sheet(wb.create_sheet(), "Past7Days",   agg_7)
    _fill_sheet(wb.create_sheet(), "Past30Days",  agg_30)
    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}: {len(pending_rows)} pending orders, "
          f"{len(agg_prev)} contracts (prev day), "
          f"{len(agg_7)} contracts (7d), {len(agg_30)} contracts (30d)")


def _fill_pending_sheet(ws, rows):
    ws.title = "Pending Order"
    yellow = PatternFill("solid", fgColor="FFFF00")
    ws.append(PENDING_HEADERS)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.fill      = yellow
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for i, w in enumerate(PENDING_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _fill_sheet(ws, title, agg_rows):
    ws.title = title
    yellow = PatternFill("solid", fgColor="FFFF00")
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.fill      = yellow
        cell.alignment = Alignment(horizontal="center")
    for row in agg_rows:
        ws.append([row.get(h, "") for h in HEADERS])
    for i, h in enumerate(HEADERS, start=1):
        values = [str(r.get(h, "")) for r in agg_rows]
        width  = max(len(h), *(len(v) for v in values)) if values else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 12), 30)
    ws.freeze_panes = "A2"


def main():
    print("Fetching pending orders from TWS...")
    orders_by_symbol = fetch_pending_orders()
    pending_rows = build_pending_rows(orders_by_symbol)
    print(f"  {len(pending_rows)} pending order rows collected.")

    ref      = send_request()
    xml_text = get_statement(ref)
    rows     = parse_trades(xml_text)
    print(f"Parsed {len(rows)} trade records.")
    if not rows:
        print("No trades found. Check the query's date range and section config.")
    write_excel(rows, pending_rows)


if __name__ == "__main__":
    main()
