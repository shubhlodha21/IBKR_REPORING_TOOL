"""
Fetch IBKR Flex Query trades and write an Excel file with sheets:
  - Dashboard         (account info + 7-day & prev-day P&L summary)
  - Pending Order     (live from TWS)
  - Running Positions (net open positions from 24 May 2026 onwards)
  - PreviousDay
  - Past7Days
  - Past30Days

Usage:
    pip install requests openpyxl ibapi
    python Repoting_Excel.py

Notes:
  - The Flex Query must cover at least 30 days.
  - TWS or IB Gateway must be running for the Pending Order sheet.
  - Set IBKR_TOKEN and IBKR_QUERY_ID as environment variables (via .env).
"""

import os
import sys
import json
import time
import threading
import datetime as dt
import xml.etree.ElementTree as ET
from collections import defaultdict

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.execution import ExecutionFilter

# ----------------------------------------------------------------------
# CONFIG  (set IBKR_TOKEN and IBKR_QUERY_ID as environment variables)
# ----------------------------------------------------------------------
TOKEN    = os.environ.get("IBKR_TOKEN", "")
QUERY_ID = os.environ.get("IBKR_QUERY_ID", "")

if not TOKEN or not QUERY_ID:
    sys.exit("Error: IBKR_TOKEN and IBKR_QUERY_ID must be set as environment variables.\n"
             "Copy .env.example to .env and fill in your credentials, then run via run.ps1.")

_REPORTS_DIR = "reports"
os.makedirs(_REPORTS_DIR, exist_ok=True)
OUTPUT_FILE = os.path.join(_REPORTS_DIR, dt.date.today().strftime("MIS_paper_%d%b%Y") + "_Ajay.xlsx")

# Persistent order ledger — accumulates trigger/limit from live open orders on
# every run, so trades can be matched to the prices set when the order was placed.
# (IBKR does NOT retain this on executed trades, so we must capture it ourselves.)
_BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
LEDGER_FILE = os.path.join(_BASE_DIR, "orders_ledger.json")

# Persistent trades cache — accumulates today's live TWS executions on every run.
# reqExecutions only ever returns the CURRENT day's fills, while the Flex
# statement lags 1-2 days behind, so the most recent trading day(s) fall into a
# gap and never appear. Persisting each day's executions bridges that gap until
# Flex finally publishes them. (See update_trades_cache.)
TRADES_CACHE_FILE = os.path.join(_BASE_DIR, "trades_cache.json")
TRADES_CACHE_DAYS = 30   # retention window — generously covers the Flex lag

# Persistent FULL trade history — unlike the trades cache (which only bridges the
# Flex lag and prunes after TRADES_CACHE_DAYS), this accumulates EVERY trade ever
# seen, from both the Flex statement and live TWS fills, and is NEVER pruned. The
# report is built from this store, so the complete record from DATA_FROM onward is
# preserved even when the Flex Query only returns a short rolling window, and a
# single empty/failed day can never shrink the sheets. (See update_trade_history.)
TRADE_HISTORY_FILE = os.path.join(_BASE_DIR, "trade_history.json")

BASE     = "https://ndcdyn.interactivebrokers.com/AccountManagement/FlexWebService"
SEND_URL = f"{BASE}/SendRequest"
GET_URL  = f"{BASE}/GetStatement"
VERSION  = "3"

MAX_POLL_ATTEMPTS = 12
POLL_WAIT_SECONDS = 5

# TWS live connection
HOST       = "127.0.0.1"
PORT       = 7497        # 7496 live TWS | 7497 paper | 4001/4002 Gateway
CLIENT_ID  = 999
ORDER_WAIT = 8           # seconds to wait for open-order callbacks
ACCT_WAIT  = 8           # seconds to wait for account-summary callbacks
EXEC_WAIT  = 15          # seconds to wait for executions (0.5 s timer + fills + commissions)

_UNSET = (0, 1.7976931348623157e308)

# Global data cutoff — the report only includes trades on/after this date.
# This is shown on the Dashboard so the reporting window is always explicit.
DATA_FROM = dt.date(2026, 5, 25)

# Running positions use the same global cutoff.
RUNNING_POSITIONS_FROM = DATA_FROM

# Risk limits shown on the Dashboard (hardcoded).
MAX_EXPOSURE   = 'NA'
DAILY_MAX_LOSS = 'NA'

# Cumulative loss attributed to bugs — a manually-maintained figure shown on the
# Dashboard. Update this value here when it changes; it stays fixed otherwise.
LOSS_DUE_TO_BUGS = 0

# IST offset for report timestamps
_IST = dt.timezone(dt.timedelta(hours=5, minutes=30))

# Account summary tags to request from TWS
ACCOUNT_SUMMARY_TAGS = (
    "NetLiquidation,TotalCashValue,SettledCash,AccruedCash,"
    "BuyingPower,AvailableFunds,ExcessLiquidity,"
    "InitMarginReq,MaintMarginReq,"
    "GrossPositionValue,UnrealizedPnL,RealizedPnL"
)

# Display labels for each account tag (label, tag_key)
ACCOUNT_DISPLAY = [
    ("Net Liquidation (Total Funds)", "NetLiquidation"),
    ("Total Cash Value",              "TotalCashValue"),
    ("Settled Cash",                  "SettledCash"),
    ("Accrued Cash",                  "AccruedCash"),
    ("Buying Power",                  "BuyingPower"),
    ("Available Funds",               "AvailableFunds"),
    ("Excess Liquidity",              "ExcessLiquidity"),
    ("Initial Margin Req.",           "InitMarginReq"),
    ("Maintenance Margin Req.",       "MaintMarginReq"),
    ("Gross Position Value",          "GrossPositionValue"),
    ("Unrealized PnL",                "UnrealizedPnL"),
    ("Realized PnL",                  "RealizedPnL"),
]


# ----------------------------------------------------------------------
# TWS: fetch live pending orders + account summary in one connection
# ----------------------------------------------------------------------
class IBOrderApp(EWrapper, EClient):
    def __init__(self):
        EClient.__init__(self, self)
        self.orders_by_symbol = defaultdict(list)
        self.account_data     = {}          # tag -> (value, currency)
        self.executions       = {}          # execId -> execution dict
        self.completed_orders = []          # today's filled/cancelled orders
        self._orders_done     = threading.Event()
        self._acct_done       = threading.Event()
        self._exec_done       = threading.Event()
        self._completed_done  = threading.Event()
        self.connected_ok     = False

    def error(self, reqId, errorCode, errorString, advancedOrderRejectJson=""):
        if errorCode in (502, 504, 1100, 1300):
            print(f"[TWS] {errorCode}: {errorString}", file=sys.stderr)
        elif errorCode not in (2104, 2106, 2107, 2119, 2158, 2100):
            print(f"[TWS] id={reqId} code={errorCode}: {errorString}", file=sys.stderr)

    def connectAck(self):
        self.connected_ok = True

    def nextValidId(self, orderId):
        self.connected_ok = True
        self.reqAllOpenOrders()
        self.reqAccountSummary(9001, "All", ACCOUNT_SUMMARY_TAGS)
        # Delay executions request by 0.5 s so the first two requests
        # don't congest the message loop, and set a 24-hour window so
        # pre-market fills and all intraday trades are captured.
        threading.Timer(0.5, self._req_executions).start()

    def _req_executions(self):
        # Empty filter → ALL of the current trading day's executions.
        # NOTE: IBKR's API only ever returns the *current day's* executions;
        # a time filter cannot reach prior days, and a malformed one can
        # suppress results entirely — so we pass an unfiltered request.
        self.reqExecutions(10001, ExecutionFilter())
        # Today's completed (filled/cancelled) orders — these carry the order's
        # limit/trigger even after they've filled, which open orders no longer do.
        self.reqCompletedOrders(False)   # False = all, not just API-placed

    # ── completed orders (today's filled/cancelled, with limit/trigger) ──
    def completedOrder(self, contract, order, orderState):
        trigger = order.auxPrice if order.auxPrice not in _UNSET else None
        limit   = order.lmtPrice if order.lmtPrice not in _UNSET else None
        oid     = getattr(order, "permId", None) or getattr(order, "orderId", None)
        self.completed_orders.append({
            "orderId":   f"C{oid}",
            "symbol":    contract.symbol,
            "action":    order.action,
            "orderType": order.orderType,
            "trigger":   trigger,
            "limit":     limit,
            "quantity":  int(order.totalQuantity) if order.totalQuantity else 0,
            "status":    orderState.status,
        })

    def completedOrdersEnd(self):
        print(f"[TWS] completedOrdersEnd - {len(self.completed_orders)} completed order(s).")
        self._completed_done.set()

    # ── pending orders ────────────────────────────────────────────────
    def openOrder(self, orderId, contract, order, orderState):
        trigger = order.auxPrice if order.auxPrice not in _UNSET else None
        limit   = order.lmtPrice if order.lmtPrice not in _UNSET else None
        self.orders_by_symbol[contract.symbol].append({
            "orderId":   orderId,
            "action":    order.action,
            "orderType": order.orderType,
            "trigger":   trigger,
            "limit":     limit,
            "quantity":  int(order.totalQuantity),
            "parentId":  order.parentId,
        })

    def openOrderEnd(self):
        self._orders_done.set()

    # ── account summary ───────────────────────────────────────────────
    def accountSummary(self, reqId, account, tag, value, currency):
        self.account_data[tag] = (value, currency)

    def accountSummaryEnd(self, reqId):
        self._acct_done.set()

    # ── today's executions ────────────────────────────────────────────
    def execDetails(self, reqId, contract, execution):
        self.executions[execution.execId] = {
            "account":    execution.acctNumber,
            "contract":   contract.symbol,
            "action":     "BUY" if execution.side in ("BOT", "BUY") else "SELL",
            "quantity":   execution.shares,
            "price":      execution.price,
            "currency":   contract.currency,
            "exchange":   execution.exchange,
            "time":       execution.time,
            "commission": None,             # filled in by commissionReport
        }

    def commissionReport(self, commissionReport):
        ex = self.executions.get(commissionReport.execId)
        if ex is not None:
            ex["commission"] = commissionReport.commission

    def execDetailsEnd(self, reqId):
        print(f"[TWS] execDetailsEnd - {len(self.executions)} execution(s) received so far.")
        self._exec_done.set()


def fetch_tws_data():
    """Connect to TWS once and fetch open orders, account summary, today's
    executions, and today's completed orders."""
    app = IBOrderApp()
    try:
        app.connect(HOST, PORT, CLIENT_ID)
    except Exception as e:
        print(f"[TWS] Could not connect ({e}) - live sheets will be empty.",
              file=sys.stderr)
        return {}, {}, {}, []
    threading.Thread(target=app.run, daemon=True).start()
    # Each feed has its own generous timeout so a slow/missing feed
    # doesn't block the others.
    app._orders_done.wait(timeout=ORDER_WAIT)
    app._acct_done.wait(timeout=ACCT_WAIT)
    app._exec_done.wait(timeout=EXEC_WAIT)
    app._completed_done.wait(timeout=EXEC_WAIT)
    # Commission reports arrive asynchronously after execDetailsEnd;
    # give TWS extra time to push them all through before we disconnect.
    time.sleep(1.5)
    app.disconnect()
    return app.orders_by_symbol, app.account_data, app.executions, app.completed_orders


# ----------------------------------------------------------------------
# Persistent order ledger: capture trigger/limit from live open orders so
# trades can later be matched to the prices set when the order was placed.
# ----------------------------------------------------------------------
def _norm_type(t):
    """Normalise order type for matching (TWS 'STP LMT' vs Flex 'STPLMT')."""
    return (t or "").upper().replace(" ", "")


def load_order_ledger():
    if os.path.exists(LEDGER_FILE):
        try:
            with open(LEDGER_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            print("[ledger] existing ledger unreadable - starting fresh.", file=sys.stderr)
    return {}


def update_order_ledger(ledger, orders_by_symbol, completed_orders=None):
    """Merge live open orders AND today's completed orders into the ledger.

    Open orders cover still-working orders; completed orders cover ones that
    already filled today (which an EOD run would otherwise miss). Both carry
    the limit/trigger prices set when the order was placed.
    """
    today = dt.date.today().isoformat()

    def _merge(oid, symbol, action, order_type, limit, trigger, quantity):
        rec = ledger.get(oid, {})
        rec.update({
            "symbol":    symbol,
            "action":    (action or "").upper(),
            "orderType": _norm_type(order_type),
            "limit":     limit,
            "trigger":   trigger,
            "quantity":  quantity,
            "last_seen": today,
        })
        rec.setdefault("first_seen", today)
        ledger[oid] = rec

    for symbol, legs in orders_by_symbol.items():
        for o in legs:
            _merge(str(o["orderId"]), symbol, o["action"], o["orderType"],
                   o["limit"], o["trigger"], o["quantity"])

    for o in (completed_orders or []):
        _merge(str(o["orderId"]), o["symbol"], o["action"], o["orderType"],
               o["limit"], o["trigger"], o["quantity"])

    return ledger


def save_order_ledger(ledger):
    try:
        with open(LEDGER_FILE, "w", encoding="utf-8") as f:
            json.dump(ledger, f, indent=2)
    except OSError as e:
        print(f"[ledger] could not save: {e}", file=sys.stderr)


def _parse_exec_date(raw):
    """Parse an IBKR execution.time string into a date (for cache pruning)."""
    if not raw:
        return None
    datepart = str(raw).split()[0]          # 'YYYYMMDD  HH:MM:SS [tz]' -> 'YYYYMMDD'
    for fmt in ("%Y%m%d", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(datepart, fmt).date()
        except ValueError:
            continue
    return None


def load_trades_cache():
    if os.path.exists(TRADES_CACHE_FILE):
        try:
            with open(TRADES_CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            print("[trades-cache] existing cache unreadable - starting fresh.", file=sys.stderr)
    return {}


def update_trades_cache(cache, executions):
    """Merge this run's live TWS executions into the persistent cache (keyed by
    execId), then prune anything older than TRADES_CACHE_DAYS.

    reqExecutions only returns the CURRENT day's fills, while the Flex statement
    lags 1-2 days behind. Without this cache, trades from the most recent day(s)
    are silently dropped — too new for Flex, too old for the live feed on the
    next run. Persisting each day's executions keeps them until Flex catches up.
    """
    for exec_id, ex in executions.items():
        rec = dict(ex)
        # A commission can arrive blank on one run and populated on a later one;
        # never overwrite a known commission with None.
        old = cache.get(exec_id)
        if old and rec.get("commission") is None and old.get("commission") is not None:
            rec["commission"] = old["commission"]
        cache[exec_id] = rec

    cutoff = dt.date.today() - dt.timedelta(days=TRADES_CACHE_DAYS)
    for exec_id in list(cache.keys()):
        d = _parse_exec_date(cache[exec_id].get("time"))
        if d is not None and d < cutoff:
            del cache[exec_id]
    return cache


def save_trades_cache(cache):
    try:
        with open(TRADES_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2)
    except OSError as e:
        print(f"[trades-cache] could not save: {e}", file=sys.stderr)


# ----------------------------------------------------------------------
# Persistent FULL trade history: accumulate every trade ever seen so the
# report always shows the complete record from DATA_FROM, regardless of how
# short a window the Flex Query returns on any single run.
# ----------------------------------------------------------------------
def _trade_history_key(row):
    """Stable string identity for one fill, shared by Flex and live-TWS rows.

    Uses the same (date, symbol, side, qty, price) fingerprint as the live/Flex
    de-dup, so a fill captured live today and re-published by Flex tomorrow maps
    to the SAME entry (no duplicate). JSON keys must be strings, so the tuple is
    joined into one."""
    k = _fill_dedup_key(row.get("symbol") or row.get("description"),
                        row.get("buySell"), row.get("quantity"),
                        row.get("tradePrice"), parse_trade_date(row))
    return "|".join(str(p) for p in k)


def load_trade_history():
    if os.path.exists(TRADE_HISTORY_FILE):
        try:
            with open(TRADE_HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            print("[trade-history] existing history unreadable - starting fresh.", file=sys.stderr)
    return {}


def update_trade_history(history, rows):
    """Merge this run's trades (Flex statement + live TWS fills) into the
    persistent full-history store, keyed by fill identity.

    Never pruned. When a fill is seen again (e.g. Flex later republishes a fill
    first captured live), its richer fields are merged in WITHOUT overwriting any
    known value with a blank, and internal computed keys (e.g. _calcRealizedPnl)
    are stripped so realized P&L is always recomputed fresh across the full set.
    """
    for r in rows:
        key   = _trade_history_key(r)
        clean = {k: v for k, v in r.items() if not k.startswith("_")}
        old   = history.get(key)
        if old:
            merged = dict(old)
            for k, v in clean.items():
                if v not in (None, ""):     # keep richer Flex fields, never blank out
                    merged[k] = v
            history[key] = merged
        else:
            history[key] = clean
    return history


def save_trade_history(history):
    try:
        with open(TRADE_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, indent=2)
    except OSError as e:
        print(f"[trade-history] could not save: {e}", file=sys.stderr)


def build_order_lookup(ledger):
    """(symbol, action, normType) -> list of (last_seen_date, limit, trigger)."""
    lookup = defaultdict(list)
    for rec in ledger.values():
        key = (rec.get("symbol"), (rec.get("action") or "").upper(),
               _norm_type(rec.get("orderType")))
        try:
            seen = dt.date.fromisoformat(rec.get("last_seen", ""))
        except (ValueError, TypeError):
            seen = dt.date.min
        lookup[key].append((seen, rec.get("limit"), rec.get("trigger")))
    return lookup


def lookup_order_prices(order_lookup, symbol, action, order_type, trade_date):
    """Return (limit, trigger) for the ledger order best matching a trade."""
    if not order_lookup:
        return None, None
    key = (symbol, (action or "").upper(), _norm_type(order_type))
    cands = order_lookup.get(key, [])
    if not cands:
        return None, None
    td = trade_date or dt.date.min
    prior = [c for c in cands if c[0] <= td]           # orders seen on/before the trade
    pick = (max(prior, key=lambda c: c[0]) if prior
            else min(cands, key=lambda c: abs((c[0] - td).days)))
    return pick[1], pick[2]


# ----------------------------------------------------------------------
# Pending order helpers
# ----------------------------------------------------------------------
def _fmt_price(v):
    return f"{v:.2f}" if v is not None else ""


def _sl_pct(entry_trigger, sl_trigger):
    if entry_trigger and sl_trigger:
        return f"{abs(entry_trigger - sl_trigger) / entry_trigger * 100:.2f}%"
    return ""


def _offset(trigger, limit_price):
    if trigger is not None and limit_price is not None:
        return f"{abs(limit_price - trigger):.2f}"
    return ""


def _total_amount(qty, trigger, limit_price):
    price = trigger if trigger is not None else limit_price
    if price is not None:
        return f"{qty * price:,.2f}"
    return ""


def build_pending_rows(orders_by_symbol):
    rows = []
    sr = 0
    for symbol, legs in orders_by_symbol.items():
        sr += 1
        parents  = [o for o in legs if not o["parentId"]]
        children = [o for o in legs if o["parentId"]]
        entry = parents[0] if parents else legs[0]
        stop  = children[0] if children else (legs[1] if len(legs) > 1 else None)

        # Entry row — show Total Amount.  Name is blank: pending orders come
        # live from TWS, which doesn't carry the Flex 'description' full name.
        rows.append([
            sr, symbol, "", entry["action"], entry["orderType"],
            _fmt_price(entry["trigger"]), _fmt_price(entry["limit"]),
            _offset(entry["trigger"], entry["limit"]),
            _sl_pct(entry["trigger"], stop["trigger"] if stop else None),
            entry["quantity"],
            _total_amount(entry["quantity"], entry["trigger"], entry["limit"]),
        ])
        # Stop/SELL row — no Total Amount.  Repeat the Contract from the entry
        # row above so it isn't left blank.
        if stop:
            rows.append([
                "", symbol, "", stop["action"], stop["orderType"],
                _fmt_price(stop["trigger"]), _fmt_price(stop["limit"]),
                _offset(stop["trigger"], stop["limit"]),
                "", stop["quantity"],
                "",
            ])
    return rows


PENDING_HEADERS = ["Sr No", "Contract", "Name", "Action", "Type",
                   "Trigger", "Limit", "Offset", "SL-Percentage", "Quantity", "Total Amount"]
PENDING_WIDTHS  = [7, 12, 30, 9, 10, 10, 10, 10, 15, 10, 14]

TODAY_HEADERS = ["Date & Time (UTC)", "Date & Time (GST)", "Sr No", "Contract", "Name",
                 "Action", "Quantity", "Price", "Trigger Price", "Limit Price",
                 "Exchange", "Order Type", "Commission", "Realized PnL", "Account"]
TODAY_WIDTHS  = [22, 22, 7, 14, 30, 10, 12, 14, 14, 14, 12, 12, 14, 14, 14]

# Gulf Standard Time (Dubai) is UTC+4 — used for the second All Trades time column.
_GST_OFFSET_HOURS = 4

# Columns whose numeric cells get green (positive) / red (negative) shading.
# Commission is intentionally excluded — it is shown with no colour.
COLOR_COLUMNS_TRADES = ["Realized PnL"]


def _parse_exec_time(raw):
    """Parse IBKR execution.time string into a display-friendly datetime."""
    raw        = str(raw).strip()
    normalised = " ".join(raw.split())
    for fmt in ("%Y%m%d %H:%M:%S", "%Y%m%d"):
        try:
            return dt.datetime.strptime(normalised, fmt).strftime("%d-%b-%Y %H:%M:%S")
        except ValueError:
            continue
    return raw


def _to_gst(utc_str):
    """Convert a formatted UTC datetime string to Gulf Standard Time (Dubai, UTC+4).
    Returns the shifted 'DD-Mon-YYYY HH:MM:SS' string, or the original unchanged
    when there is no parseable time component to shift."""
    if not utc_str:
        return ""
    try:
        d = dt.datetime.strptime(str(utc_str), "%d-%b-%Y %H:%M:%S")
    except ValueError:
        return utc_str   # date-only or unparseable — leave as-is
    return (d + dt.timedelta(hours=_GST_OFFSET_HOURS)).strftime("%d-%b-%Y %H:%M:%S")


def _trade_price_key(dt_str, symbol, action, qty_str, price_str):
    """Stable identity for a single fill, used to carry forward manual prices.
    Built from the exact formatted strings written to the All Trades sheet."""
    return (str(dt_str), str(symbol), str(action), str(qty_str), str(price_str))


def build_individual_trade_rows(trade_rows, date_filter=None, order_lookup=None,
                                manual_prices=None):
    """
    Build per-trade rows from Flex Query data.

    date_filter   : a collection of dates to include (e.g. {today, yesterday}).
                    If None, ALL trades are included (full history).
    order_lookup  : optional ledger lookup (see build_order_lookup) used to fill
                    the Trigger / Limit columns from previously-captured orders.
    manual_prices : optional {trade_key: (trigger, limit)} carried forward from
                    the previous report, where the user typed prices by hand.
                    These take precedence over the ledger / Flex values.
    """
    if date_filter is not None:
        sel = [r for r in trade_rows if parse_trade_date(r) in date_filter]
    else:
        sel = list(trade_rows)

    # Sort newest-first (descending) by parsed date, then raw timestamp.
    sel.sort(key=lambda r: (parse_trade_date(r) or dt.date.min,
                            r.get("dateTime") or r.get("tradeDate") or ""),
             reverse=True)

    def _fmt(v, dp=2):
        """Absolute-value format (for qty/price/commission)."""
        try:
            return f"{abs(float(v)):.{dp}f}" if v not in (None, "") else ""
        except (ValueError, TypeError):
            return str(v) if v else ""

    def _amt(v, signed=False):
        """Numeric amount (so the cell can carry the 1,234.00 number format).
        Returns a rounded float (abs unless signed) or '' when not numeric."""
        try:
            x = round(float(v), 2) if v not in (None, "") else ""
        except (ValueError, TypeError):
            return str(v) if v else ""
        return x if (x == "" or signed) else abs(x)

    rows = []
    for sr, r in enumerate(sel, start=1):
        symbol   = r.get("symbol") or r.get("description") or ""
        raw_side = r.get("buySell", "").upper().strip()
        action   = ("BUY"  if raw_side in ("BUY", "BOT", "B") else
                    "SELL" if raw_side in ("SELL", "SLD", "S") else raw_side)
        account  = (r.get("accountId") or r.get("acctId") or
                    r.get("account") or "")
        commission = r.get("ibCommission") or r.get("commission")
        name       = r.get("description") or r.get("underlyingSymbol") or ""
        # Trigger / limit are order-level fields, absent from Flex trades.
        # Prefer the captured order ledger; fall back to any Flex fields.
        led_lim, led_trg = lookup_order_prices(order_lookup, symbol, action,
                                               r.get("orderType"), parse_trade_date(r))
        trigger = (led_trg if led_trg is not None
                   else r.get("triggerPrice") or r.get("auxPrice") or r.get("stopPrice"))
        limit   = (led_lim if led_lim is not None
                   else r.get("limitPrice") or r.get("lmtPrice") or r.get("orderLimitPrice"))

        dt_str    = parse_trade_datetime(r)
        qty_str   = _fmt(r.get("quantity"),   0)
        price_str = _fmt(r.get("tradePrice"), 4)
        trigger_cell = _fmt(trigger, 4)
        limit_cell   = _fmt(limit,   4)

        # Carry forward any trigger/limit the user typed into the previous
        # report for this same fill — manual edits win over ledger/Flex values.
        if manual_prices:
            m_trg, m_lim = manual_prices.get(
                _trade_price_key(dt_str, symbol, action, qty_str, price_str),
                (None, None))
            if m_trg not in (None, ""):
                trigger_cell = m_trg
            if m_lim not in (None, ""):
                limit_cell = m_lim

        rows.append([
            dt_str,             # Date & Time (UTC)
            _to_gst(dt_str),    # Date & Time (GST) — Dubai, UTC+4
            sr,
            symbol,
            name,
            action,
            qty_str,
            price_str,
            trigger_cell,
            limit_cell,
            r.get("exchange", ""),
            r.get("orderType", ""),
            _amt(commission),
            _amt(_row_realized_pnl(r), signed=True),
            account,
        ])
    return rows


def build_today_rows_from_flex(trade_rows):
    """Today Orders fallback: per-trade rows for today and yesterday only."""
    today = dt.date.today()
    return build_individual_trade_rows(trade_rows, {today, today - dt.timedelta(days=1)})


def _list_previous_reports():
    """All prior report files in the reports folder, newest first (by the date
    encoded in the filename, then modified time), excluding today's OUTPUT_FILE.
    Accepts the current "MIS_paper_" prefix and the legacy "MIS_"/"Reports_"
    prefixes."""
    try:
        candidates = []
        for fn in os.listdir(_REPORTS_DIR):
            if not fn.lower().endswith(".xlsx") or fn.startswith("~$"):
                continue
            path = os.path.join(_REPORTS_DIR, fn)
            if os.path.abspath(path) == os.path.abspath(OUTPUT_FILE):
                continue
            d = dt.date.min
            for prefix in ("MIS_paper_", "MIS_", "Reports_"):
                try:
                    d = dt.datetime.strptime(fn, prefix + "%d%b%Y.xlsx").date()
                    break
                except ValueError:
                    continue
            candidates.append((d, os.path.getmtime(path), path))
    except OSError:
        return []
    candidates.sort(key=lambda c: (c[0], c[1]), reverse=True)
    return [c[2] for c in candidates]


def _read_alltrades_prices(wb):
    """Build {trade_key: (trigger, limit)} from a workbook's 'All Trades' sheet.

    Columns are resolved by HEADER NAME (row 2), not fixed positions, so prices
    carry forward regardless of the file's column layout — older 14-column
    reports (single "Date & Time") and the current 15-column UTC/GST layout both
    read correctly, and future column changes won't silently break this."""
    prices = {}
    # Accept the current "paper_All Trades" tab and the legacy "All Trades" one
    # so prices still carry forward from reports written before the tab rename.
    sheet = next((nm for nm in ("paper_All Trades", "All Trades")
                  if nm in wb.sheetnames), None)
    if sheet is None:
        return prices
    rows = list(wb[sheet].iter_rows(min_row=2, values_only=True))
    if not rows:
        return prices
    header = [str(c) if c is not None else "" for c in rows[0]]

    def _col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    c_dt  = _col("Date & Time (UTC)", "Date & Time")
    c_con = _col("Contract")
    c_act = _col("Action")
    c_qty = _col("Quantity")
    c_prc = _col("Price")
    c_trg = _col("Trigger Price")
    c_lim = _col("Limit Price")
    if None in (c_dt, c_con, c_act, c_qty, c_prc, c_trg, c_lim):
        return prices

    for row in rows[1:]:
        if not row:
            continue
        g = lambda i: row[i] if i < len(row) else None
        trigger, limit = g(c_trg), g(c_lim)
        if trigger in (None, "") and limit in (None, ""):
            continue
        key = _trade_price_key(g(c_dt) or "", g(c_con) or "", g(c_act) or "",
                               g(c_qty) or "", g(c_prc) or "")
        prices[key] = (
            str(trigger) if trigger not in (None, "") else None,
            str(limit)   if limit   not in (None, "") else None,
        )
    return prices


def load_previous_report():
    """Open the closest prior report (e.g. MIS_04Jun2026.xlsx) and pull forward
    the manual Trigger/Limit prices from the 'All Trades' sheet, which would
    otherwise be lost because every run writes a brand-new dated file. Prices are
    keyed by (datetime, contract, action, qty, price) so they re-attach to the
    same fill.

    Returns manual_prices: dict.
    """
    from openpyxl import load_workbook

    reports = _list_previous_reports()
    if not reports:
        return {}
    prev_path = reports[0]

    try:
        wb = load_workbook(prev_path, read_only=True, data_only=True)
    except Exception as e:                                  # noqa: BLE001
        print(f"  [Carry-forward] could not open {prev_path}: {e}", file=sys.stderr)
        return {}

    # Trigger/limit prices from the most recent report.
    manual_prices = _read_alltrades_prices(wb)

    wb.close()

    # Self-heal: if the latest report carries no trigger/limit prices at all
    # (e.g. a column-layout change once broke the chain), fall back to the most
    # recent older report that still has them so the prices aren't lost.
    if not manual_prices:
        for older in reports[1:]:
            try:
                w2 = load_workbook(older, read_only=True, data_only=True)
            except Exception:                              # noqa: BLE001
                continue
            mp = _read_alltrades_prices(w2)
            w2.close()
            if mp:
                manual_prices = mp
                print(f"  [Carry-forward] recovered {len(mp)} trigger/limit "
                      f"price(s) from {os.path.basename(older)}.")
                break
    print(f"  [Carry-forward] {len(manual_prices)} trigger/limit price(s) "
          f"from {os.path.basename(prev_path)}.")
    return manual_prices


def build_trade_rows(executions):
    """Supplementary: live fills received from reqExecutions this session.
    Aligned to TODAY_HEADERS; Order Type / Realized PnL aren't carried by the
    execution feed, so those cells are left blank for live rows."""
    rows = []
    for sr, ex in enumerate(
        sorted(executions.values(), key=lambda e: str(e["time"])), start=1
    ):
        exec_time = _parse_exec_time(ex["time"])
        rows.append([
            exec_time,                                       # Date & Time (UTC)
            _to_gst(exec_time),                              # Date & Time (GST) — Dubai, UTC+4
            sr,
            ex["contract"],
            "",                                              # Name (not in exec feed)
            ex["action"],
            ex["quantity"],
            f"{ex['price']:.4f}"      if ex["price"]      is not None else "",
            "",                                              # Trigger Price (n/a)
            "",                                              # Limit Price (n/a)
            ex["exchange"],
            "",                                              # Order Type (n/a)
            round(abs(ex["commission"]), 2) if ex["commission"] is not None else "",
            "",                                              # Realized PnL (n/a)
            ex["account"],
        ])
    return rows


def _exec_to_flex_row(ex):
    """Map a live TWS execution into a Flex-style trade dict so it flows through
    the same aggregation / realized-PnL pipeline as Flex trades. Commission is
    stored as a negative cost to match the Flex 'ibCommission' convention, and
    fifoPnlRealized is left blank so the FIFO engine computes it."""
    comm = ex.get("commission")
    return {
        "symbol":          ex.get("contract", ""),
        "description":     ex.get("contract", ""),
        "buySell":         ex.get("action", ""),
        "quantity":        ex.get("quantity", 0),
        "tradePrice":      ex.get("price", 0),
        "commission":      (-abs(_flt(comm)) if comm is not None else None),
        "exchange":        ex.get("exchange", ""),
        "accountId":       ex.get("account", ""),
        "dateTime":        " ".join(str(ex.get("time", "")).split()),
        "orderType":       "",
        "fifoPnlRealized": "",
    }


def _fill_dedup_key(symbol, action, qty, price, d):
    """Stable identity for one fill across the Flex and TWS feeds."""
    a = str(action).upper().strip()
    a = "BUY" if a in ("BUY", "BOT", "B") else "SELL" if a in ("SELL", "SLD", "S") else a
    return (d, str(symbol), a, round(_flt(qty), 4), round(_flt(price), 6))


def merge_live_executions(flex_rows, executions):
    """Combine the Flex statement's trades with cached TWS executions.

    IBKR's Flex batch lags — the most recent day(s) of fills usually aren't in
    it yet — so without this the report drops them. `executions` is the persisted
    trades cache (today's fills plus those captured on recent runs). Each cached
    execution the Flex statement hasn't already published (matched on date,
    contract, side, quantity and price) is appended as a Flex-style row so it
    appears on every sheet. Returns the merged list."""
    flex_keys = {
        _fill_dedup_key(r.get("symbol") or r.get("description"), r.get("buySell"),
                        r.get("quantity"), r.get("tradePrice"), parse_trade_date(r))
        for r in flex_rows
    }
    merged = list(flex_rows)
    added  = 0
    for ex in executions.values():
        fr = _exec_to_flex_row(ex)
        if _fill_dedup_key(fr["symbol"], fr["buySell"], fr["quantity"],
                           fr["tradePrice"], parse_trade_date(fr)) in flex_keys:
            continue                       # already in the Flex batch
        merged.append(fr)
        added += 1
    if added:
        print(f"  [Live fills] merged {added} TWS execution(s) not yet in the "
              f"Flex statement.")
    return merged


# ----------------------------------------------------------------------
# STEP 1: request the statement -> reference code
# ----------------------------------------------------------------------
def send_request():
    r = requests.get(SEND_URL, params={"t": TOKEN, "q": QUERY_ID, "v": VERSION}, timeout=30)
    r.raise_for_status()
    root = ET.fromstring(r.content)
    status = root.findtext("Status")
    if status != "Success":
        code = root.findtext("ErrorCode")
        msg  = root.findtext("ErrorMessage")
        sys.exit(f"SendRequest failed [{code}]: {msg}")
    ref = root.findtext("ReferenceCode")
    print(f"Reference code: {ref}")
    return ref


# ----------------------------------------------------------------------
# STEP 2: poll for the statement XML
# ----------------------------------------------------------------------
def get_statement(ref):
    for attempt in range(1, MAX_POLL_ATTEMPTS + 1):
        r = requests.get(GET_URL, params={"t": TOKEN, "q": ref, "v": VERSION}, timeout=60)
        r.raise_for_status()
        text = r.text
        if "<FlexQueryResponse" in text:
            print("Statement retrieved.")
            return text
        try:
            root = ET.fromstring(r.content)
            code = root.findtext("ErrorCode")
            msg  = root.findtext("ErrorMessage")
            if code and code not in ("1019",):
                sys.exit(f"GetStatement failed [{code}]: {msg}")
            print(f"  not ready yet (attempt {attempt}/{MAX_POLL_ATTEMPTS}), waiting...")
        except ET.ParseError:
            pass
        time.sleep(POLL_WAIT_SECONDS)
    sys.exit("Statement not ready after maximum retries.")


# ----------------------------------------------------------------------
# STEP 3: parse trades + account ID
# ----------------------------------------------------------------------
def parse_trades(xml_text):
    root   = ET.fromstring(xml_text)
    trades = root.findall(".//Trade")
    if not trades:
        trades = root.findall(".//Order")
    return [dict(t.attrib) for t in trades]


def parse_account_id(xml_text):
    root = ET.fromstring(xml_text)
    stmt = root.find(".//FlexStatement")
    if stmt is not None:
        return stmt.get("accountId", "N/A")
    return "N/A"


def parse_trade_date(row):
    """Extract a date from common IBKR datetime attributes."""
    raw = row.get("dateTime") or row.get("tradeDate") or row.get("reportDate") or ""
    raw = raw.strip()
    if not raw:
        return None
    raw = raw.replace(",", " ").replace(";", " ")
    datepart = raw.split()[0]
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(datepart, fmt).date()
        except ValueError:
            continue
    return None


def parse_trade_datetime(row):
    """Extract a full datetime string from IBKR attributes, formatted for display.

    IBKR Flex sends dateTime as 'DD/MM/YYYY;HHMMSS' (e.g. '20/05/2026;080617'),
    but other configs use 'YYYYMMDD;HHMMSS' or ISO. We normalise separators
    then parse the date and time parts independently.
    """
    raw = (row.get("dateTime") or row.get("tradeDate") or "").strip()
    if not raw:
        return ""
    norm  = raw.replace(";", " ").replace(",", " ")
    parts = norm.split()
    datepart = parts[0]
    timepart = parts[1] if len(parts) > 1 else ""

    d = None
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            d = dt.datetime.strptime(datepart, fmt)
            break
        except ValueError:
            continue
    if d is None:
        return raw

    if timepart:
        for tfmt in ("%H%M%S", "%H:%M:%S"):
            try:
                t = dt.datetime.strptime(timepart, tfmt).time()
                return dt.datetime.combine(d.date(), t).strftime("%d-%b-%Y %H:%M:%S")
            except ValueError:
                continue
    return d.strftime("%d-%b-%Y")


# ----------------------------------------------------------------------
# STEP 4: aggregate raw trades into per-contract summary rows
# ----------------------------------------------------------------------
HEADERS = [
    "First Trade Date (UTC)", "First Trade Date (GST)",
    "Last Trade Date (UTC)",  "Last Trade Date (GST)",
    "Contract", "Name", "Buys", "Sells", "Net",
    "Avg (bought)", "Avg (sold)",
    "Total (bought)", "Total (sold)",
    "Exchange List",
    "Commission",
    "PnL", "Unrealized PnL",
]

# Trade Summary columns shaded green (positive) / red (negative).
# Commission is intentionally excluded — it is shown with no colour.
COLOR_COLUMNS_SUMMARY = ["PnL", "Unrealized PnL"]

# Money columns formatted as 1,234.00 on the aggregate sheets.
AMOUNT_COLUMNS_SUMMARY = ["Avg (bought)", "Avg (sold)", "Total (bought)",
                          "Total (sold)", "Commission", "PnL", "Unrealized PnL"]


def _flt(val):
    try:
        return float(val) if val not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def _fmt_limit(val):
    """Format a risk-limit value as 1,234.00 when numeric, else show it as-is
    (e.g. the placeholder 'NA')."""
    try:
        return f"{float(val):,.2f}"
    except (ValueError, TypeError):
        return str(val)


def aggregate(trade_rows):
    data = defaultdict(lambda: {
        "buy_qty": 0.0, "sell_qty": 0.0,
        "buy_value": 0.0, "sell_value": 0.0,
        "commission": 0.0, "pnl": 0.0, "unrealized": 0.0,
        "exchanges": set(),
        "dates": [],
        "name": "",
    })

    for r in trade_rows:
        symbol     = r.get("symbol") or r.get("description") or "UNKNOWN"
        name       = r.get("description") or r.get("underlyingSymbol") or ""
        side       = r.get("buySell", "").upper().strip()
        qty        = abs(_flt(r.get("quantity", 0)))
        price      = abs(_flt(r.get("tradePrice", 0)))
        money      = abs(_flt(r.get("tradeMoney", 0))) or qty * price
        # IBKR Flex uses 'ibCommission'; the TWS API uses 'commission'.
        commission = _flt(r.get("ibCommission") or r.get("commission") or 0)
        pnl        = _row_realized_pnl(r)
        unrealized = _flt(r.get("mtmPnl", 0)) or _flt(r.get("unrealizedPnl", 0))
        exchange   = r.get("exchange", "")

        c = data[symbol]
        if side in ("BUY", "B"):
            c["buy_qty"]   += qty
            c["buy_value"] += money
        elif side in ("SELL", "S"):
            c["sell_qty"]   += qty
            c["sell_value"] += money
        c["commission"] += commission
        c["pnl"]        += pnl
        c["unrealized"] += unrealized
        if exchange:
            c["exchanges"].add(exchange)
        if name and not c["name"]:
            c["name"] = name
        d = parse_trade_date(r)
        if d:
            c["dates"].append((d, parse_trade_datetime(r)))

    result = []
    for symbol, c in data.items():
        avg_b     = c["buy_value"]  / c["buy_qty"]  if c["buy_qty"]  else 0.0
        avg_s     = c["sell_value"] / c["sell_qty"] if c["sell_qty"] else 0.0
        datetimes = sorted(c["dates"], key=lambda x: x[0])
        result.append({
            "Contract":             symbol,
            "Name":                 c["name"],
            "Buys":                 round(c["buy_qty"], 4),
            "Sells":                round(c["sell_qty"], 4),
            "Net":                  round(c["buy_qty"] - c["sell_qty"], 4),
            "Avg (bought)":         round(avg_b, 6),
            "Avg (sold)":           round(avg_s, 6),
            "Total (bought)":       round(c["buy_value"], 2),
            "Total (sold)":         round(c["sell_value"], 2),
            "Exchange List":        ", ".join(sorted(c["exchanges"])),
            "Commission":           round(c["commission"], 2),
            "PnL":                  round(c["pnl"], 2),
            "Unrealized PnL":       round(c["unrealized"], 2),
            "First Trade Date (UTC)": datetimes[0][1]  if datetimes else "",
            "First Trade Date (GST)": _to_gst(datetimes[0][1])  if datetimes else "",
            "Last Trade Date (UTC)":  datetimes[-1][1] if datetimes else "",
            "Last Trade Date (GST)":  _to_gst(datetimes[-1][1]) if datetimes else "",
        })
    result.sort(key=lambda x: x["Contract"])
    return result


def _trade_sort_dt(row):
    """A real datetime for chronological (FIFO) ordering of fills.
    Falls back to datetime.min when the row carries no parseable date."""
    d = parse_trade_date(row)
    if d is None:
        return dt.datetime.min
    raw   = (row.get("dateTime") or row.get("tradeDate") or "").strip()
    parts = raw.replace(";", " ").replace(",", " ").split()
    timepart = parts[1] if len(parts) > 1 else ""
    for tfmt in ("%H%M%S", "%H:%M:%S"):
        try:
            t = dt.datetime.strptime(timepart, tfmt).time()
            return dt.datetime.combine(d, t)
        except ValueError:
            continue
    return dt.datetime.combine(d, dt.time.min)


def compute_realized_pnl(trade_rows):
    """Calculate FIFO realized P&L for each fill, annotating the trade dict
    in place under '_calcRealizedPnl'.

    IBKR's Flex 'fifoPnlRealized' field comes back empty on paper accounts, so
    realized P&L is computed here: each closing fill is matched against earlier
    opposite-side open lots, first-in-first-out, and the realized amount
    (close price − open price) × matched qty is assigned to that closing fill.
    Because the value lands on the closing trade, per-trade figures sum cleanly
    up into the per-contract Trade Summary and the Dashboard totals. Figures are
    gross of commissions, which stay in their own column."""
    from collections import deque

    lots_by_symbol = defaultdict(deque)   # symbol -> deque of [qty, price, sign]
    for r in sorted(trade_rows, key=_trade_sort_dt):
        symbol = r.get("symbol") or r.get("description") or "UNKNOWN"
        side   = r.get("buySell", "").upper().strip()
        qty    = abs(_flt(r.get("quantity", 0)))
        price  = abs(_flt(r.get("tradePrice", 0)))
        sign   = (1 if side in ("BUY", "BOT", "B") else
                  -1 if side in ("SELL", "SLD", "S") else 0)
        if qty == 0 or sign == 0:
            r["_calcRealizedPnl"] = 0.0
            continue

        lots      = lots_by_symbol[symbol]
        realized  = 0.0
        remaining = qty
        # Close opposite-side open lots first (FIFO).
        while remaining > 1e-9 and lots and lots[0][2] == -sign:
            lot     = lots[0]
            matched = min(remaining, lot[0])
            if lot[2] == 1:        # closing a long lot with a sell
                realized += (price - lot[1]) * matched
            else:                  # closing a short lot with a buy
                realized += (lot[1] - price) * matched
            lot[0]    -= matched
            remaining -= matched
            if lot[0] <= 1e-9:
                lots.popleft()
        # Any leftover quantity opens a new lot on this side.
        if remaining > 1e-9:
            lots.append([remaining, price, sign])
        r["_calcRealizedPnl"] = round(realized, 2)


def _row_realized_pnl(row):
    """Realized P&L for one fill: the official Flex 'fifoPnlRealized' when the
    feed reports a real (non-zero) figure, otherwise our FIFO-computed
    '_calcRealizedPnl'. Paper accounts send 'fifoPnlRealized=0' for every fill,
    so a zero/blank Flex value is treated as "not provided" and we fall back to
    the calculated value rather than masking it with 0."""
    flex = _flt(row.get("fifoPnlRealized"))
    if flex != 0:
        return flex
    return _flt(row.get("_calcRealizedPnl", 0))


# ----------------------------------------------------------------------
# STEP 5: write Excel
# ----------------------------------------------------------------------

# Style constants — clean corporate palette: navy headers, neutral text
_NAVY        = PatternFill("solid", fgColor="1F3864")   # dark navy (title + header)
_TITLE_FONT  = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
_HDR_FONT    = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
_DATA_FONT   = Font(size=10, color="262626", name="Calibri")   # near-black body text
_LABEL_FONT  = Font(bold=True, size=10, color="1F3864", name="Calibri")
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="center")
_RIGHT       = Alignment(horizontal="right",  vertical="center")

# PnL cell highlights: green for profit, faint red for loss.
_PNL_POS_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green (positive)
_PNL_NEG_FILL = PatternFill("solid", fgColor="FFC7CE")   # faint red (negative)

# Thin grey gridline border applied around every cell in a table.
_GRID_SIDE   = Side(style="thin", color="BFBFBF")
_CELL_BORDER = Border(left=_GRID_SIDE, right=_GRID_SIDE, top=_GRID_SIDE, bottom=_GRID_SIDE)

# Uniform amount/money display across every sheet, e.g. 1,234.00 (and -1,234.00).
_AMOUNT_FMT = "#,##0.00"


def _apply_amount_format(ws, col, start_row=3):
    """Apply the 1,234.00 number format to numeric cells in a column."""
    for r in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = _AMOUNT_FMT

# Green (Excel selection green) border drawn around each sheet's title banner.
_TITLE_SIDE   = Side(style="medium", color="00B050")
_TITLE_BORDER = Border(left=_TITLE_SIDE, right=_TITLE_SIDE, top=_TITLE_SIDE, bottom=_TITLE_SIDE)


def _apply_cell_borders(ws, n_cols, start_row=2):
    """Draw a thin border around every cell from start_row..max_row, cols 1..n_cols."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.border = _CELL_BORDER


def _color_pnl_cell(cell):
    """Shade a cell green if its numeric value is positive, faint red if negative."""
    try:
        v = float(str(cell.value).replace(",", ""))
    except (ValueError, TypeError, AttributeError):
        return
    if v > 0:
        cell.fill = _PNL_POS_FILL
    elif v < 0:
        cell.fill = _PNL_NEG_FILL

# Every worksheet tab name carries this prefix (paper-trading report).
_TAB_PREFIX = "paper_"


def _tab(name):
    """Worksheet tab name: the internal sheet key prefixed with 'paper_'."""
    return f"{_TAB_PREFIX}{name}"


# Sheet display titles
_FROM_STR = DATA_FROM.strftime("%d-%b-%Y")
_SHEET_TITLES = {
    "Dashboard":       "ACCOUNT DASHBOARD",
    "Pending Order":   "PENDING ORDERS",
    "All Trades":      f"ALL TRADES  (From {_FROM_STR})",
    "Open Position":   f"OPEN POSITIONS  (From {_FROM_STR})",
    "Trade Summary":    f"TRADE SUMMARY  (From {_FROM_STR})",
}


def _write_title_row(ws, text, n_cols):
    """Row 1: merged dark-navy title spanning all columns, framed by a green border."""
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A1:{end_col}1")
    cell = ws["A1"]
    cell.value     = text
    cell.font      = _TITLE_FONT
    cell.fill      = _NAVY
    cell.alignment = _CENTER
    ws.row_dimensions[1].height = 24
    # Green frame around the whole merged title banner. For a merged cell the
    # border must be set on every underlying cell so all four outer edges show.
    for col in range(1, n_cols + 1):
        ws.cell(row=1, column=col).border = _TITLE_BORDER


def _write_header_row(ws, headers, row=2):
    """Row 2: per-column headers in dark navy."""
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font      = _HDR_FONT
        cell.fill      = _NAVY
        cell.alignment = _CENTER
    ws.row_dimensions[row].height = 30


def _style_data_rows(ws, start_row, n_cols):
    """Apply data font and alignment to all data rows."""
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=n_cols):
        for cell in row:
            cell.font      = _DATA_FONT
            cell.alignment = _CENTER


def _apply_autofilter(ws, n_cols, header_row=2):
    """Enable Excel column filters on the header row across all columns."""
    last_col = get_column_letter(n_cols)
    last_row = max(ws.max_row, header_row)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


def write_excel(rows, pending_rows, trade_rows, account_id, account_data, order_lookup=None):
    today     = dt.date.today()
    yesterday = today - dt.timedelta(days=1)
    cutoff_7  = today - dt.timedelta(days=7)

    # ── Global cutoff: keep only trades on/after DATA_FROM (25-May-2026) ──
    rows = [r for r in rows if (parse_trade_date(r) or dt.date.min) >= DATA_FROM]
    print(f"  [Filter] {len(rows)} trade(s) on/after {DATA_FROM:%d-%b-%Y}.")

    # Calculate FIFO realized P&L per fill (annotates each row in place) so the
    # All Trades, Trade Summary and Dashboard all reflect a computed P&L even
    # when IBKR's Flex feed omits it (e.g. on paper accounts).
    compute_realized_pnl(rows)

    # Today's set mirrors the Today Orders sheet: today's fills, or yesterday's
    # as a fallback when IBKR's Flex batch hasn't published today's yet.
    rows_today, rows_7 = [], []
    for row in rows:
        d = parse_trade_date(row)
        if d is None:
            continue
        if d in (today, yesterday):
            rows_today.append(row)
        if d >= cutoff_7:
            rows_7.append(row)

    agg_today = aggregate(rows_today)
    agg_7     = aggregate(rows_7)
    agg_all   = aggregate(rows)             # all trades since DATA_FROM

    # Carry forward the manual trigger/limit prices (All Trades) from the
    # previous report.
    manual_prices = load_previous_report()

    # All Trades: every individual trade since DATA_FROM (ledger trigger/limit,
    # with manually-entered prices from the previous report taking precedence).
    all_trade_rows = build_individual_trade_rows(rows, order_lookup=order_lookup,
                                                 manual_prices=manual_prices)
    print(f"  [All Trades] {len(all_trade_rows)} individual trade(s) since {DATA_FROM:%d-%b-%Y}.")

    # Sheet sequence: Index, Dashboard, Pending Order, Open Position, All Trades,
    # Trade Summary.
    wb = Workbook()
    _fill_index_sheet(wb.active,
                      ["Dashboard", "Pending Order", "Open Position",
                       "All Trades", "Trade Summary"],
                      account_id)
    _fill_dashboard_sheet(wb.create_sheet(), account_id, account_data, pending_rows,
                          agg_7, agg_today, agg_all, len(all_trade_rows))
    _fill_pending_sheet(wb.create_sheet(),            pending_rows)
    _fill_running_positions_sheet(wb.create_sheet(),  rows)
    _fill_trade_list_sheet(wb.create_sheet(), "All Trades", all_trade_rows)
    _fill_sheet(wb.create_sheet(), "Trade Summary",   agg_all)
    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}: {len(pending_rows)} pending, "
          f"{len(all_trade_rows)} all trades, "
          f"{len(agg_today)} contracts (today), "
          f"{len(agg_7)} contracts (7d), {len(agg_all)} contracts (summary)")


# ── Index sheet (links to every other sheet) ──────────────────────────
def _ordinal(n):
    if 11 <= (n % 100) <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def _fill_index_sheet(ws, sheet_titles, account_id="N/A"):
    ws.title = _tab("Index")

    title_fill = PatternFill("solid", fgColor="DDEBF7")   # light blue
    head_fill  = PatternFill("solid", fgColor="DDD9C4")   # beige
    data_fill  = PatternFill("solid", fgColor="DDEBF7")   # light blue
    thin       = Side(style="thin", color="1F3864")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    link_font  = Font(name="Calibri", size=11, bold=True, color="0563C1", underline="single")
    head_font  = Font(name="Calibri", size=11, bold=True, color="1F3864")
    title_font = Font(name="Calibri", size=12, bold=True, color="1F3864")
    label_font = Font(name="Calibri", size=11, bold=True, color="1F3864")
    value_font = Font(name="Calibri", size=11, color="1F3864")

    today = dt.date.today()
    report_title = f"IBKR - MIS Report as on {_ordinal(today.day)} {today:%B %Y}"

    # Row 3: report title (merged B:C)
    ws.merge_cells("B3:C3")
    t = ws["B3"]
    t.value, t.font, t.fill = report_title, title_font, title_fill
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Row 4: "Index" (merged B:C)
    ws.merge_cells("B4:C4")
    ix = ws["B4"]
    ix.value, ix.font, ix.fill = "Index", head_font, title_fill
    ix.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5: column headers
    for col, label in ((2, "SI"), (3, "Description")):
        c = ws.cell(row=5, column=col, value=label)
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Rows 6+: one linked row per sheet
    for i, title in enumerate(sheet_titles, start=1):
        r = 5 + i
        sl = ws.cell(row=r, column=2, value=i)
        sl.fill = data_fill
        sl.alignment = Alignment(horizontal="center", vertical="center")
        sl.font = Font(name="Calibri", size=11, color="1F3864")

        link = ws.cell(row=r, column=3, value=_tab(title))
        link.hyperlink  = f"#'{_tab(title)}'!A1"
        link.font       = link_font
        link.fill       = data_fill
        link.alignment  = Alignment(horizontal="center", vertical="center")

    # Borders around the whole table (rows 3..last, cols B:C)
    last_row = 5 + len(sheet_titles)
    for row in ws.iter_rows(min_row=3, max_row=last_row, min_col=2, max_col=3):
        for cell in row:
            cell.border = border

    # Green frame around the report title banner (B3:C3), matching other sheets.
    ws["B3"].border = _TITLE_BORDER
    ws["C3"].border = _TITLE_BORDER

    # ── Report Details box (cols E:F), alongside the index table ──────
    details = [
        ("Account ID",         account_id),
        ("Report Date",        today.strftime("%d-%b-%Y")),
        ("Data From (cutoff)", DATA_FROM.strftime("%d-%b-%Y")),
    ]
    ws.merge_cells("E4:F4")
    dt_title = ws["E4"]
    dt_title.value, dt_title.font, dt_title.fill = "Report Details", head_font, title_fill
    dt_title.alignment = Alignment(horizontal="center", vertical="center")
    for i, (label, value) in enumerate(details, start=1):
        r = 4 + i
        lc = ws.cell(row=r, column=5, value=label)
        lc.font, lc.fill = label_font, data_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=6, value=value)
        vc.font, vc.fill = value_font, data_fill
        vc.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=4, max_row=4 + len(details), min_col=5, max_col=6):
        for cell in row:
            cell.border = border

    # ── Login Details box (cols E:F), below the Report Details box ─────
    # Report Details fills rows 4..(4+len(details)); leave one blank row, then
    # start the credentials box beneath it.
    cred_title_row = 4 + len(details) + 2          # one-row gap after Report Details
    credentials = [
        ("Username", "gautam575"),
        ("Password", "algoggt1))"),
    ]
    ws.merge_cells(start_row=cred_title_row, start_column=5,
                   end_row=cred_title_row, end_column=6)
    ct = ws.cell(row=cred_title_row, column=5)
    ct.value, ct.font, ct.fill = "Login Details (Paper Account)", head_font, title_fill
    ct.alignment = Alignment(horizontal="center", vertical="center")
    for i, (label, value) in enumerate(credentials, start=1):
        r = cred_title_row + i
        lc = ws.cell(row=r, column=5, value=label)
        lc.font, lc.fill = label_font, data_fill
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = ws.cell(row=r, column=6, value=value)
        vc.font, vc.fill = value_font, data_fill
        vc.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=cred_title_row,
                            max_row=cred_title_row + len(credentials),
                            min_col=5, max_col=6):
        for cell in row:
            cell.border = border

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.sheet_view.showGridLines = False


# ── Dashboard sheet ────────────────────────────────────────────────────
def _fill_dashboard_sheet(ws, account_id, account_data, pending_rows,
                          agg_7, agg_today, agg_all, num_all_trades):
    ws.title  = _tab("Dashboard")
    now_ist   = dt.datetime.now(_IST)
    n_cols    = 2   # Parameter | Value

    _write_title_row(ws, _SHEET_TITLES["Dashboard"], n_cols)
    _write_header_row(ws, ["Parameter", "Value"], row=2)

    # ── Report info ───────────────────────────────────────────────────
    info = [
        ("Account ID",         account_id,                     ""),
        ("Report Date",        now_ist.strftime("%d-%b-%Y"),    ""),
        ("Data From (cutoff)", DATA_FROM.strftime("%d-%b-%Y"),  ""),
    ]

    # ── All-trades summary (everything since DATA_FROM) ───────────────
    at_buys   = sum(r["Total (bought)"]       for r in agg_all)
    at_sells  = sum(r["Total (sold)"]         for r in agg_all)
    at_comm   = sum(r["Commission"]           for r in agg_all)
    at_pnl    = sum(r["PnL"]                  for r in agg_all)
    at_unreal = sum(r["Unrealized PnL"]       for r in agg_all)

    summary_all = [
        ("Total Trades (fills)", num_all_trades,            ""),
        ("Contracts Traded",     len(agg_all),              ""),
        ("Total Bought Value",   f"{at_buys:,.2f}",          ""),
        ("Total Sold Value",     f"{at_sells:,.2f}",         ""),
        ("Commission",           f"{at_comm:,.2f}",          ""),
        ("Realized PnL",         f"{at_pnl:,.2f}",           ""),
        ("Unrealized PnL",       f"{at_unreal:,.2f}",        ""),
        ("Max Exposure",         _fmt_limit(MAX_EXPOSURE),    ""),
        ("Daily Max Loss",       _fmt_limit(DAILY_MAX_LOSS),  ""),
        ("Loss due to bugs",     f"{LOSS_DUE_TO_BUGS:,.2f}", ""),
    ]

    # Dashboard value cells shaded green (positive) / red (negative).
    # Commission is intentionally excluded — it is shown with no colour.
    color_labels = {"Loss due to bugs", "Unrealized PnL", "Realized PnL"}

    # ── Assemble all sections with dividers ───────────────────────────
    sections = [
        ("REPORT INFO",                          info),
        (f"ALL TRADES SUMMARY (From {DATA_FROM:%d-%b-%Y})", summary_all),
    ]

    r_idx = 3
    for section_title, section_rows in sections:
        # Blank gap before each section (except first)
        if r_idx > 3:
            ws.row_dimensions[r_idx].height = 6
            r_idx += 1

        # Section sub-heading row
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=r_idx, column=col,
                           value=section_title if col == 1 else "")
            cell.font      = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
            cell.fill      = _NAVY
            cell.alignment = _LEFT
        ws.row_dimensions[r_idx].height = 18
        r_idx += 1

        # Data rows (Currency column dropped — values are USD as shown in TWS)
        for label, value, _currency in section_rows:
            ca = ws.cell(row=r_idx, column=1, value=label)
            cb = ws.cell(row=r_idx, column=2, value=value)
            ca.font = _LABEL_FONT
            cb.font = _DATA_FONT
            cb.alignment = _RIGHT
            if label in color_labels:
                _color_pnl_cell(cb)
            ws.row_dimensions[r_idx].height = 16
            r_idx += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    ws.freeze_panes = "A3"


# ── Per-trade sheets (All Trades) ─────────────────────────────────────
def _fill_trade_list_sheet(ws, sheet_key, rows):
    ws.title  = _tab(sheet_key)
    n_cols    = len(TODAY_HEADERS)

    _write_title_row(ws, _SHEET_TITLES[sheet_key], n_cols)
    _write_header_row(ws, TODAY_HEADERS, row=2)

    for row in rows:
        ws.append(row)

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    # Amount columns: uniform 1,234.00 number format.
    for h in ("Commission", "Realized PnL"):
        _apply_amount_format(ws, TODAY_HEADERS.index(h) + 1)

    # Shade the Commission / Realized PnL cells green (positive) / red (negative).
    for h in COLOR_COLUMNS_TRADES:
        col = TODAY_HEADERS.index(h) + 1
        for r in range(3, ws.max_row + 1):
            _color_pnl_cell(ws.cell(row=r, column=col))

    for col, width in enumerate(TODAY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _apply_cell_borders(ws, n_cols, start_row=2)
    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


# ── Pending Order sheet ────────────────────────────────────────────────
def _fill_pending_sheet(ws, rows):
    ws.title  = _tab("Pending Order")
    n_cols    = len(PENDING_HEADERS)

    _write_title_row(ws, _SHEET_TITLES["Pending Order"], n_cols)
    _write_header_row(ws, PENDING_HEADERS, row=2)

    for row in rows:
        ws.append(row)

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    for col, width in enumerate(PENDING_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


# ── Open Position sheet ───────────────────────────────────────────────
def _fill_running_positions_sheet(ws, trade_rows):
    filtered = [r for r in trade_rows
                if (parse_trade_date(r) or dt.date.min) >= RUNNING_POSITIONS_FROM]
    agg     = aggregate(filtered)
    running = [r for r in agg if r["Net"] != 0]
    _fill_sheet(ws, "Open Position", running)


# ── Generic trade aggregation sheet ──────────────────────────────────
def _fill_sheet(ws, title, agg_rows):
    ws.title = _tab(title)
    n_cols   = len(HEADERS)

    _write_title_row(ws, _SHEET_TITLES.get(title, title.upper()), n_cols)
    _write_header_row(ws, HEADERS, row=2)

    for row in agg_rows:
        ws.append([row.get(h, "") for h in HEADERS])

    _style_data_rows(ws, start_row=3, n_cols=n_cols)

    # Amount columns: uniform 1,234.00 number format.
    for h in AMOUNT_COLUMNS_SUMMARY:
        if h in HEADERS:
            _apply_amount_format(ws, HEADERS.index(h) + 1)

    # Shade money columns green (positive) / faint red (negative).
    for h in COLOR_COLUMNS_SUMMARY:
        col = HEADERS.index(h) + 1
        for r in range(3, ws.max_row + 1):
            _color_pnl_cell(ws.cell(row=r, column=col))

    # Auto-fit column widths based on header + data content
    for i, h in enumerate(HEADERS, start=1):
        values = [str(r.get(h, "")) for r in agg_rows]
        width  = max(len(h), *(len(v) for v in values)) if values else len(h)
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 3, 13), 32)

    _apply_autofilter(ws, n_cols)
    ws.freeze_panes = "A3"


def main():
    print("Connecting to TWS - fetching open/completed orders, executions, and account summary...")
    orders_by_symbol, account_data, executions, completed_orders = fetch_tws_data()
    pending_rows = build_pending_rows(orders_by_symbol)
    trade_rows   = build_trade_rows(executions)
    print(f"  {len(pending_rows)} pending order rows | "
          f"{len(trade_rows)} today's executions | "
          f"{len(completed_orders)} completed orders | "
          f"{len(account_data)} account tags collected.")

    # Persist trigger/limit from both open AND completed orders, then build the
    # lookup used to back-fill those prices onto matching trades.
    ledger = load_order_ledger()
    update_order_ledger(ledger, orders_by_symbol, completed_orders)
    save_order_ledger(ledger)
    order_lookup = build_order_lookup(ledger)
    print(f"  [Order Ledger] {len(ledger)} order(s) tracked for trigger/limit history "
          f"({LEDGER_FILE}).")

    # Persist today's live executions so the most recent trading day(s) aren't
    # lost while the Flex statement lags behind. The cache (today's fills plus
    # those captured on recent runs) is what we merge into the Flex trades.
    trades_cache = load_trades_cache()
    update_trades_cache(trades_cache, executions)
    save_trades_cache(trades_cache)
    print(f"  [Trades Cache] {len(trades_cache)} execution(s) retained to bridge the "
          f"Flex lag ({TRADES_CACHE_FILE}).")

    ref        = send_request()
    xml_text   = get_statement(ref)
    rows       = parse_trades(xml_text)
    account_id = parse_account_id(xml_text)
    print(f"Account: {account_id} | Parsed {len(rows)} trade records from Flex.")

    # The Flex statement lags — the most recent day(s) of fills usually aren't in
    # it yet. Merge in the cached TWS executions (today's plus any captured on
    # recent runs) so no trading day is lost while Flex catches up.
    rows = merge_live_executions(rows, trades_cache)

    # Accumulate every trade into the persistent full-history store, then build
    # the report from the COMPLETE history. This way the sheets always show the
    # full record from DATA_FROM even when the Flex Query returns only a short
    # rolling window, and a single empty/failed run can never shrink them.
    history = load_trade_history()
    update_trade_history(history, rows)
    save_trade_history(history)
    rows = list(history.values())
    print(f"  [Trade History] {len(rows)} trade(s) in the full-history store "
          f"({TRADE_HISTORY_FILE}).")

    if not rows:
        print("No trades found. Check the query's date range and section config.")
    write_excel(rows, pending_rows, trade_rows, account_id, account_data, order_lookup)


if __name__ == "__main__":
    main()
